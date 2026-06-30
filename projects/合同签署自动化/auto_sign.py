#!/usr/bin/env python3
"""
京东电子签约平台 - 批量合同填报自动化脚本 v2
============================================
使用方式：
  python auto_sign.py              # 正常模式（27家全跑）
  python auto_sign.py --test       # 测试模式（只跑第1家，每步暂停）
  python auto_sign.py --start 5    # 从第5家开始

环境要求:
  pip install playwright openpyxl
  playwright install chromium
"""

import os
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ==================== 配置 ====================

EXCEL_FILE = "合同字段提取表_27家供应商_已核验.xlsx"
SHEET_NAME = "合同字段提取表"
PLATFORM_URL = "https://xingkong.jd.com/qudao/793559463196729346"

SCREENSHOT_DIR = Path("./screenshots")
SCREENSHOT_DIR.mkdir(exist_ok=True)
LOG_FILE = Path("./sign_log.txt")

# 运行模式
TEST_MODE = "--test" in sys.argv
START_INDEX = 1
for i, arg in enumerate(sys.argv):
    if arg == "--start" and i + 1 < len(sys.argv):
        START_INDEX = int(sys.argv[i + 1])

# ==================== 日志 ====================

def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")

def screenshot(page, name):
    path = SCREENSHOT_DIR / f"{datetime.now().strftime('%H%M%S')}_{name}.png"
    page.screenshot(path=str(path), full_page=False)
    log(f"  截图: {path.name}")
    return path

def debug_pause(msg=""):
    if TEST_MODE:
        print(f"\n[DEBUG] {msg}")
        print("按回车继续，输入 'q' 退出...")
        resp = input().strip()
        if resp == "q":
            sys.exit(0)

# ==================== Excel ====================

def load_excel(filepath):
    """按列名读取Excel，返回list[dict]"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[SHEET_NAME]

    # 读表头
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val).strip() if val else f"col_{col}")

    log(f"表头: {len(headers)} 列")

    records = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        if not name:
            continue
        record = {}
        for col_idx, header in enumerate(headers):
            val = ws.cell(row=row, column=col_idx + 1).value
            record[header] = str(val).strip() if val is not None else ""
        records.append(record)

    log(f"加载 {len(records)} 条记录")
    return records

# ==================== 工具函数 ====================

def try_selectors(page, selectors, timeout=5000):
    """尝试多个选择器，返回第一个匹配的locator"""
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.is_visible(timeout=1000):
                return loc
        except Exception:
            continue
    return None

def find_input_near_label(page, label_text, timeout=5000):
    """通过label文本找到附近的input/textarea"""
    strategies = [
        # 策略1: label + for属性关联
        f"label:has-text('{label_text}') >> .. >> input",
        f"label:has-text('{label_text}') >> .. >> textarea",
        # 策略2: 包含label文本的form-item内的input
        f".ant-form-item:has-text('{label_text}') input",
        f".ant-form-item:has-text('{label_text}') textarea",
        f"[class*='form-item']:has-text('{label_text}') input",
        # 策略3: label文本后面的同级input
        f"text='{label_text}' >> xpath=../..//input",
        f"text='{label_text}' >> xpath=../..//textarea",
        # 策略4: 更宽泛
        f"text='{label_text}' >> xpath=../../..//input",
    ]
    return try_selectors(page, strategies, timeout)

def find_select_near_label(page, label_text, timeout=5000):
    """通过label文本找到附近的下拉框"""
    strategies = [
        f".ant-form-item:has-text('{label_text}') .ant-select",
        f".ant-form-item:has-text('{label_text}') .ant-select-selector",
        f"[class*='form-item']:has-text('{label_text}') [class*='select']",
        f"text='{label_text}' >> xpath=../..//*[contains(@class,'select')]",
        f"text='{label_text}' >> xpath=../../..//*[contains(@class,'select')]",
    ]
    return try_selectors(page, strategies, timeout)

def click_option(page, option_text, timeout=5000):
    """点击下拉选项（在弹出的dropdown中）"""
    strategies = [
        f".ant-select-dropdown .ant-select-item:has-text('{option_text}')",
        f".ant-select-dropdown [title='{option_text}']",
        f".ant-select-item:has-text('{option_text}')",
        f".rc-virtual-list-holder-inner >> text='{option_text}'",
        f"[class*='dropdown'] >> text='{option_text}'",
        f"text='{option_text}'",
    ]
    return try_selectors(page, strategies, timeout)

def safe_fill(page, label_text, value, log_prefix=""):
    """通过label填写input/textarea"""
    if not value:
        return True
    loc = find_input_near_label(page, label_text)
    if loc:
        try:
            loc.click()
            loc.fill("")
            loc.fill(str(value))
            log(f"  {log_prefix}填写 '{label_text}' = {value}")
            return True
        except Exception as e:
            log(f"  {log_prefix}[WARN] 填写失败 '{label_text}': {e}")
    else:
        log(f"  {log_prefix}[WARN] 未找到 '{label_text}' 输入框")
    return False

def safe_select(page, label_text, option_text, log_prefix=""):
    """通过label选择下拉选项"""
    if not option_text:
        return True
    trigger = find_select_near_label(page, label_text)
    if trigger:
        try:
            trigger.click()
            time.sleep(0.8)
            opt = click_option(page, option_text)
            if opt:
                opt.click()
                log(f"  {log_prefix}选择 '{label_text}' -> {option_text}")
                return True
            else:
                log(f"  {log_prefix}[WARN] 未找到选项 '{option_text}'")
                # 按Esc关闭下拉
                page.keyboard.press("Escape")
        except Exception as e:
            log(f"  {log_prefix}[WARN] 选择失败 '{label_text}': {e}")
            page.keyboard.press("Escape")
    else:
        log(f"  {log_prefix}[WARN] 未找到 '{label_text}' 下拉框")
    return False

def safe_date(page, label_text, date_str, log_prefix=""):
    """填写日期"""
    if not date_str:
        return True
    # 标准化日期格式
    date_str = date_str.replace("/", "-")
    loc = find_input_near_label(page, label_text)
    if loc:
        try:
            loc.click()
            time.sleep(0.3)
            # 清除已有内容
            page.keyboard.press("Control+a")
            page.keyboard.press("Backspace")
            time.sleep(0.2)
            page.keyboard.type(date_str, delay=50)
            time.sleep(0.3)
            page.keyboard.press("Enter")
            time.sleep(0.3)
            log(f"  {log_prefix}日期 '{label_text}' = {date_str}")
            return True
        except Exception as e:
            log(f"  {log_prefix}[WARN] 日期失败 '{label_text}': {e}")
    else:
        log(f"  {log_prefix}[WARN] 未找到 '{label_text}' 日期框")
    return False

def safe_radio(page, label_text, option_text, log_prefix=""):
    """点击单选按钮"""
    if not option_text:
        return True
    strategies = [
        f".ant-radio-group:has-text('{label_text}') >> text='{option_text}'",
        f"text='{label_text}' >> xpath=../..//label:has-text('{option_text}')",
        f"text='{option_text}'",
    ]
    loc = try_selectors(page, strategies)
    if loc:
        try:
            loc.click()
            log(f"  {log_prefix}单选 '{label_text}' -> {option_text}")
            return True
        except Exception as e:
            log(f"  {log_prefix}[WARN] 单选失败: {e}")
    return False

def click_button(page, text, timeout=10000):
    """点击按钮"""
    strategies = [
        f"button:has-text('{text}')",
        f".ant-btn:has-text('{text}')",
        f"[class*='btn']:has-text('{text}')",
        f"a:has-text('{text}')",
    ]
    loc = try_selectors(page, strategies, timeout)
    if loc:
        try:
            loc.click()
            log(f"  点击按钮: '{text}'")
            return True
        except Exception as e:
            log(f"  [WARN] 点击按钮失败 '{text}': {e}")
    else:
        log(f"  [WARN] 未找到按钮 '{text}'")
    return False

# ==================== 核心流程 ====================

def fill_single_contract(page, record):
    """填写单个合同，返回 (success: bool, errors: list)"""
    errors = []
    idx = record.get("序号", "?")
    name = record.get("供应商名称", "未知")
    prefix = f"[{idx}-{name}] "

    log(f"\n{'='*60}")
    log(f"开始填写: {idx}. {name}")
    log(f"{'='*60}")

    # ----- 1. 点击"新建e签" -----
    log(f"{prefix}步骤1: 新建e签")
    if not click_button(page, "新建e签"):
        # 尝试其他文本
        if not click_button(page, "新建一签"):
            errors.append("找不到'新建e签'按钮")
            return False, errors
    time.sleep(3)
    screenshot(page, f"{idx}_step1_new")

    # ----- 2. 合作意向 -----
    log(f"{prefix}步骤2: 合作意向")
    coop_intent = record.get("合作意向", "")
    if coop_intent:
        # 找到合作意向的输入框/搜索框
        coop_loc = find_input_near_label(page, "合作意向")
        if not coop_loc:
            # 尝试更宽泛的搜索
            coop_loc = find_select_near_label(page, "合作意向")
        if coop_loc:
            try:
                coop_loc.click()
                time.sleep(0.5)
                coop_loc.fill("")
                # 输入供应商名称进行搜索
                coop_loc.fill(name)
                time.sleep(1)
                page.keyboard.press("Enter")
                time.sleep(2)
                # 从下拉中选择匹配的合作意向
                opt = click_option(page, coop_intent)
                if opt:
                    opt.click()
                    log(f"  {prefix}合作意向选择: {coop_intent}")
                else:
                    # 如果精确匹配不到，尝试模糊匹配
                    opt = click_option(page, "存量续签")
                    if opt:
                        opt.click()
                        log(f"  {prefix}合作意向选择: 模糊匹配'存量续签'")
                    else:
                        errors.append(f"合作意向下拉未找到匹配项")
                        log(f"  {prefix}[WARN] 合作意向下拉无匹配")
            except Exception as e:
                errors.append(f"合作意向填写异常: {e}")
                log(f"  {prefix}[WARN] 合作意向异常: {e}")
        else:
            errors.append("未找到合作意向输入框")
            log(f"  {prefix}[WARN] 未找到合作意向输入框")
    else:
        log(f"  {prefix}[WARN] Excel中无合作意向值，跳过")
    time.sleep(1)

    if TEST_MODE:
        debug_pause("合作意向填写完成，检查页面")

    # ----- 3. 合作模式 -----
    log(f"{prefix}步骤3: 合作模式")
    coop_mode = record.get("合作模式", "渠道代理")
    safe_select(page, "合作模式", coop_mode, prefix)
    time.sleep(1)

    # ----- 4. 合同模板 -----
    log(f"{prefix}步骤4: 合同模板")
    template = record.get("合同模板", "")
    if template:
        safe_select(page, "合同模板", template, prefix)
    time.sleep(2)
    screenshot(page, f"{idx}_step4_template")

    if TEST_MODE:
        debug_pause("合同模板选择完成，检查页面")

    # ----- 5. 合同信息 -----
    log(f"{prefix}步骤5: 合同信息")

    # 合同类型
    safe_select(page, "合同类型", record.get("合同类型", ""), prefix)

    # 资金流向
    safe_select(page, "资金流向", record.get("资金流向", "支出"), prefix)

    # 合同金额
    amount = record.get("合同金额(元)", "0")
    safe_fill(page, "合同金额", amount, prefix)

    # 合同期限
    safe_select(page, "合同期限", record.get("合同期限", "固定期限"), prefix)

    # 合同开始/结束时间
    safe_date(page, "合同开始时间", record.get("合同开始时间", ""), prefix)
    safe_date(page, "合同结束时间", record.get("合同结束时间", ""), prefix)

    # 是否E签
    safe_radio(page, "是否E签", record.get("是否E签", "是"), prefix)

    # 是否宿迁承接对账
    safe_radio(page, "是否宿迁承接对账", "是", prefix)

    # 合同标的
    safe_fill(page, "合同标的", record.get("合同标的", "电销服务"), prefix)

    # 合同说明
    safe_fill(page, "合同说明", record.get("合同说明", ""), prefix)

    time.sleep(1)
    screenshot(page, f"{idx}_step5_contract_info")

    if TEST_MODE:
        debug_pause("合同信息填写完成，检查页面")

    # ----- 6. 合同模板变量 -----
    log(f"{prefix}步骤6: 模板变量")
    template_var_keys = [
        "${party_b_name}", "${party_b_address}", "${party_b_phone}",
        "${party_b_contact_email}", "${party_b_contact_person}",
        "${trial_period}",
        "${trial_start_year}", "${trial_start_month}", "${trial_start_day}",
        "${trial_end_year}", "${trial_end_month}", "${trial_end_day}",
        "${renewal_start_year}", "${renewal_start_month}", "${renewal_start_day}",
        "${renewal_end_year}", "${renewal_end_month}", "${renewal_end_day}",
        "${party_b_email}", "${account_name}", "${account_number}",
        "${bank_name}", "${uniform_credit_code}",
    ]
    for var_key in template_var_keys:
        value = record.get(var_key, "")
        if value:
            # 模板变量label通常是 ${xxx} 格式
            safe_fill(page, var_key, value, prefix)
    time.sleep(1)
    screenshot(page, f"{idx}_step6_template_vars")

    if TEST_MODE:
        debug_pause("模板变量填写完成，检查页面")

    # ----- 7. 签约主体 -----
    log(f"{prefix}步骤7: 签约主体")

    # 客方（乙方）
    partner_name = record.get("客方主体", name)
    partner_code = record.get("客方统一信用代码", record.get("统一信用代码", ""))
    partner_is_applicant = record.get("客方是否申请主体", "否")
    partner_sign_id = record.get("客方签约方标识", "乙方")

    log(f"  客方: {partner_name} | {partner_code} | {partner_is_applicant} | {partner_sign_id}")
    # 签约主体通常是表格形式，需要点击"添加"或直接在行内填写
    # 这部分DOM结构差异大，优先尝试通用策略
    safe_fill(page, "客方主体", partner_name, prefix)
    safe_fill(page, "客方统一信用代码", partner_code, prefix)

    # 我方（甲方）
    our_name = record.get("我方主体", "京东科技控股股份有限公司")
    our_code = record.get("我方统一信用代码", "91110302053604529E")
    our_is_applicant = record.get("我方是否申请主体", "是")
    our_sign_id = record.get("我方签约方标识", "甲方")

    log(f"  我方: {our_name} | {our_code} | {our_is_applicant} | {our_sign_id}")
    safe_fill(page, "我方主体", our_name, prefix)
    safe_fill(page, "我方统一信用代码", our_code, prefix)

    time.sleep(1)
    screenshot(page, f"{idx}_step7_parties")

    if TEST_MODE:
        debug_pause("签约主体填写完成，检查页面")

    # ----- 8. 发票信息 -----
    log(f"{prefix}步骤8: 发票信息")
    safe_select(page, "发票类型", record.get("发票类型", "增值税专用发票"), prefix)
    safe_select(page, "开票税率", record.get("开票税率", "6%"), prefix)
    time.sleep(1)
    screenshot(page, f"{idx}_step8_invoice")

    # ----- 9. 保存 -----
    log(f"{prefix}步骤9: 保存")
    click_button(page, "保存")
    time.sleep(3)
    screenshot(page, f"{idx}_step9_saved")

    success = len(errors) == 0
    if success:
        log(f"{prefix}✓ 填写完成")
    else:
        log(f"{prefix}⚠ 填写完成但有 {len(errors)} 个问题:")
        for e in errors:
            log(f"  - {e}")

    return success, errors

# ==================== 主程序 ====================

def main():
    log("=" * 60)
    log("京东电子签约平台 - 批量合同填报")
    log(f"模式: {'测试（仅第1家）' if TEST_MODE else '正常'}")
    log(f"起始序号: {START_INDEX}")
    log("=" * 60)

    # 加载数据
    script_dir = Path(__file__).parent
    excel_path = script_dir / EXCEL_FILE
    if not excel_path.exists():
        log(f"Excel文件不存在: {excel_path}")
        return

    records = load_excel(str(excel_path))
    if not records:
        log("无数据，退出")
        return

    # 过滤起始序号
    if START_INDEX > 1:
        records = [r for r in records if int(r.get("序号", 0)) >= START_INDEX]
        log(f"从第{START_INDEX}家开始，剩余 {len(records)} 条")

    # 测试模式只跑第1条
    if TEST_MODE:
        records = records[:1]
        log("测试模式: 只处理第1家")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        page = context.new_page()

        # 打开平台
        log(f"打开平台: {PLATFORM_URL}")
        page.goto(PLATFORM_URL, wait_until="networkidle", timeout=30000)
        time.sleep(3)
        screenshot(page, "00_platform")

        # 登录确认
        print("\n" + "=" * 60)
        print("请确认浏览器中已登录平台。")
        print("如未登录，请在浏览器窗口中手动登录。")
        print("登录完成后，按回车继续...")
        print("=" * 60)
        input()

        # 批量处理
        success_count = 0
        fail_count = 0
        results = []

        for record in records:
            try:
                ok, errs = fill_single_contract(page, record)
                if ok:
                    success_count += 1
                else:
                    fail_count += 1

                results.append({
                    "name": record.get("供应商名称", ""),
                    "success": ok,
                    "errors": errs,
                })

                if not ok and not TEST_MODE:
                    print(f"\n{record.get('供应商名称', '')} 有问题，是否继续？(y/n)")
                    if input().strip().lower() != "y":
                        break

                # 回到列表页（如果保存后没有自动跳回）
                time.sleep(1)

            except Exception as e:
                log(f"❌ 严重错误: {record.get('供应商名称', '')}: {e}")
                log(traceback.format_exc())
                screenshot(page, f"ERROR_{record.get('序号', '?')}")
                fail_count += 1

                if not TEST_MODE:
                    print(f"\n发生严重错误，是否继续？(y/n)")
                    if input().strip().lower() != "y":
                        break

        # 汇总
        log(f"\n{'='*60}")
        log(f"完成: 成功 {success_count}, 失败 {fail_count}, 总计 {success_count + fail_count}")
        log(f"截图: {SCREENSHOT_DIR.absolute()}")
        log(f"日志: {LOG_FILE.absolute()}")
        log(f"{'='*60}")

        print("\n执行完毕。按回车关闭浏览器...")
        input()
        browser.close()


if __name__ == "__main__":
    main()
