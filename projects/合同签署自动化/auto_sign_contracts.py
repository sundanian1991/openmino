#!/usr/bin/env python3
"""
京东电子签约平台 - 批量合同填报自动化脚本
============================================
使用方式：
  1. 先手动登录 xingkong 平台（保持登录态）
  2. 运行: python auto_sign_contracts.py
  3. 脚本会自动读取同目录下的 Excel 文件，逐条填写并保存
  4. 每填完一家自动截图留档，供您人工复核

环境要求:
  pip install playwright openpyxl
  playwright install chromium
"""

import os
import sys
import time
import json
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ==================== 配置区域 ====================
# TODO: 请根据实际页面确认后修改以下配置

EXCEL_FILE = "合同字段提取表_27家供应商_已核验.xlsx"
SHEET_NAME = "合同字段提取表"

# 平台URL
PLATFORM_URL = "https://xingkong.jd.com/qudao/793559463196729346"

# 合同信息统一值
CONTRACT_TYPE = "渠道主协议"
FUND_DIRECTION = "支出"
CONTRACT_AMOUNT = "0"
CONTRACT_TERM = "固定期限"
IS_ESIGN = "是"
IS_SUQIAN_RECONCILE = "是"           # 是否宿迁承接对账
CONTRACT_SUBJECT = "电销服务"
CONTRACT_DESC = "2026年BPO供应商合作协议换签，本次为供应商换签新协议，无新引入供应商；"
CONTRACT_START = "2026-06-01"
CONTRACT_END = "2027-05-31"

# 发票信息
INVOICE_TYPE = "增值税专用发票"
INVOICE_TAX_RATE = "6%"

# 合同模板（下拉选择）
CONTRACT_TEMPLATE = "客户服务外包合同-202604"

# 签约主体
OUR_COMPANY = "京东科技控股股份有限公司"
OUR_CREDIT_CODE = "91110302053604529E"
OUR_IS_APPLICANT = "是"
OUR_SIGN_ID = "甲方"
PARTNER_IS_APPLICANT = "否"
PARTNER_SIGN_ID = "乙方"

# 测试期字段（无测试期统一填/）
TRIAL_FIELDS = {
    "trial_period": "/",
    "trial_start_year": "/", "trial_start_month": "/", "trial_start_day": "/",
    "trial_end_year": "/", "trial_end_month": "/", "trial_end_day": "/",
}

# 续约期字段
RENEWAL_FIELDS = {
    "renewal_start_year": "2026", "renewal_start_month": "6", "renewal_start_day": "1",
    "renewal_end_year": "2027", "renewal_end_month": "5", "renewal_end_day": "31",
}

# 截图保存目录
SCREENSHOT_DIR = Path("./screenshots")
SCREENSHOT_DIR.mkdir(exist_ok=True)

# 日志文件
LOG_FILE = Path("./sign_log.txt")

# 重试次数
MAX_RETRY = 2

# ==================== 日志工具 ====================

def log(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")

def screenshot(page, name):
    path = SCREENSHOT_DIR / f"{datetime.now().strftime('%H%M%S')}_{name}.png"
    page.screenshot(path=path, full_page=False)
    log(f"  截图已保存: {path}")
    return path

# ==================== Excel读取 ====================

def load_excel_data(filepath):
    """从Excel加载合同数据"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[SHEET_NAME]
    
    records = []
    # 数据从第2行开始，第1行是表头
    for row in range(2, ws.max_row + 1):
        # 如果供应商名称为空则跳过
        name = ws.cell(row=row, column=2).value
        if not name:
            continue
        
        record = {
            "index": ws.cell(row=row, column=1).value,
            "party_b_name": name,
            "party_b_address": ws.cell(row=row, column=3).value or "",
            "party_b_phone": ws.cell(row=row, column=4).value or "",
            "party_b_contact_person": ws.cell(row=row, column=5).value or "",
            "party_b_contact_email": ws.cell(row=row, column=6).value or "",
            "party_b_email": ws.cell(row=row, column=7).value or "",
            "account_name": ws.cell(row=row, column=8).value or "",
            "account_number": ws.cell(row=row, column=9).value or "",
            "bank_name": ws.cell(row=row, column=10).value or "",
            "uniform_credit_code": ws.cell(row=row, column=11).value or "",
        }
        records.append(record)
    
    log(f"从Excel加载了 {len(records)} 条记录")
    return records

# ==================== Playwright操作封装 ====================

class ContractFiller:
    def __init__(self, page):
        self.page = page
        self.errors = []
    
    def safe_click(self, selector, timeout=5000):
        """安全点击，带重试"""
        for attempt in range(MAX_RETRY):
            try:
                self.page.locator(selector).first.click(timeout=timeout)
                return True
            except Exception as e:
                log(f"  点击失败(尝试{attempt+1}/{MAX_RETRY}): {selector} -> {e}")
                time.sleep(1)
        self.errors.append(f"点击失败: {selector}")
        return False
    
    def safe_fill(self, selector, value, timeout=5000):
        """安全填写，先clear再fill"""
        for attempt in range(MAX_RETRY):
            try:
                loc = self.page.locator(selector).first
                loc.fill("", timeout=timeout)  # clear
                loc.fill(str(value), timeout=timeout)
                # 验证
                actual = loc.input_value(timeout=2000)
                if actual == str(value):
                    return True
                log(f"  填写验证失败: 期望'{value}' 实际'{actual}'")
            except Exception as e:
                log(f"  填写失败(尝试{attempt+1}/{MAX_RETRY}): {selector} -> {e}")
                time.sleep(1)
        self.errors.append(f"填写失败: {selector} = {value}")
        return False
    
    def select_dropdown_by_text(self, label_text, option_text, timeout=10000):
        """
        通过label文本找到下拉框，选择指定选项
        策略：点击下拉框 -> 等待选项出现 -> 点击选项
        """
        try:
            # 策略1：通过label关联找到select或模拟下拉
            # 先找到包含label_text的元素，然后在同一父元素内找input或下拉触发器
            label_loc = self.page.locator(f"text='{label_text}'").first
            if label_loc.count() == 0:
                log(f"  未找到label: {label_text}")
                return False
            
            # 点击label附近的下拉框（通常是label的下一个兄弟元素或父元素的某个子元素）
            # 这里使用通用策略：找到包含该label的表单项，然后点击其中的下拉触发器
            parent = label_loc.locator("xpath=../../..")  # 向上找几层到表单项容器
            trigger = parent.locator("input, .ant-select, [class*='select']").first
            trigger.click(timeout=timeout)
            time.sleep(0.5)
            
            # 点击选项
            option = self.page.locator(f"text='{option_text}'").first
            option.click(timeout=timeout)
            log(f"  下拉选择: {label_text} -> {option_text}")
            return True
        except Exception as e:
            log(f"  下拉选择失败: {label_text} -> {option_text}: {e}")
            self.errors.append(f"下拉选择失败: {label_text}")
            return False
    
    def select_date(self, label_text, date_str, timeout=10000):
        """选择日期"""
        try:
            # 找到日期选择器并点击
            label_loc = self.page.locator(f"text='{label_text}'").first
            parent = label_loc.locator("xpath=../../..")
            trigger = parent.locator("input[placeholder*='选择'], .ant-picker").first
            trigger.click(timeout=timeout)
            time.sleep(0.5)
            
            # 输入日期
            self.page.keyboard.insert_text(date_str)
            self.page.keyboard.press("Enter")
            log(f"  日期选择: {label_text} -> {date_str}")
            return True
        except Exception as e:
            log(f"  日期选择失败: {label_text} -> {date_str}: {e}")
            self.errors.append(f"日期选择失败: {label_text}")
            return False
    
    def click_radio(self, label_text, option_text, timeout=5000):
        """点击单选按钮"""
        try:
            # 找到包含option_text的radio
            loc = self.page.locator(f"text='{option_text}'").first
            loc.click(timeout=timeout)
            log(f"  单选: {label_text} -> {option_text}")
            return True
        except Exception as e:
            log(f"  单选失败: {label_text} -> {option_text}: {e}")
            return False
    
    def click_button_by_text(self, text, timeout=10000):
        """通过文本点击按钮"""
        try:
            loc = self.page.locator(f"button:has-text('{text}'), .ant-btn:has-text('{text}'), [class*='btn']:has-text('{text}')").first
            loc.click(timeout=timeout)
            log(f"  点击按钮: {text}")
            return True
        except Exception as e:
            log(f"  点击按钮失败: {text}: {e}")
            return False

# ==================== 主流程 ====================

def fill_contract(page, filler, record):
    """填写单个合同"""
    name = record["party_b_name"]
    idx = record["index"]
    log(f"\n{'='*60}")
    log(f"开始填写第 {idx} 家: {name}")
    log(f"{'='*60}")
    
    errors_before = len(filler.errors)
    
    # ----- 1. 新建e签 -----
    log("步骤1: 点击'新建e签'")
    filler.click_button_by_text("新建e签")
    time.sleep(2)
    screenshot(page, f"{idx:02d}_{name}_step1_new")
    
    # ----- 2. 选择合作意向 -----
    log("步骤2: 选择合作意向")
    # TODO: 需要确认合作意向的名称规则
    # 策略A：如果合作意向名称可从供应商名称推导（如"XX-企金-存量续签"），在此实现
    # 策略B：如果需要在平台搜索，使用select_dropdown_by_text
    # 当前留空，需要用户根据实际DOM补充selector
    log("  [TODO] 合作意向选择逻辑待补充")
    time.sleep(1)
    
    # ----- 3. 选择合作模板 -----
    log("步骤3: 选择合作模板")
    filler.select_dropdown_by_text("合作模板", COOPERATION_TEMPLATE)
    time.sleep(1)
    
    # ----- 4. 选择合同模板 -----
    log("步骤4: 选择合同模板")
    filler.select_dropdown_by_text("合同模板", CONTRACT_TEMPLATE)
    time.sleep(1)
    screenshot(page, f"{idx:02d}_{name}_step4_template")
    
    # ----- 5. 填写合同信息 -----
    log("步骤5: 填写合同信息")
    
    # 合同名称（如已自动生成则跳过）
    # TODO: 确认合同名称是否需要脚本填写
    
    # 合同类型
    filler.select_dropdown_by_text("合同类型", CONTRACT_TYPE)
    
    # 资金流向
    filler.select_dropdown_by_text("资金流向", FUND_DIRECTION)
    
    # 合同金额
    # TODO: 确认金额输入框的selector
    
    # 合同期限
    filler.select_dropdown_by_text("合同期限", CONTRACT_TERM)
    
    # 合同开始/结束时间
    filler.select_date("合同开始时间", CONTRACT_START)
    filler.select_date("合同结束时间", CONTRACT_END)
    
    # 是否E签
    filler.click_radio("是否E签", IS_ESIGN)
    
    # 是否宿迁承接对账
    # TODO: 确认字段名称，勾选"是"
    # filler.click_radio("是否宿迁承接对账", "是")
    
    # 合同标的
    # TODO: 确认输入框selector
    # filler.safe_fill("input[placeholder*='合同标的']", CONTRACT_SUBJECT)
    
    # 合同说明
    # TODO: 确认输入框selector
    # filler.safe_fill("textarea[placeholder*='合同说明']", CONTRACT_DESC)
    
    time.sleep(1)
    screenshot(page, f"{idx:02d}_{name}_step5_contract_info")
    
    # ----- 6. 填写合同模板变量 -----
    log("步骤6: 填写合同模板变量")
    # 这些字段名与Word模板变量一一对应
    template_fields = {
        "party_b_name": record["party_b_name"],
        "party_b_address": record["party_b_address"],
        "party_b_phone": record["party_b_phone"],
        "party_b_contact_email": record["party_b_contact_email"],
        "party_b_contact_person": record["party_b_contact_person"],
        "party_b_email": record["party_b_email"],
        "account_name": record["account_name"],
        "account_number": record["account_number"],
        "bank_name": record["bank_name"],
        "uniform_credit_code": record["uniform_credit_code"],
    }
    
    # 合并日期字段
    template_fields.update(TRIAL_FIELDS)
    template_fields.update(RENEWAL_FIELDS)
    
    for field_name, value in template_fields.items():
        # 策略：通过label文本找到对应输入框
        # TODO: 需要根据实际DOM结构调整selector
        try:
            # 尝试通过包含字段名的label找到输入框
            label = page.locator(f"text='${{{field_name}}}'").first
            if label.count() > 0:
                parent = label.locator("xpath=../..")
                inp = parent.locator("input, textarea").first
                inp.fill(str(value), timeout=3000)
                log(f"  填写 ${{{field_name}}} = {value}")
            else:
                log(f"  [WARN] 未找到字段: ${{{field_name}}}")
        except Exception as e:
            log(f"  [WARN] 填写失败 ${{{field_name}}}: {e}")
    
    time.sleep(1)
    screenshot(page, f"{idx:02d}_{name}_step6_template_vars")
    
    # ----- 7. 签约主体 -----
    log("步骤7: 填写签约主体")
    
    # 客方（乙方）
    # TODO: 确认DOM结构，通常是在表格行中填写
    log("  [TODO] 客方主体填写逻辑待补充")
    
    # 我方（甲方）
    log("  [TODO] 我方主体填写逻辑待补充")
    
    time.sleep(1)
    screenshot(page, f"{idx:02d}_{name}_step7_parties")
    
    # ----- 8. 发票信息 -----
    log("步骤8: 填写发票信息")
    filler.select_dropdown_by_text("发票类型", INVOICE_TYPE)
    filler.select_dropdown_by_text("开票税率", INVOICE_TAX_RATE)
    time.sleep(1)
    screenshot(page, f"{idx:02d}_{name}_step8_invoice")
    
    # ----- 9. 保存（不提交）-----
    log("步骤9: 点击保存")
    filler.click_button_by_text("保存")
    time.sleep(3)
    screenshot(page, f"{idx:02d}_{name}_step9_saved")
    
    # 检查结果
    errors_after = len(filler.errors)
    if errors_after > errors_before:
        log(f"⚠ 第 {idx} 家 ({name}) 填写过程中出现 {errors_after - errors_before} 个错误")
        return False
    else:
        log(f"✓ 第 {idx} 家 ({name}) 填写完成")
        return True


def main():
    # 加载数据
    records = load_excel_data(EXCEL_FILE)
    if not records:
        log("未加载到任何记录，退出")
        return
    
    log(f"准备批量填写 {len(records)} 份合同")
    log("=" * 60)
    
    with sync_playwright() as p:
        # 启动浏览器（headful模式，复用用户登录态）
        # TODO: 如需复用Chrome用户数据目录，请修改以下launch参数
        browser = p.chromium.launch(
            headless=False,  # 有界面模式，便于观察
            # args=["--user-data-dir=/path/to/your/chrome/profile"]  # 如需复用Chrome登录态
        )
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
        )
        page = context.new_page()
        
        # 访问平台
        log(f"打开平台: {PLATFORM_URL}")
        page.goto(PLATFORM_URL)
        time.sleep(3)
        screenshot(page, "00_login_page")
        
        # 如果未登录，提示用户手动登录
        # TODO: 需要根据实际情况判断是否已登录
        print("\n" + "=" * 60)
        print("请确认是否已登录平台。如未登录，请在本窗口手动登录。")
        print("登录完成后，按回车键继续...")
        print("=" * 60)
        input()
        
        # 主循环
        success_count = 0
        fail_count = 0
        filler = ContractFiller(page)
        
        for record in records:
            try:
                ok = fill_contract(page, filler, record)
                if ok:
                    success_count += 1
                else:
                    fail_count += 1
                    # 出错时暂停，让用户决定是否继续
                    print(f"\n⚠ {record['party_b_name']} 填写出错，是否继续？(y/n)")
                    cont = input().strip().lower()
                    if cont != 'y':
                        log("用户中断执行")
                        break
            except Exception as e:
                log(f"❌ 严重错误: {record['party_b_name']}: {e}")
                fail_count += 1
                screenshot(page, f"ERROR_{record['party_b_name']}")
                print(f"\n发生严重错误，是否继续？(y/n)")
                cont = input().strip().lower()
                if cont != 'y':
                    break
        
        # 汇总
        log(f"\n{'='*60}")
        log(f"批量填写完成: 成功 {success_count} 家, 失败 {fail_count} 家, 总计 {len(records)} 家")
        log(f"截图保存在: {SCREENSHOT_DIR.absolute()}")
        log(f"日志保存在: {LOG_FILE.absolute()}")
        log(f"{'='*60}")
        
        # 保持浏览器打开，便于用户查看
        print("\n执行完毕。浏览器保持打开，按回车键关闭...")
        input()
        browser.close()


if __name__ == "__main__":
    main()
