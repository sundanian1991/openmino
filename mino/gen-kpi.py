from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "KPI考核表"

# 标题
ws['A1'] = 'BPO供应商KPI月度考核表'
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A1:H1')

# 基本信息
ws['A3'] = '供应商名称：'
ws['C3'] = '考核月份：'
ws['E3'] = '考核人：'
ws['G3'] = '日期：'

# 表头
headers = ['考核维度', '考核指标', '目标值', '权重', '实际值', '得分', '数据来源', '备注']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col, value=header)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 数据行
kpi_data = [
    ['服务响应时效', '工作时间内首响时间', '≤5分钟', 20, '', '=IF(E7="","",IF(E7<=5,E7/5*D7,D7-(E7-5)*2))', '企微后台数据', '超时每分钟扣2分'],
    ['', '非工作时间首响时间', '≤30分钟', 10, '', '=IF(E8="","",IF(E8<=30,E8/30*D8,D8-(E8-30)*0.5))', '企微后台数据', '超时每分钟扣0.5分'],
    ['服务质量指标', '客户满意度', '≥90%', 25, '', '=IF(E9="","",IF(E9>=90,D9,E9/90*D9))', '满意度调研', '低于90%按比例扣分'],
    ['', '投诉率', '≤1%', 20, '', '=IF(E10="","",IF(E10<=1,D10,D10-(E10-1)*100))', '客服系统', '每超0.1%扣10分'],
    ['', '合规违规事件', '0起', 15, '', '=IF(E11="","",IF(E11=0,D11,0))', '审计记录', '每起扣15分'],
    ['综合表现', '加分项（月度优秀案例等）', '-', 10, '', '=IF(E12="",0,MIN(E12,D12))', '运营团队评估', '最高10分'],
]

for row_idx, row_data in enumerate(kpi_data, 6):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if col_idx in [1, 2]:  # 维度、指标列加粗
            cell.font = Font(bold=True)

# 合计行
total_row = 12
ws.cell(row=total_row, column=1, value='合计')
ws.cell(row=total_row, column=1).font = Font(bold=True, size=11)
ws.cell(row=total_row, column=4, value=100)
ws.cell(row=total_row, column=4).font = Font(bold=True)
ws.cell(row=total_row, column=6, value='=SUM(F7:F12)')
ws.cell(row=total_row, column=6).font = Font(bold=True, size=12)
ws.cell(row=total_row, column=6).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

# 评级标准
ws['A14'] = '评级标准：'
ws['A14'].font = Font(bold=True)
ws['A15'] = 'A级（优秀）：90-100分'
ws['A16'] = 'B级（良好）：80-89分'
ws['A17'] = 'C级（合格）：60-79分'
ws['A18'] = 'D级（不合格）：<60分'

# 评级结果
ws['D14'] = '本次评级：'
ws['D14'].font = Font(bold=True, size=12)
ws.merge_cells('E14:F14')
ws['E14'] = '=IF(F12>=90,"A级（优秀）",IF(F12>=80,"B级（良好）",IF(F12>=60,"C级（合格）","D级（不合格）")))'
ws['E14'].font = Font(bold=True, size=12, color='C00000')
ws['E14'].alignment = Alignment(horizontal='center')

# 签名区域
ws['A20'] = '供应商确认：'
ws['D20'] = '运营部门审核：'
ws['G20'] = '管理层审批：'

ws['A21'] = '签字：'
ws['D21'] = '签字：'
ws['G21'] = '签字：'

ws['A22'] = '日期：'
ws['D22'] = '日期：'
ws['G22'] = '日期：'

# 设置列宽
col_widths = [15, 25, 12, 8, 10, 12, 15, 20]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 设置行高
for row in range(6, 12):
    ws.row_dimensions[row].height = 35

# 添加边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws['A5:H12']:
    for cell in row:
        cell.border = thin_border

# 保存
wb.save('BPO供应商KPI考核表.xlsx')
print("Excel考核表生成成功！")
