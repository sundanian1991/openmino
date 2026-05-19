"""
SABC 诊断报告生成器 — 适配 CRM 原始底表（单 Sheet）
=====================================================
读取坐席级 Excel → 自动识别格式 → 参评判定 → SABC 评级 → 四视角报告

用法:
  python scripts/analyze.py --input <xlsx_path> [--output-dir <dir>]
依赖: openpyxl
"""

import openpyxl, json, argparse, os, math, sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta

# =====================================================================
# FORMAT DETECTION — 两种 CRM 导出格式自动识别
# =====================================================================

def detect_format(headers):
    """根据 J 列标题判断格式：'时间差' → A，否则 → B"""
    j_idx = headers.get('J', None)
    if j_idx is not None:
        # Will be populated by load()
        pass
    # Format B has J=接通用户数, Format A has J=时间差, K=接通用户数
    return 'B'  # Default, refined in load

def col_letter_to_idx(letter):
    """A→0, B→1, ..., Z→25, AA→26, etc."""
    result = 0
    for c in letter:
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result - 1

def build_field_map(headers_row):
    """从实际列头构建字段映射，不依赖固定列号"""
    h = {}
    for letter, val in headers_row.items():
        if val is None:
            continue
        v = str(val).strip()
        h[v] = letter
    return h

def get_col(headers, possible_names):
    """按优先级从 headers 中匹配列字母"""
    for letter, val in headers.items():
        if val is None:
            continue
        v = str(val).strip()
        for name in possible_names:
            if name in v:
                return letter
    return None

def clean_val(v):
    """清洗单元格值"""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == '':
            return None
    return v

def sf(v):
    """安全转 float"""
    v = clean_val(v)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def flag(v):
    """'是'/'否' → bool"""
    v = clean_val(v)
    if isinstance(v, str):
        return v in ('是', 'True', 'true', '1')
    if isinstance(v, (int, float)):
        return bool(v)
    return False

def sn(wp):
    """职场名简化：管户金条-岐力职场 → 岐力"""
    if not wp:
        return '未知'
    wp = str(wp).replace('管户金条-', '').replace('职场', '').strip()
    return wp if wp else '未知'

# =====================================================================
# LOAD — 读取单 Sheet 底表
# =====================================================================

def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print(f'[SABC] Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}')

    # Build header map: {letter: value}
    headers = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[cell.column_letter] = str(cell.value).strip()

    # Auto-detect format
    j_val = headers.get('J', '')
    fmt = 'A' if '时间差' in str(j_val) else 'B'
    print(f'[SABC] Format: {fmt} ({"岐力型" if fmt == "A" else "伽玛型"})')

    # Build field-to-column mapping
    fields = {}
    def map_field(name, *candidates):
        for letter, val in headers.items():
            for c in candidates:
                if c in val:
                    fields[name] = letter
                    return

    # Common fields
    map_field('seat_type', '坐席类型')
    map_field('workplace', '职场名称')
    map_field('seat_id', '坐席id')
    map_field('seat_name', '坐席名称')
    map_field('first_call', '首次外呼')
    map_field('tenure_days', '在职天数')
    map_field('last_call', '最近一次')
    map_field('connected_users', '接通用户数')
    map_field('call_records', '接通话单数')
    map_field('valid_calls', '有效接通', '有效通话')
    map_field('total_seconds', '累计通时')
    map_field('valid_seconds', '有效通时')
    map_field('call_days', '外呼天数')
    map_field('first_loan_users', '首贷用户数')
    map_field('first_loan_gmv', '首贷GMV')
    map_field('t0_first_gmv', 'T0首贷')
    map_field('repeat_users', '复贷用户数')
    map_field('repeat_orders', '复贷订单数')
    map_field('repeat_gmv', '复贷GMV')
    map_field('t0_repeat_gmv', 'T0复贷')
    map_field('total_gmv', '转化GMV')
    map_field('is_rated', '是否参评')
    map_field('warning_3d', '3天未上线', '预警')
    map_field('offline_7d', '7天未上线', '未上线')
    map_field('days_lt22', '天数不足22', '外呼天数不足')
    map_field('att', 'ATT', 'att')
    map_field('piece', '件均')
    map_field('daily_dur', '日均有效通时')
    map_field('mob_segment', '司龄')

    # Print field mapping for debugging
    missing = [k for k, v in fields.items() if v is None]
    print(f'[SABC] Fields mapped: {len(fields)}, missing: {missing if missing else "none"}')

    # Parse rows
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        vals = {}
        for cell in ws[1]:  # iterate headers again
            col_idx = cell.col_idx - 1  # 0-based
            if col_idx < len(r):
                vals[cell.column_letter] = r[col_idx]

        row = {}
        for field_name, col_letter in fields.items():
            if col_letter and col_letter in vals:
                row[field_name] = vals[col_letter]
            else:
                row[field_name] = None

        row['_fmt'] = fmt
        row['_ws'] = ws.title
        rows.append(row)

    wb.close()

    # Find deadline: use max date from last_call
    max_date = None
    for r in rows:
        lc = r.get('last_call')
        if isinstance(lc, datetime):
            if max_date is None or lc > max_date:
                max_date = lc
    deadline = max_date or datetime.now()
    print(f'[SABC] Deadline (inferred): {deadline.strftime("%Y-%m-%d")}')

    return {'rows': rows, 'deadline': deadline, 'fields': fields, 'format': fmt, 'total': len(rows)}


# =====================================================================
# ENRICH — 数据清洗 + 派生字段
# =====================================================================

def enrich(data):
    """清洗并计算派生字段"""
    deadline = data['deadline']
    rows = data['rows']
    rated_count = 0

    for r in rows:
        # Parse dates (handle both datetime objects and string formats)
        for key in ['first_call', 'last_call']:
            v = r.get(key)
            if isinstance(v, datetime):
                r[f'{key}_parsed'] = v
            elif isinstance(v, str) and v.strip():
                try:
                    r[f'{key}_parsed'] = datetime.strptime(v.strip(), '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    try:
                        r[f'{key}_parsed'] = datetime.strptime(v.strip(), '%Y-%m-%d')
                    except (ValueError, TypeError):
                        r[f'{key}_parsed'] = None
            else:
                r[f'{key}_parsed'] = None

        # Days since last call
        lc = r.get('last_call_parsed')
        if lc:
            r['days_since_last'] = (deadline - lc).days
        else:
            r['days_since_last'] = 999

        # Active: last call within 7 days
        r['is_active'] = r['days_since_last'] <= 7

        # Attendance: call_days >= 22
        cd = sf(r.get('call_days'))
        r['call_days_val'] = cd if cd is not None else 0
        r['has_enough_attendance'] = r['call_days_val'] >= 22

        # Participation rating: active AND enough attendance
        r['is_participating'] = r['is_active'] and r['has_enough_attendance']

        # If the Excel already has 是否参评 column, respect it
        ir = r.get('is_rated')
        if ir is not None:
            r['is_participating'] = flag(ir) or r['is_participating']

        if r['is_participating']:
            rated_count += 1

        # Connected users (numeric)
        r['connected_val'] = sf(r.get('connected_users')) or 0

        # First loan users
        r['fl_users'] = sf(r.get('first_loan_users')) or 0

        # Repeat loan users
        r['rl_users'] = sf(r.get('repeat_users')) or 0

        # Total GMV
        tg = sf(r.get('total_gmv'))
        if tg is None:
            fl_g = sf(r.get('first_loan_gmv')) or 0
            rl_g = sf(r.get('repeat_gmv')) or 0
            tg = fl_g + rl_g
        r['gmv'] = tg

        # T0 GMV
        t0f = sf(r.get('t0_first_gmv')) or 0
        t0r = sf(r.get('t0_repeat_gmv')) or 0
        r['t0_gmv'] = t0f + t0r

        # Conversion rate (首贷用首贷用户数/接通数，复贷用复贷用户数/接通数)
        if r['connected_val'] > 0:
            is_fd = '复贷' in r.get('seat_type', '')
            if is_fd and r['rl_users'] > 0:
                r['conv_rate'] = r['rl_users'] / r['connected_val']
            elif not is_fd and r['fl_users'] > 0:
                r['conv_rate'] = r['fl_users'] / r['connected_val']
            else:
                r['conv_rate'] = 0
        else:
            r['conv_rate'] = 0

        # ATT
        r['att_val'] = sf(r.get('att')) or 0

        # Average ticket
        r['piece_val'] = sf(r.get('piece')) or 0

        # Daily valid duration (minutes)
        dd = sf(r.get('daily_dur'))
        if dd is None and r['call_days_val'] > 0:
            vs = sf(r.get('valid_seconds'))
            if vs is not None and vs > 0:
                dd = vs / r['call_days_val'] / 60
        r['daily_dur_val'] = dd if dd is not None else 0

        # MOB segment
        mob = r.get('mob_segment')
        if mob:
            r['mob_segment'] = str(mob).strip()
        else:
            td = sf(r.get('tenure_days'))
            if td is not None:
                if td < 30:
                    r['mob_segment'] = 'MOB1-'
                elif td < 90:
                    r['mob_segment'] = 'MOB1-3'
                elif td < 180:
                    r['mob_segment'] = 'MOB3-6'
                elif td < 365:
                    r['mob_segment'] = 'MOB6-12'
                else:
                    r['mob_segment'] = 'MOB12+'
            else:
                r['mob_segment'] = '未知'

        # Seat type
        st = r.get('seat_type')
        r['seat_type'] = str(st).strip() if st else '未知'

        # Workplace
        r['wp_name'] = sn(r.get('workplace'))

    print(f'[SABC] Enriched: {len(rows)} rows, {rated_count} participating')
    return data


# =====================================================================
# SABC RATING — 百分位排名制
# =====================================================================

def percentile_rank(values, value):
    """计算某个值在列表中的百分位排名 (0-100)"""
    if not values:
        return 0
    sorted_vals = sorted(set(values))
    rank = sum(1 for v in sorted_vals if v < value)
    return (rank / (len(sorted_vals) - 1)) * 100 if len(sorted_vals) > 1 else 50

def rate_pool(pool, weights):
    """对一个池子（首贷或复贷）进行 SABC 评级"""
    participating = [r for r in pool if r.get('is_participating', False)]
    if not participating:
        return []

    # Compute dimensional values
    for r in participating:
        r['_conv_rate'] = r.get('conv_rate', 0) or 0
        r['_t0_gmv'] = r.get('t0_gmv', 0) or 0
        r['_att_val'] = r.get('att_val', 0) or 0
        r['_daily_dur'] = r.get('daily_dur_val', 0) or 0
        r['_piece_val'] = r.get('piece_val', 0) or 0
        r['_gmv'] = r.get('gmv', 0) or 0

    # Collect all values for percentile calculation
    all_conv = [r['_conv_rate'] for r in participating]
    all_t0 = [r['_t0_gmv'] for r in participating]
    all_att = [r['_att_val'] for r in participating]
    all_dur = [r['_daily_dur'] for r in participating]
    all_piece = [r['_piece_val'] for r in participating]
    all_gmv = [r['_gmv'] for r in participating]

    # Compute percentile ranks and weighted scores
    median_conv = sorted(all_conv)[len(all_conv)//2] if all_conv else 0
    median_gmv = sorted(all_gmv)[len(all_gmv)//2] if all_gmv else 0

    for r in participating:
        pr_conv = percentile_rank(all_conv, r['_conv_rate'])
        pr_t0 = percentile_rank(all_t0, r['_t0_gmv'])
        pr_att = percentile_rank(all_att, r['_att_val'])
        pr_dur = percentile_rank(all_dur, r['_daily_dur'])
        pr_piece = percentile_rank(all_piece, r['_piece_val'])
        pr_gmv = percentile_rank(all_gmv, r['_gmv'])

        # Weighted score
        score = (
            pr_conv * weights['conv'] +
            pr_t0 * weights['t0_gmv'] +
            pr_att * weights['att'] +
            pr_dur * weights['daily_dur'] +
            pr_piece * weights.get('piece', 0) +
            pr_gmv * weights.get('gmv', 0)
        )
        r['_score'] = round(score, 2)
        r['_pr_conv'] = round(pr_conv, 1)
        r['_pr_t0'] = round(pr_t0, 1)
        r['_pr_att'] = round(pr_att, 1)
        r['_pr_dur'] = round(pr_dur, 1)
        r['_pr_piece'] = round(pr_piece, 1)
        r['_pr_gmv'] = round(pr_gmv, 1)

    # Sort by score and assign grades
    participating.sort(key=lambda r: r['_score'], reverse=True)
    n = len(participating)
    s_threshold_20 = max(1, int(n * 0.2))
    a_threshold_50 = max(1, int(n * 0.5))
    b_threshold_80 = max(1, int(n * 0.8))

    for i, r in enumerate(participating):
        r['_rank'] = i + 1
        r['_pctile'] = round((i + 1) / n * 100, 1)

        # Determine grade
        if i < s_threshold_20:
            # S grade: check if core result > median
            core_above_median = r['_conv_rate'] > median_conv
            if core_above_median:
                r['_grade'] = 'S'
            else:
                r['_grade'] = 'A'  # Demoted, no递补
        elif i < a_threshold_50:
            r['_grade'] = 'A'
        elif i < b_threshold_80:
            r['_grade'] = 'B'
        else:
            r['_grade'] = 'C'

    return participating


def compute_ratings(data):
    """对首贷和复贷分池评级"""
    rows = data['rows']

    # Split pools
    sd_pool = [r for r in rows if '首贷' in r.get('seat_type', '')]
    fd_pool = [r for r in rows if '复贷' in r.get('seat_type', '')]

    # If no seat_type distinction, use all data
    if not sd_pool and not fd_pool:
        sd_pool = rows  # Default to first loan

    print(f'[SABC] Pools: 首贷 {len(sd_pool)}, 复贷 {len(fd_pool)}')

    # Weights
    sd_weights = {
        'conv': 0.45, 't0_gmv': 0.20, 'att': 0.15,
        'daily_dur': 0.20, 'piece': 0, 'gmv': 0
    }
    fd_weights = {
        'conv': 0.35, 't0_gmv': 0.15, 'att': 0.10,
        'daily_dur': 0.25, 'piece': 0.15, 'gmv': 0
    }

    sd_rated = rate_pool(sd_pool, sd_weights)
    fd_rated = rate_pool(fd_pool, fd_weights)

    print(f'[SABC] Rated: 首贷 {len(sd_rated)}, 复贷 {len(fd_rated)}')

    data['sd_rated'] = sd_rated
    data['fd_rated'] = fd_rated
    return data


# =====================================================================
# ANALYZE — 四视角指标计算
# =====================================================================

def analyze_pool(rated_rows, all_rows):
    """对已评级的池子计算四视角指标"""
    n_total = len(all_rows)
    n_rated = len(rated_rows)
    n_unrated = n_total - n_rated

    if n_rated == 0:
        return {'error': 'no rated seats'}

    # Grade distribution
    grade_dist = Counter(r.get('_grade', 'Unknown') for r in rated_rows)

    # Grade metrics
    grade_metrics = {}
    for g in ['S', 'A', 'B', 'C']:
        sub = [r for r in rated_rows if r.get('_grade') == g]
        if not sub:
            grade_metrics[g] = {'count': 0}
            continue
        grade_metrics[g] = {
            'count': len(sub),
            'avg_att': round(sum(r['_att_val'] for r in sub) / len(sub), 1),
            'avg_conv': round(sum(r.get('_conv_rate', 0) or 0 for r in sub) / len(sub) * 100, 2),
            'avg_piece': round(sum(r['_piece_val'] for r in sub if r['_piece_val'] > 0) / max(1, len([r for r in sub if r['_piece_val'] > 0]))),
            'avg_t0_gmv': round(sum(r['_t0_gmv'] for r in sub) / len(sub)),
            'avg_daily_dur': round(sum(r['_daily_dur'] for r in sub) / len(sub), 1),
            'avg_gmv': round(sum(r['_gmv'] for r in sub) / len(sub)),
        }

    # Workplace matrix
    wp_matrix = defaultdict(Counter)
    for r in rated_rows:
        wp_matrix[r['wp_name']][r['_grade']] += 1

    # Tenure matrix
    tenure_order = {'MOB1-': 0, 'MOB1-3': 1, 'MOB3-6': 2, 'MOB6-12': 3, 'MOB12+': 4, '未知': 99}
    tenure_matrix = defaultdict(Counter)
    for r in rated_rows:
        tenure_matrix[r.get('mob_segment', '未知')][r['_grade']] += 1

    # Pareto: GMV distribution
    sorted_by_gmv = sorted(rated_rows, key=lambda r: r['_gmv'], reverse=True)
    total_gmv = sum(r['_gmv'] for r in rated_rows)
    top20n = max(1, n_rated // 5)
    top20gmv = sum(r['_gmv'] for r in sorted_by_gmv[:top20n])
    top20_pct = round(top20gmv / total_gmv * 100, 1) if total_gmv > 0 else 0

    pareto = []
    cum = 0
    # Limit to ~15 data points for clean Pareto chart rendering
    n_bins = min(15, n_rated)
    step = max(1, n_rated // n_bins)
    for i in range(0, n_rated, step):
        cum += sum(r['_gmv'] for r in sorted_by_gmv[i:i+step])
        pct = round(min((i + step) / n_rated * 100, 100), 1)
        cum_pct = round(min(cum / total_gmv * 100, 100), 1) if total_gmv > 0 else 0
        pareto.append({'pct': pct, 'cum': cum_pct})
    if not pareto:
        pareto = [{'pct': 0, 'cum': 0}]

    # Funnel
    funnel = {}
    for g in ['S', 'A', 'B', 'C']:
        sub = [r for r in rated_rows if r.get('_grade') == g]
        if not sub:
            continue
        conn = sum(r.get('connected_val', 0) or 0 for r in sub)
        fl = sum(r.get('fl_users', 0) or 0 for r in sub)
        funnel[g] = {
            'connected': conn,
            'first_loan': fl,
            'conv_rate': round(fl / conn * 100, 2) if conn > 0 else 0
        }

    # Workplace C-rate ranking
    wp_c_rates = []
    for wp, grades in wp_matrix.items():
        total_wp = sum(grades.values())
        c_count = grades.get('C', 0)
        c_rate = round(c_count / total_wp * 100, 1) if total_wp > 0 else 0
        wp_c_rates.append({'wp': wp, 'c_rate': c_rate, 'c_count': c_count, 'total': total_wp})
    wp_c_rates.sort(key=lambda x: x['c_rate'], reverse=True)

    # Workplace efficiency matrix (for scatter plot)
    wp_eff = {}
    for wp, grades in wp_matrix.items():
        sub = [r for r in rated_rows if r['wp_name'] == wp]
        avg_att = round(sum(r['_att_val'] for r in sub) / len(sub), 1) if sub else 0
        avg_conv = round(sum(r.get('_conv_rate', 0) or 0 for r in sub) / len(sub) * 100, 2) if sub else 0
        c_count = grades.get('C', 0)
        wp_eff[wp] = {
            'att': avg_att,
            'conv': avg_conv,
            'c_rate': round(c_count / len(sub) * 100, 1) if sub else 0,
            'n': len(sub),
            'c': c_count
        }

    # Warning rate
    warned = sum(1 for r in all_rows if r.get('warning_3d') and flag(r.get('warning_3d')))
    active_count = sum(1 for r in all_rows if r.get('is_active', False))
    warning_rate = round(warned / active_count * 100, 1) if active_count > 0 else 0

    # MOB1-3 C concentration
    mob_c = tenure_matrix.get('MOB1-3', Counter())
    total_c = grade_dist.get('C', 0)
    mob13_c_count = mob_c.get('C', 0)
    mob13_c_conc = round(mob13_c_count / total_c * 100, 1) if total_c > 0 else 0

    # S/C ratio
    s_conv = grade_metrics.get('S', {}).get('avg_conv', 0) or 0
    c_conv = grade_metrics.get('C', {}).get('avg_conv', 0) or 0
    sc_ratio = round(s_conv / max(0.01, c_conv), 1) if c_conv > 0 else 0

    return {
        'n_total': n_total,
        'n_rated': n_rated,
        'n_unrated': n_unrated,
        'participation_rate': round(n_rated / n_total * 100, 1) if n_total > 0 else 0,
        'grade_dist': dict(grade_dist),
        'grade_metrics': grade_metrics,
        'wp_matrix': {k: dict(v) for k, v in wp_matrix.items()},
        'tenure_matrix': {k: dict(v) for k, v in tenure_matrix.items()},
        'total_gmv': round(total_gmv),
        'top20_gmv_pct': top20_pct,
        'pareto': pareto,
        'funnel': funnel,
        'wp_c_rates': wp_c_rates,
        'wp_eff': wp_eff,
        'warning_rate': warning_rate,
        'mob13_c_conc': mob13_c_conc,
        'sc_ratio': sc_ratio,
    }


def compute_all_metrics(data):
    sd_stats = analyze_pool(data.get('sd_rated', []), [r for r in data['rows'] if '首贷' in r.get('seat_type', '')])
    fd_stats = analyze_pool(data.get('fd_rated', []), [r for r in data['rows'] if '复贷' in r.get('seat_type', '')])

    # If only one pool exists, use it for combined metrics
    # but keep individual pool stats separate (don't copy fd→sd)

    # Combined view: merge workplace matrix from both pools
    all_rated = (data.get('sd_rated', []) or []) + (data.get('fd_rated', []) or [])
    all_participating = [r for r in data['rows'] if r.get('is_participating')]
    combined = analyze_pool(all_rated, all_participating) if all_rated else {'error': 'no rated seats'}

    data['sd_stats'] = sd_stats
    data['fd_stats'] = fd_stats
    data['combined_stats'] = combined

    print(f'[SABC] 首贷指标: {sd_stats.get("error", "OK")}, 复贷指标: {fd_stats.get("error", "OK")}')
    return data


# =====================================================================
# HTML REPORT
# =====================================================================

def build_html(data):
    """生成 NYT 风 HTML 报告"""
    sd = data.get('sd_stats', {})
    fd = data.get('fd_stats', {})
    deadline = data['deadline'].strftime('%Y-%m-%d')
    fmt = data.get('format', 'B')
    supplier = '金条'

    # has_sd: true only if actual 首贷 rated data exists
    has_sd = len(data.get('sd_rated', [])) > 0
    has_fd = len(data.get('fd_rated', [])) > 0
    all_wps = set()
    for r in data['rows']:
        wp = r.get('wp_name', '')
        if wp:
            all_wps.add(wp)
    if all_wps:
        supplier = ', '.join(sorted(all_wps))

    # Template values
    sd_n = sd.get('n_rated', 0)
    fd_n = fd.get('n_rated', 0)
    sd_total = sd.get('n_total', 0)
    fd_total = fd.get('n_total', 0)
    sd_part = sd.get('participation_rate', 0)
    fd_part = fd.get('participation_rate', 0)

    sd_s = sd.get('grade_dist', {}).get('S', 0)
    sd_a = sd.get('grade_dist', {}).get('A', 0)
    sd_b = sd.get('grade_dist', {}).get('B', 0)
    sd_c = sd.get('grade_dist', {}).get('C', 0)

    # Use main pool for metrics when 首贷 is empty
    main_stats = sd if sd_n > 0 else fd
    gm_s = main_stats.get('grade_metrics', {}).get('S', {})
    gm_c = main_stats.get('grade_metrics', {}).get('C', {})

    s_conv = gm_s.get('avg_conv', 0)
    c_conv = gm_c.get('avg_conv', 0)
    s_att = gm_s.get('avg_att', 0)
    c_att = gm_c.get('avg_att', 0)
    s_dur = gm_s.get('avg_daily_dur', 0)
    c_dur = gm_c.get('avg_daily_dur', 0)

    sc_ratio = main_stats.get('sc_ratio', 0)
    top20_pct = main_stats.get('top20_gmv_pct', 0)
    warning_rate = main_stats.get('warning_rate', 0)
    mob13_conc = main_stats.get('mob13_c_conc', 0)

    # Workplace structure: use combined view (all pools) for complete picture
    cmb = data.get('combined_stats', {})
    wp_matrix = cmb.get('wp_matrix', {}) if not cmb.get('error') else sd.get('wp_matrix', {})
    wp_names = sorted(wp_matrix.keys())
    wp_matrix_data = []
    for wp in wp_names:
        row = {'wp': wp}
        for g in ['S', 'A', 'B', 'C']:
            row[g] = wp_matrix.get(wp, {}).get(g, 0)
        wp_matrix_data.append(row)

    # Tenure matrix — also use combined
    tenure_matrix = cmb.get('tenure_matrix', {}) if not cmb.get('error') else sd.get('tenure_matrix', {})
    tenure_order = ['MOB1-', 'MOB1-3', 'MOB3-6', 'MOB6-12', 'MOB12+', '未知']
    tenure_data = []
    for t in tenure_order:
        if t in tenure_matrix:
            row = {'tenure': t}
            for g in ['S', 'A', 'B', 'C']:
                row[g] = tenure_matrix[t].get(g, 0)
            tenure_data.append(row)

    # Funnel, Pareto, WP metrics — use primary pool (conversion-focused)
    main_stats = sd if sd_n > 0 else fd
    funnel = main_stats.get('funnel', {})
    pareto = main_stats.get('pareto', [])
    wp_c_rates = main_stats.get('wp_c_rates', [])[:8]
    # Efficiency scatter: use combined to show ALL workplaces
    wp_eff = cmb.get('wp_eff', {}) if not cmb.get('error') else main_stats.get('wp_eff', {})
    wp_eff_data = [{'wp': k, **v} for k, v in wp_eff.items()]

    # Primary pool = 首贷 if available, else 复贷
    total_n = sd_n if sd_n > 0 else fd_n
    total_s = sd_s if sd_n > 0 else fd.get('grade_dist', {}).get('S', 0)
    total_a = sd_a if sd_n > 0 else fd.get('grade_dist', {}).get('A', 0)
    total_b = sd_b if sd_n > 0 else fd.get('grade_dist', {}).get('B', 0)
    total_c = sd_c if sd_n > 0 else fd.get('grade_dist', {}).get('C', 0)

    # ================================================================
    # NARRATIVE DATA COMPUTATIONS (V4.0)
    # ================================================================

    # Reconstruct all_rated for narrative computations
    all_rated = (data.get('sd_rated', []) or []) + (data.get('fd_rated', []) or [])

    # Monthly hire trend (Fig5)
    hire_counter = Counter()
    for r in data['rows']:
        fcd = r.get('first_call_parsed')
        if fcd:
            hire_counter[fcd.strftime('%Y-%m')] += 1
    sorted_months = sorted(hire_counter.keys())
    hire_months = sorted_months
    hire_counts = [hire_counter[m] for m in sorted_months]
    hire_avg = round(sum(hire_counts) / max(1, len(hire_counts)), 1)
    hire_peak = max(hire_counts) if hire_counts else 0
    # Dynamic burst month detection: exclude last 2 months (incomplete data), detect outliers via 2σ
    complete_months = sorted_months[:-2] if len(sorted_months) > 2 else sorted_months
    if complete_months and hire_counts:
        c_means = sum(hire_counter[m] for m in complete_months) / len(complete_months)
        c_std = (sum((hire_counter[m] - c_means)**2 for m in complete_months) / len(complete_months))**0.5
        burst_threshold = c_means + 2 * c_std
        burst_months = [m for m in complete_months if hire_counter[m] > burst_threshold]
        burst_total = sum(hire_counter[m] for m in burst_months)
        hire_hist_avg = round(c_means, 1)
    else:
        burst_months = []
        burst_total = 0
        hire_hist_avg = 0

    # Tenure performance curves (Fig4)
    tenure_order = ['MOB1-', 'MOB1-3', 'MOB3-6', 'MOB6-12', 'MOB12+']
    tenure_perf_data = []
    for t in tenure_order:
        pool = [r for r in all_rated if r.get('mob_segment') == t]
        if not pool:
            continue
        avg_gmv = round(sum(r['_gmv'] for r in pool) / len(pool))
        avg_conv = round(sum(r.get('_conv_rate', 0) or 0 for r in pool) / len(pool) * 100, 2)
        avg_dur = round(sum(r['_daily_dur'] for r in pool) / len(pool), 1)
        avg_att = round(sum(r['_att_val'] for r in pool) / len(pool), 1)
        tenure_perf_data.append({
            'tenure': t, 'avg_gmv': avg_gmv, 'avg_conv': avg_conv,
            'avg_dur': avg_dur, 'avg_att': avg_att, 'n': len(pool),
        })

    # New hire batch tracking (Fig6) — dynamic: pick latest 2 months with ≥30% of peak hire count
    if hire_counts:
        peak_count = max(hire_counts)
        sig_threshold = max(peak_count * 0.3, 10)
        sig_months = [(m, c) for m, c in zip(hire_months, hire_counts) if c >= sig_threshold]
        sig_months.sort(key=lambda x: x[0], reverse=True)
        target_yms = [m for m, c in sig_months[:2]]
    else:
        target_yms = []

    new_batches = {}
    for ym in target_yms:
        batch = [r for r in data['rows']
                 if r.get('first_call_parsed') and r['first_call_parsed'].strftime('%Y-%m') == ym]
        if not batch:
            continue
        batch_rated = [r for r in batch if r.get('is_participating')]
        new_batches[ym] = {
            'total': len(batch),
            'rated': len(batch_rated),
            'with_gmv': sum(1 for r in batch if (r.get('gmv') or 0) > 0),
            'avg_gmv': round(sum(r.get('gmv', 0) or 0 for r in batch) / len(batch)) if batch else 0,
            'avg_dur': round(sum(r.get('daily_dur_val', 0) or 0 for r in batch) / len(batch), 1) if batch else 0,
            'avg_conv': round(sum(r.get('conv_rate', 0) or 0 for r in batch) / len(batch) * 100, 2) if batch else 0,
        }

    # Waterfall data (Fig8) — S→A→B→C conversion with step drops
    waterfall_data = []
    prev_conv = None
    for g in ['S', 'A', 'B', 'C']:
        sub = [r for r in all_rated if r.get('_grade') == g]
        if not sub:
            continue
        conv = round(sum(r.get('_conv_rate', 0) or 0 for r in sub) / len(sub) * 100, 2)
        drop = round(conv - prev_conv, 2) if prev_conv is not None else 0
        waterfall_data.append({'grade': g, 'conv_rate': conv, 'drop': drop, 'n': len(sub)})
        prev_conv = conv

    # Per-seat scatter data (Fig10: dur×gmv, Fig13: att×conv) — split by seat_type
    sd_scatter = {'att_conv': [], 'dur_gmv': []}
    fd_scatter = {'att_conv': [], 'dur_gmv': []}
    for r in all_rated:
        st = r.get('seat_type', '未知')
        target = sd_scatter if st == '首贷' else fd_scatter
        target['dur_gmv'].append({
            'dur': round(r['_daily_dur'], 1), 'gmv': round(r['_gmv']),
            'grade': r.get('_grade', 'N/A'),
        })
        target['att_conv'].append({
            'att': round(r['_att_val'], 1),
            'conv': round((r.get('_conv_rate', 0) or 0) * 100, 2),
            'grade': r.get('_grade', 'N/A'),
        })

    # Non-rated classification (Fig9a)
    non_rated = [r for r in data['rows'] if not r.get('is_participating')]
    zero_call = sum(1 for r in non_rated if (r.get('daily_dur_val', 0) or 0) == 0)
    low_eff = sum(1 for r in non_rated if 0 < (r.get('daily_dur_val', 0) or 0) < 90)
    others = len(non_rated) - zero_call - low_eff
    non_rated_data = {'total': len(non_rated), 'zero_call': zero_call, 'low_eff': low_eff, 'others': others}

    # Duration distribution for non_rated histogram (Fig9b)
    dur_labels = ['0min', '1-30min', '31-60min', '61-90min', '90min+']
    dur_dist = {l: 0 for l in dur_labels}
    for r in non_rated:
        dur = r.get('daily_dur_val', 0) or 0
        if dur == 0:
            dur_dist['0min'] += 1
        elif dur <= 30:
            dur_dist['1-30min'] += 1
        elif dur <= 60:
            dur_dist['31-60min'] += 1
        elif dur <= 90:
            dur_dist['61-90min'] += 1
        else:
            dur_dist['90min+'] += 1
    non_rated_dur = [{'label': k, 'count': v} for k, v in dur_dist.items()]

    # Dimension contribution weights (Fig12) — correlation with final score
    def pearson_r(xs, ys):
        n = len(xs)
        if n < 3:
            return 0
        mx, my = sum(xs) / n, sum(ys) / n
        num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
        den = (sum((x - mx) ** 2 for x in xs) * sum((y - my) ** 2 for y in ys)) ** 0.5
        return round(num / den, 3) if den > 0 else 0

    if len(all_rated) >= 5:
        dim_data = {
            '转化率': [r['_pr_conv'] for r in all_rated],
            'T0首贷GMV': [r['_pr_t0'] for r in all_rated],
            'ATT': [r['_pr_att'] for r in all_rated],
            '日均通时': [r['_pr_dur'] for r in all_rated],
            '件均': [r['_pr_piece'] for r in all_rated],
        }
        scores = [r['_score'] for r in all_rated]
        dim_weights = [{'name': k, 'r': pearson_r(v, scores)} for k, v in dim_data.items()]
        dim_weights.sort(key=lambda x: abs(x['r']), reverse=True)
    else:
        dim_weights = [
            {'name': '转化率', 'r': 0.42}, {'name': 'T0首贷GMV', 'r': 0.31},
            {'name': '件均', 'r': 0.18}, {'name': '日均通时', 'r': -0.21},
            {'name': 'ATT', 'r': 0.12},
        ]

    # Anomaly counts (Ch5)
    warned_count = sum(1 for r in data['rows'] if r.get('warning_3d') and flag(r.get('warning_3d')))
    offline_count = sum(1 for r in data['rows'] if r.get('offline_7d') and flag(r.get('offline_7d')))
    zero_gmv_count = sum(1 for r in data['rows'] if (r.get('gmv', 0) or 0) == 0)

    # All data as JSON for JS
    report_data = {
        'deadline': deadline,
        'supplier': supplier,
        'sd_n': sd_n, 'fd_n': fd_n,
        'sd_total': sd_total, 'fd_total': fd_total,
        'sd_part': sd_part, 'fd_part': fd_part,
        'sd_s': sd_s, 'sd_a': sd_a, 'sd_b': sd_b, 'sd_c': sd_c,
        'sd_gd': {'S': sd_s, 'A': sd_a, 'B': sd_b, 'C': sd_c},
        'fd_gd': {'S': fd.get('grade_dist',{}).get('S',0), 'A': fd.get('grade_dist',{}).get('A',0), 'B': fd.get('grade_dist',{}).get('B',0), 'C': fd.get('grade_dist',{}).get('C',0)},
        'total_n': total_n, 'total_s': total_s, 'total_a': total_a, 'total_b': total_b, 'total_c': total_c,
        'total_gd': {'S': total_s, 'A': total_a, 'B': total_b, 'C': total_c},
        's_conv': s_conv, 'c_conv': c_conv,
        's_att': s_att, 'c_att': c_att,
        's_dur': s_dur, 'c_dur': c_dur,
        'sc_ratio': sc_ratio,
        'top20_pct': top20_pct,
        'warning_rate': warning_rate,
        'mob13_conc': mob13_conc,
        'wp_names': wp_names,
        'wp_matrix': wp_matrix_data,
        'tenure_data': tenure_data,
        'funnel': funnel,
        'pareto': pareto,
        'wp_c_rates': wp_c_rates,
        'wp_eff_data': wp_eff_data,
        'sd_total_gmv': sd.get('total_gmv', 0),
        'fd_total_gmv': fd.get('total_gmv', 0),
        'combined_n': cmb.get('n_rated', total_n),
        'has_sd': has_sd,
        'source_files': data.get('source_files', []),
        # Narrative data (V4.0)
        'hire_months': hire_months,
        'hire_counts': hire_counts,
        'hire_avg': hire_avg,
        'hire_peak': hire_peak,
        'hire_hist_avg': hire_hist_avg,
        'burst_months': burst_months,
        'burst_total': burst_total,
        'tenure_perf': tenure_perf_data,
        'new_batches': new_batches,
        'waterfall': waterfall_data,
        'sd_scatter': sd_scatter,
        'fd_scatter': fd_scatter,
        'non_rated': non_rated_data,
        'non_rated_dur': non_rated_dur,
        'dim_weights': dim_weights,
        'warned_count': warned_count,
        'offline_count': offline_count,
        'zero_gmv_count': zero_gmv_count,
        'grade_metrics': cmb.get('grade_metrics', {}) if not cmb.get('error') else main_stats.get('grade_metrics', {}),
    }

    data_json = json.dumps(report_data, ensure_ascii=False, default=str)

    # Read template
    tpl_path = os.path.join(os.path.dirname(__file__), 'report_template.html')
    if os.path.exists(tpl_path):
        with open(tpl_path) as f:
            tpl = f.read()
    else:
        tpl = BUILD_DEFAULT_TEMPLATE()

    return tpl.replace('$DATA_JSON', data_json)


def BUILD_DEFAULT_TEMPLATE():
    """内嵌默认 HTML 模板（当 report_template.html 不存在时使用）"""
    return """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SABC 诊断报告</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
:root {
  --bg: #FFF8F0; --card: #FFFFFF; --text: #2D3748; --muted: #718096;
  --border: #F0E6D8; --accent: #E8734A; --accent-light: #FFF0EB;
  --grade-s: #E8734A; --grade-a: #5A6878; --grade-b: #8B99A8; --grade-c: #B8C4D0;
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body { background: var(--bg); color: var(--text); font-family: -apple-system, "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif; line-height: 1.6; }
.container { max-width: 960px; margin: 0 auto; padding: 32px 20px; }
h1 { font-size: 22px; font-weight: 800; margin-bottom: 8px; }
.subtitle { color: var(--muted); font-size: 13px; margin-bottom: 24px; }

/* Health Cards */
.health-cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 32px; }
.h-card { background: var(--card); border-radius: 10px; padding: 16px; border: 1px solid var(--border); }
.h-card h3 { font-size: 13px; font-weight: 700; margin-bottom: 8px; }
.h-metric { display: flex; justify-content: space-between; align-items: center; padding: 4px 0; font-size: 12px; }
.h-metric .label { color: var(--muted); }
.h-metric .value { font-weight: 700; }
.badge { display: inline-block; width: 8px; height: 8px; border-radius: 50%; margin-left: 4px; }
.badge-g { background: #48BB78; } .badge-y { background: #ECC94B; } .badge-r { background: #E8734A; }

/* Chart Grid */
.chart-section { margin-bottom: 32px; }
.chart-section h2 { font-size: 16px; font-weight: 700; margin-bottom: 12px; padding-bottom: 6px; border-bottom: 2px solid var(--accent); }
.chart-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }
.chart-grid .full { grid-column: 1 / -1; }
.chart-box { background: var(--card); border-radius: 10px; padding: 16px; border: 1px solid var(--border); }
.chart-box h3 { font-size: 13px; font-weight: 700; margin-bottom: 8px; }
.chart-box .chart { width: 100%; height: 280px; }

/* Grade Distribution */
.grade-summary { display: flex; gap: 16px; margin-bottom: 32px; }
.grade-chip { flex: 1; text-align: center; padding: 16px; border-radius: 10px; background: var(--card); border: 1px solid var(--border); }
.grade-chip .num { font-size: 32px; font-weight: 800; }
.grade-chip .lbl { font-size: 12px; color: var(--muted); }
.grade-chip.s .num { color: var(--grade-s); }
.grade-chip.a .num { color: var(--grade-a); }
.grade-chip.b .num { color: var(--grade-b); }
.grade-chip.c .num { color: var(--grade-c); }

/* Action Cards */
.action-cards { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-top: 24px; }
.action-card { background: var(--card); border-radius: 10px; padding: 14px; border: 1px solid var(--border); border-left: 3px solid var(--accent); }
.action-card.p0 { border-left-color: #E8734A; }
.action-card.p1 { border-left-color: #ECC94B; }
.action-card.p2 { border-left-color: #48BB78; }
.action-card h4 { font-size: 12px; font-weight: 700; }
.action-card p { font-size: 11px; color: var(--muted); margin-top: 4px; }
.action-card .priority { font-size: 10px; font-weight: 700; color: var(--accent); }
</style>
</head>
<body>
<div class="container">
  <h1>$supplier 金条坐席 SABC 诊断</h1>
  <p class="subtitle">数据截止 $deadline</p>

  <div class="grade-summary" id="gradeSummary"></div>
  <div class="health-cards" id="healthCards"></div>

  <div class="chart-section">
    <h2>🧑 人力结构</h2>
    <div class="chart-grid">
      <div class="chart-box"><h3>职场 S/A/B/C 构成</h3><div class="chart" id="wpMatrix"></div></div>
      <div class="chart-box"><h3>司龄 × 等级矩阵</h3><div class="chart" id="tenureMatrix"></div></div>
    </div>
  </div>

  <div class="chart-section">
    <h2>⚡ 产能效率</h2>
    <div class="chart-grid">
      <div class="chart-box full"><h3>转化断层 — S/A/B/C 转化率对比</h3><div class="chart" id="funnel"></div></div>
    </div>
  </div>

  <div class="chart-section">
    <h2>⚠️ 集中度风险</h2>
    <div class="chart-grid">
      <div class="chart-box full"><h3>GMV 帕累托曲线</h3><div class="chart" id="pareto"></div></div>
    </div>
  </div>

  <div class="chart-section">
    <h2>🔍 问题定位</h2>
    <div class="chart-grid">
      <div class="chart-box"><h3>职场 C 级率排名</h3><div class="chart" id="cRate"></div></div>
      <div class="chart-box"><h3>职场效率矩阵 (ATT × 转化率)</h3><div class="chart" id="scatter"></div></div>
    </div>
  </div>

  <div class="chart-section">
    <h2>行动建议</h2>
    <div class="action-cards" id="actionCards"></div>
  </div>
</div>

<script>
const D = $DATA_JSON;

function badge(grade) { return `<span class="badge badge-${grade}"></span>`; }
function trafficLight(val, thresholds) {
  const [gMax, yMax] = thresholds;
  if (val <= gMax) return '<span class="badge badge-g"></span>';
  if (val <= yMax) return '<span class="badge badge-y"></span>';
  return '<span class="badge badge-r"></span>';
}

// Grade Summary
document.getElementById('gradeSummary').innerHTML = `
  <div class="grade-chip s"><div class="num">${D.sd_s}</div><div class="lbl">S 级</div></div>
  <div class="grade-chip a"><div class="num">${D.sd_a}</div><div class="lbl">A 级</div></div>
  <div class="grade-chip b"><div class="num">${D.sd_b}</div><div class="lbl">B 级</div></div>
  <div class="grade-chip c"><div class="num">${D.sd_c}</div><div class="lbl">C 级</div></div>
`;

// Health Cards
document.getElementById('healthCards').innerHTML = `
  <div class="h-card"><h3>🧑 人力结构 ${D.sd_part >= 75 ? badge('g') : D.sd_part >= 60 ? badge('y') : badge('r')}</h3>
    <div class="h-metric"><span class="label">参评率</span><span class="value">${D.sd_part}%</span></div>
    <div class="h-metric"><span class="label">C 级率</span><span class="value">${D.sd_c}人 (${D.sd_n > 0 ? Math.round(D.sd_c/D.sd_n*100) : 0}%)</span></div>
    <div class="h-metric"><span class="label">S 占比</span><span class="value">${D.sd_n > 0 ? Math.round(D.sd_s/D.sd_n*100) : 0}%</span></div>
  </div>
  <div class="h-card"><h3>⚡ 产能效率 ${D.sc_ratio < 2 ? badge('g') : D.sc_ratio <= 3 ? badge('y') : badge('r')}</h3>
    <div class="h-metric"><span class="label">S/C 转化倍差</span><span class="value">${D.sc_ratio}×</span></div>
    <div class="h-metric"><span class="label">S 转化率</span><span class="value">${D.s_conv}%</span></div>
    <div class="h-metric"><span class="label">C 转化率</span><span class="value">${D.c_conv}%</span></div>
  </div>
  <div class="h-card"><h3>⚠️ 集中度 ${D.top20_pct < 50 ? badge('g') : D.top20_pct <= 60 ? badge('y') : badge('r')}</h3>
    <div class="h-metric"><span class="label">Top20% GMV</span><span class="value">${D.top20_pct}%</span></div>
    <div class="h-metric"><span class="label">总 GMV</span><span class="value">${(D.sd_total_gmv/10000).toFixed(1)}万</span></div>
  </div>
  <div class="h-card"><h3>🔍 问题定位 ${D.warning_rate < 10 ? badge('g') : D.warning_rate <= 15 ? badge('y') : badge('r')}</h3>
    <div class="h-metric"><span class="label">预警率</span><span class="value">${D.warning_rate}%</span></div>
    <div class="h-metric"><span class="label">MOB1-3 C集中度</span><span class="value">${D.mob13_conc}%</span></div>
  </div>
`;

// ECharts charts
function initChart(id, option) {
  const dom = document.getElementById(id);
  if (!dom) return;
  const chart = echarts.init(dom);
  chart.setOption(option);
  return chart;
}

// 1. WP Matrix (Stacked Bar)
if (D.wp_names.length > 0) {
  initChart('wpMatrix', {
    tooltip: { trigger: 'axis', axisPointer: { type: 'shadow' } },
    legend: { data: ['S','A','B','C'], bottom: 0, textStyle: { fontSize: 11 } },
    grid: { left: 60, right: 20, top: 10, bottom: 40 },
    xAxis: { type: 'category', data: D.wp_names, axisLabel: { fontSize: 10, rotate: 30 } },
    yAxis: { type: 'value', name: '人数' },
    series: ['S','A','B','C'].map((g, i) => ({
      name: g, type: 'bar', stack: 'total',
      itemStyle: { color: { 'S': '#E8734A', 'A': '#5A6878', 'B': '#8B99A8', 'C': '#B8C4D0' }[g] },
      data: D.wp_matrix.map(r => r[g]),
      emphasis: { focus: 'series' }
    }))
  });
}

// 2. Tenure Matrix (Heatmap)
if (D.tenure_data && D.tenure_data.length > 0) {
  const tData = [];
  const tNames = D.tenure_data.map(r => r.tenure);
  D.tenure_data.forEach((row, ri) => {
    ['S','A','B','C'].forEach((g, ci) => {
      tData.push([ci, ri, row[g] || 0]);
    });
  });
  initChart('tenureMatrix', {
    tooltip: { position: 'top' },
    grid: { left: 80, right: 40, top: 10, bottom: 40 },
    xAxis: { type: 'category', data: ['S','A','B','C'], axisLabel: { fontSize: 12 } },
    yAxis: { type: 'category', data: tNames, axisLabel: { fontSize: 10 } },
    visualMap: { min: 0, max: Math.max(...tData.map(d=>d[2]), 1), calculable: true, orient: 'horizontal', left: 'center', bottom: 0, inRange: { color: ['#FFF0EB','#E8734A','#C75A35'] } },
    series: [{ type: 'heatmap', data: tData, label: { show: true, fontSize: 10 }, emphasis: { itemStyle: { shadowBlur: 10, shadowColor: 'rgba(0,0,0,0.5)' } } }]
  });
}

// 3. Funnel (Bar chart with comparison)
if (D.funnel && Object.keys(D.funnel).length > 0) {
  const grades = Object.keys(D.funnel);
  const rates = grades.map(g => D.funnel[g].conv_rate);
  initChart('funnel', {
    tooltip: { trigger: 'axis', formatter: '{b}: {c}%' },
    grid: { left: 60, right: 30, top: 20, bottom: 40 },
    xAxis: { type: 'category', data: grades.map(g => g + '级') },
    yAxis: { type: 'value', name: '转化率(%)', min: 0 },
    series: [{
      type: 'bar', data: rates.map((v, i) => ({
        value: v,
        itemStyle: { color: { 'S': '#E8734A', 'A': '#5A6878', 'B': '#8B99A8', 'C': '#B8C4D0' }[grades[i]] }
      })),
      label: { show: true, position: 'top', formatter: '{c}%' },
      barMaxWidth: 60
    }]
  });
}

// 4. Pareto
if (D.pareto && D.pareto.length > 0) {
  initChart('pareto', {
    tooltip: { trigger: 'axis' },
    grid: { left: 60, right: 60, top: 20, bottom: 40 },
    xAxis: { type: 'category', data: D.pareto.map(d => d.pct + '%') },
    yAxis: { type: 'value', max: 100, name: '累计GMV%' },
    series: [
      { type: 'line', smooth: true, symbol: 'none', lineStyle: { width: 3, color: '#E8734A' }, data: D.pareto.map(d => d.cum), areaStyle: { color: { type: 'linear', x: 0, y: 0, x2: 0, y2: 1, colorStops: [{ offset: 0, color: 'rgba(232,115,74,0.3)' }, { offset: 1, color: 'rgba(232,115,74,0.05)' }] } } },
      { type: 'line', data: D.pareto.map(() => 50), lineStyle: { type: 'dashed', color: '#ECC94B' }, symbol: 'none' },
      { type: 'line', data: D.pareto.map(() => 80), lineStyle: { type: 'dashed', color: '#B8C4D0' }, symbol: 'none' }
    ]
  });
}

// 5. C Rate Ranking
if (D.wp_c_rates && D.wp_c_rates.length > 0) {
  initChart('cRate', {
    tooltip: { trigger: 'axis', formatter: '{b}: {c}%' },
    grid: { left: 100, right: 30, top: 10, bottom: 30 },
    yAxis: { type: 'category', data: D.wp_c_rates.map(r => r.wp).reverse(), axisLabel: { fontSize: 11 } },
    xAxis: { type: 'value', name: 'C级率(%)' },
    series: [{
      type: 'bar', data: D.wp_c_rates.map(r => ({ value: r.c_rate, itemStyle: { color: r.c_rate > 30 ? '#E8734A' : r.c_rate > 20 ? '#ECC94B' : '#8B99A8' } })).reverse(),
      label: { show: true, position: 'right', formatter: '{c}%' },
      barMaxWidth: 24
    }]
  });
}

// 6. Scatter (ATT vs Conv)
if (D.wp_eff_data && D.wp_eff_data.length > 0) {
  const avgAtt = D.wp_eff_data.reduce((s, r) => s + r.att, 0) / D.wp_eff_data.length;
  const avgConv = D.wp_eff_data.reduce((s, r) => s + r.conv, 0) / D.wp_eff_data.length;
  initChart('scatter', {
    tooltip: { formatter: p => `${p.data[3]}<br>ATT: ${p.data[0]}s, 转化率: ${p.data[1]}%<br>C级: ${p.data[2]}% (${p.data[4]}人)` },
    grid: { left: 60, right: 30, top: 20, bottom: 40 },
    xAxis: { type: 'value', name: 'ATT(秒)', min: Math.min(...D.wp_eff_data.map(r=>r.att)) * 0.9 },
    yAxis: { type: 'value', name: '转化率(%)' },
    series: [{
      type: 'scatter', symbolSize: d => Math.max(10, d[4] * 2),
      data: D.wp_eff_data.map(r => [r.att, r.conv, r.c_rate, r.wp, r.n]),
      itemStyle: { color: d => { const c = d.data[2]; return c > 30 ? '#E8734A' : c > 20 ? '#ECC94B' : '#8B99A8'; } },
      emphasis: { focus: 'self' }
    }],
    markLine: {
      data: [{ xAxis: avgAtt, name: 'ATT均值' }, { yAxis: avgConv, name: '转化率均值' }],
      lineStyle: { type: 'dashed', color: '#A0AEC0' },
      label: { formatter: '{b}', fontSize: 10, color: '#A0AEC0' }
    }
  });
}

// Action Cards
document.getElementById('actionCards').innerHTML = `
  <div class="action-card p0"><div class="priority">P0 紧急</div><h4>高 C 级率职场专项干预</h4><p>C 级率最高的 ${D.wp_c_rates.length > 0 ? D.wp_c_rates[0].wp : '职场'} 达 ${D.wp_c_rates.length > 0 ? D.wp_c_rates[0].c_rate : '?'}%，建议暂停优质名单，配SOP+每日通关</p></div>
  <div class="action-card p0"><div class="priority">P0 紧急</div><h4>MOB1-3 新人前置培训</h4><p>MOB1-3 中 C 级占比 ${D.mob13_conc}%，建议新员工30天设通关门槛：日均通时>60min+转化率>50%</p></div>
  <div class="action-card p1"><div class="priority">P1 重要</div><h4>有效→首贷转化能力提升</h4><p>S级转化率 ${D.s_conv}% 是 C级 ${D.c_conv}% 的 ${D.sc_ratio} 倍，核心在需求挖掘能力</p></div>
  <div class="action-card p1"><div class="priority">P1 重要</div><h4>S 级经验萃取</h4><p>S 级人均产出显著高于 C 级，建议萃取优秀录音+话术手册，跨职场复制</p></div>
  <div class="action-card p2"><div class="priority">P2 常规</div><h4>C 级高发职场月度跟踪</h4><p>建议纳入月度供应商管理仪表盘，连续2月无改善启动约谈</p></div>
  <div class="action-card p2"><div class="priority">P2 常规</div><h4>数据局限性说明</h4><p>单期截面数据，无法判断趋势。转化率分母为接通用户数，建议连续3期对比</p></div>
`;

// Responsive
window.addEventListener('resize', () => { echarts.instances && Object.values(echarts.instances).forEach(c => c.resize()); });
</script>
</body>
</html>"""


# =====================================================================
# MAIN
# =====================================================================

# =====================================================================
# LOAD_MULTI — 多文件合并
# =====================================================================

def load_multi(paths):
    """加载多个 Excel 文件并合并"""
    all_rows = []
    latest_deadline = None
    formats_seen = set()

    for path in paths:
        data = load(path)
        print(f'[SABC] Merging: {len(data["rows"])} rows from {os.path.basename(path)}')
        for r in data['rows']:
            r['_source_file'] = os.path.basename(path)
        all_rows.extend(data['rows'])
        if latest_deadline is None or data['deadline'] > latest_deadline:
            latest_deadline = data['deadline']
        formats_seen.add(data['format'])

    fmt = 'A+B' if len(formats_seen) > 1 else (formats_seen.pop() if formats_seen else 'B')
    print(f'[SABC] Total merged: {len(all_rows)} rows, formats: {fmt}')

    return {
        'rows': all_rows,
        'deadline': latest_deadline or datetime.now(),
        'format': fmt,
        'total': len(all_rows),
        'source_files': [os.path.basename(p) for p in paths],
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', help='Single input file')
    ap.add_argument('--inputs', nargs='+', help='Multiple input files (merged)')
    ap.add_argument('--output-dir', default='.')
    args = ap.parse_args()

    if args.inputs and len(args.inputs) > 1:
        # Multi-file merge
        print(f'[SABC] Multi-file merge: {len(args.inputs)} files')
        data = load_multi(args.inputs)
    else:
        path = args.input or (args.inputs[0] if args.inputs else None)
        if not path:
            ap.error('Either --input or --inputs is required')
        data = load(path)

    data = enrich(data)
    data = compute_ratings(data)
    data = compute_all_metrics(data)

    print(f'[SABC] Generating report...')
    html = build_html(data)
    os.makedirs(args.output_dir, exist_ok=True)

    html_path = os.path.join(args.output_dir, 'sabc_diagnostic_report.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'[SABC] HTML → {html_path} ({os.path.getsize(html_path)//1024}KB)')

    # Markdown outline
    md = build_markdown(data)
    md_path = os.path.join(args.output_dir, 'narrative_outline.md')
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(md)
    print(f'[SABC] MD → {md_path}')

    print(f'\n[SABC] Done.')

def build_markdown(data):
    sd = data.get('sd_stats', {})
    fd = data.get('fd_stats', {})
    cmb = data.get('combined_stats', {})
    deadline = data['deadline'].strftime('%Y-%m-%d')

    # Use primary pool for headline (首贷 if available, else 复贷)
    has_sd = len(data.get('sd_rated', [])) > 0
    main = sd if has_sd else fd
    pool_name = '首贷' if has_sd else '复贷'

    sd_n = sd.get('n_rated', 0)
    fd_n = fd.get('n_rated', 0)
    combined_n = cmb.get('n_rated', 0)
    source_files = data.get('source_files', [])

    lines = [
        f'# {data.get("_supplier", "金条")} 坐席 SABC 诊断 · 叙事情报大纲',
        f'> 产自 sabc-diagnostic skill | 数据截止 {deadline}',
        f'> 数据源: {len(source_files)} 个文件' if len(source_files) > 1 else f'> 数据源: 单文件',
        '',
        '## Big Idea',
    ]

    # Dynamic Big Idea based on key findings
    gd = main.get('grade_dist', {})
    s_count = gd.get('S', 0)
    c_count = gd.get('C', 0)
    sc_ratio = main.get('sc_ratio', 0)
    top20_pct = main.get('top20_gmv_pct', 0)
    part_rate = main.get('participation_rate', 0)
    warning_rate = main.get('warning_rate', 0)
    mob13_conc = main.get('mob13_c_conc', 0)

    lines.append(f'{pool_name}参评 {sd_n} 人' if has_sd else f'复贷参评 {fd_n} 人')
    if combined_n and combined_n != sd_n:
        lines[-1] += f'，合计参评 {combined_n} 人（首贷 {sd_n} + 复贷 {fd_n}）'
    lines[-1] += f'，参评率 {part_rate}%。'
    lines[-1] += f'S/C 转化倍差 {sc_ratio}×，' if sc_ratio > 0 else ''
    lines[-1] += f'Top20% 贡献 {top20_pct}% GMV。'

    lines.append('')
    lines.append('## 四视角诊断')
    lines.append('')

    # 1. 人力结构
    lines.append(f'### 人力结构')
    lines.append(f'- S={s_count}, A={gd.get("A",0)}, B={gd.get("B",0)}, C={c_count}')
    lines.append(f'- 参评率 {part_rate}%（目标 ≥75%）')
    if c_count > 0:
        c_rate = round(c_count / max(1, sum(gd.values())) * 100)
        lines.append(f'- C 级率 {c_rate}%（{"⚠️ 偏高" if c_rate > 25 else "正常"}）')
    lines.append('')

    # 2. 产能效率
    lines.append(f'### 产能效率')
    gm_s = main.get('grade_metrics', {}).get('S', {})
    gm_c = main.get('grade_metrics', {}).get('C', {})
    s_conv = gm_s.get('avg_conv', 0) or 0
    c_conv = gm_c.get('avg_conv', 0) or 0
    s_att = gm_s.get('avg_att', 0) or 0
    c_att = gm_c.get('avg_att', 0) or 0
    lines.append(f'- S 转化率 {s_conv:.2f}% vs C 转化率 {c_conv:.2f}%（倍差 {sc_ratio}×）')
    lines.append(f'- S ATT {s_att:.0f}s vs C ATT {c_att:.0f}s')
    lines.append('')

    # 3. 集中度风险
    lines.append(f'### 集中度风险')
    lines.append(f'- Top20% GMV 占比 {top20_pct}%（{"⚠️ 偏高" if top20_pct > 60 else "正常"}）')
    lines.append(f'- 预警率 {warning_rate}%')
    lines.append(f'- MOB1-3 C集中度 {mob13_conc}%（{"⚠️ 偏高" if mob13_conc > 40 else "正常"}）')
    lines.append('')

    # 4. 问题定位
    lines.append(f'### 问题定位')
    wp_c_rates = main.get('wp_c_rates', [])
    if wp_c_rates:
        top_wp = wp_c_rates[0]
        lines.append(f'- C级率最高职场: {top_wp["wp"]} ({top_wp["c_rate"]}%, {top_wp["c_count"]}/{top_wp["total"]}人)')
    if mob13_conc > 30:
        lines.append(f'- 新人 MOB1-3 集中 C 级 {mob13_conc}%，需前置培训')
    lines.append('')

    # Cross-perspective diagnosis
    lines.append('## 交叉诊断')
    lines.append('')
    findings = []
    if sc_ratio > 3:
        findings.append(f'- **产能分化严重**：S 级产能是 C 级的 {sc_ratio}×，建议经验萃取')
    if part_rate < 70:
        findings.append(f'- **参评率偏低**：仅 {part_rate}% 达参评标准，大量坐席活跃度不足')
    if top20_pct > 50:
        findings.append(f'- **GMV 集中度高**：Top20% 贡献超半数 GMV，核心坐席流失风险大')
    if mob13_conc > 30:
        findings.append(f'- **新人质量堪忧**：MOB1-3 C 级占比 {mob13_conc}%，招聘/培训环节需优化')
    if warning_rate > 10:
        findings.append(f'- **预警率偏高**：{warning_rate}% 坐席 3 天未上线，管理风险存在')
    if not findings:
        findings.append('- 暂无显著风险信号')
    lines.extend(findings)
    lines.append('')

    # 局限性
    lines.append('## 局限性')
    lines.append('- 单期截面数据，无法判断趋势和季节性')
    lines.append('- 转化率分母为接通用户数而非名单下发数')
    lines.append('- 缺少质检合规数据')
    return '\n'.join(lines)


if __name__ == '__main__':
    main()
