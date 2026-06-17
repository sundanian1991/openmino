from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# Color scheme
HEADER_FILL = PatternFill('solid', start_color='1F3864')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
SUBHEAD_FILL = PatternFill('solid', start_color='D6E4F0')
SUBHEAD_FONT = Font(name='Calibri', bold=True, color='1F3864', size=11)
INPUT_FONT = Font(name='Calibri', color='0000FF', size=11)
FORMULA_FONT = Font(name='Calibri', color='000000', size=11)
LINK_FONT = Font(name='Calibri', color='008000', size=11)
A_FILL = PatternFill('solid', start_color='C6EFCE')
B_FILL = PatternFill('solid', start_color='FFEB9C')
C_FILL = PatternFill('solid', start_color='FFC7CE')
LIGHT_FILL = PatternFill('solid', start_color='F2F2F2')
WHITE_FILL = PatternFill('solid', start_color='FFFFFF')
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1F3864')
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def style_data_cell(ws, row, col, font=None, fill=None, center=False):
    cell = ws.cell(row=row, column=col)
    cell.font = font or FORMULA_FONT
    if fill: cell.fill = fill
    cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
    cell.border = BORDER
    return cell

def add_table_borders(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER

# ════════════════════════════════════════════════════════
# Sheet 1: 供应商基础信息
# ════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '供应商基础信息'
ws1.sheet_properties.tabColor = '1F3864'

ws1.merge_cells('A1:F1')
ws1['A1'].value = '供应商基础信息'
ws1['A1'].font = TITLE_FONT
ws1['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[1].height = 30

headers1 = ['序号', '供应商名称', '供应商编码', '业务线', '合作起始日期', '当前等级']
col_widths1 = [8, 25, 18, 15, 18, 15]
for i, (h, w) in enumerate(zip(headers1, col_widths1), 1):
    ws1.cell(row=3, column=i, value=h)
    ws1.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws1, 3, len(headers1))

# Sample rows (33 suppliers)
sample_names = ['毅航', '毛毛虫', '伽玛', '赛维斯', '岐力', '翰锐', '供应商G', '供应商H', '供应商I']
business_lines = ['金条', '企金', '信用卡', '财富']
for i in range(9):
    r = 4 + i
    ws1.cell(row=r, column=1, value=i+1)
    ws1.cell(row=r, column=2, value=sample_names[i] if i < len(sample_names) else f'供应商{chr(65+i)}')
    ws1.cell(row=r, column=3, value=f'SUP-{str(i+1).zfill(3)}')
    ws1.cell(row=r, column=4, value=business_lines[i % 4])
    ws1.cell(row=r, column=5, value='2024-01-01')
    ws1.cell(row=r, column=6, value='B')
    for c in range(1, 7):
        cell = ws1.cell(row=r, column=c)
        cell.font = INPUT_FONT if c in [2,4,5,6] else FORMULA_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

for r in range(4, 36):
    ws1.row_dimensions[r].height = 22

# ════════════════════════════════════════════════════════
# Sheet 2: 业务督导评分（60分）
# ════════════════════════════════════════════════════════
ws2 = wb.create_sheet('业务督导评分(60分)')
ws2.sheet_properties.tabColor = '2E75B6'

ws2.merge_cells('A1:H1')
ws2['A1'].value = '业务督导评分表（定量 · 满分60分）'
ws2['A1'].font = TITLE_FONT
ws2.row_dimensions[1].height = 30

# Scoring rules reference
ws2.merge_cells('A3:H3')
ws2['A3'].value = '评分规则参考'
ws2['A3'].font = SUBHEAD_FONT
ws2['A3'].fill = SUBHEAD_FILL

rules = [
    ['产能达成率(25分)', '≥100%=25分, 90-99%=比例分, 80-89%=20分, 70-79%=15分, <70%=10分'],
    ['人均产出(15分)', '≥目标=15分, 90-99%=比例分, 80-89%=12分, 70-79%=9分, <70%=6分'],
    ['交付时效率(10分)', '≥95%=10分, 90-94%=8分, 85-89%=6分, 80-84%=4分, <80%=2分'],
    ['任务完成率(5分)', '100%=5分, 95-99%=4分, 90-94%=3分, 85-89%=2分, <85%=1分'],
    ['业务增长率(5分)', '≥10%=5分, 5-9%=4分, 0-4%=3分, 下滑0-5%=2分, 下滑>5%=1分'],
]
for i, (name, rule) in enumerate(rules):
    ws2.cell(row=4+i, column=1, value=name).font = Font(name='Calibri', bold=True, size=10)
    ws2.cell(row=4+i, column=2, value=rule).font = Font(name='Calibri', size=10)
    ws2.merge_cells(start_row=4+i, start_column=2, end_row=4+i, end_column=8)

# Data entry headers
data_row = 10
headers2 = ['供应商名称', '评估季度', '目标产能', '实际产能', '产能达成率', '产能得分(25分)',
            '人均产出', '人均得分(15分)', '按时工单数', '总工单数', '交付效率(%)', '时效得分(10分)',
            '完成任务数', '下发任务数', '完成率(%)', '任务得分(5分)',
            '上期产能', '增长率(%)', '增长得分(5分)', '业务督导总分']
col_widths2 = [18, 15, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 15]
for i, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    ws2.cell(row=data_row, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws2, data_row, len(headers2))

# Formulas for first supplier (row 11 = Excel row)
r = 11
# A: supplier name (link from sheet1)
ws2.cell(row=r, column=1).value = "='供应商基础信息'!B4"
# B: quarter (input)
ws2.cell(row=r, column=2).value = '2026-Q1'
ws2.cell(row=r, column=2).font = INPUT_FONT
# C: target capacity (input)
ws2.cell(row=r, column=3).font = INPUT_FONT
# D: actual capacity (input)
ws2.cell(row=r, column=4).font = INPUT_FONT
# E: capacity rate
ws2.cell(row=r, column=5).value = '=IF(C11=0,0,D11/C11)'
ws2.cell(row=r, column=5).number_format = '0.0%'
# F: capacity score using IFS
ws2.cell(row=r, column=6).value = '=IF(E11>=1,25,IF(E11>=0.9,E11*25,IF(E11>=0.8,20,IF(E11>=0.7,15,10))))'
# G: per capita output (input)
ws2.cell(row=r, column=7).font = INPUT_FONT
# H: per capita score
ws2.cell(row=r, column=8).value = '=IF(G11>=1,15,IF(G11>=0.9,G11*15,IF(G11>=0.8,12,IF(G11>=0.7,9,6))))'
# I: on-time orders (input)
ws2.cell(row=r, column=9).font = INPUT_FONT
# J: total orders (input)
ws2.cell(row=r, column=10).font = INPUT_FONT
# K: delivery rate
ws2.cell(row=r, column=11).value = '=IF(J11=0,0,I11/J11)'
ws2.cell(row=r, column=11).number_format = '0.0%'
# L: delivery score
ws2.cell(row=r, column=12).value = '=IF(K11>=0.95,10,IF(K11>=0.9,8,IF(K11>=0.85,6,IF(K11>=0.8,4,2))))'
# M: completed tasks (input)
ws2.cell(row=r, column=13).font = INPUT_FONT
# N: assigned tasks (input)
ws2.cell(row=r, column=14).font = INPUT_FONT
# O: completion rate
ws2.cell(row=r, column=15).value = '=IF(N11=0,0,M11/N11)'
ws2.cell(row=r, column=15).number_format = '0.0%'
# P: task score
ws2.cell(row=r, column=16).value = '=IF(O11>=1,5,IF(O11>=0.95,4,IF(O11>=0.9,3,IF(O11>=0.85,2,1))))'
# Q: last period capacity (input)
ws2.cell(row=r, column=17).font = INPUT_FONT
# R: growth rate
ws2.cell(row=r, column=18).value = '=IF(Q11=0,0,(D11-Q11)/Q11)'
ws2.cell(row=r, column=18).number_format = '0.0%'
# S: growth score
ws2.cell(row=r, column=19).value = '=IF(R11>=0.1,5,IF(R11>=0.05,4,IF(R11>=0,3,IF(R11>=-0.05,2,1))))'
# T: total score
ws2.cell(row=r, column=20).value = '=F11+H11+L11+P11+S11'
ws2.cell(row=r, column=20).font = Font(name='Calibri', bold=True, color='1F3864', size=12)

# Apply styles to row 11
for c in range(1, 21):
    cell = ws2.cell(row=r, column=c)
    if c not in [6,8,12,16,19,20,5,11,15,18]:
        cell.font = FORMULA_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER

# Extend formulas to 33 suppliers
for r in range(12, 37):
    src = r - 1
    for c in range(1, 21):
        old_cell = get_column_letter(c) + str(src)
        new_cell = get_column_letter(c) + str(r)
        old_formula = ws2.cell(row=src, column=c).value
        if old_formula and isinstance(old_formula, str) and old_formula.startswith('='):
            # Adjust row references
            import re
            def adj_row(match):
                ref = match.group(0)
                # Extract row number
                row_match = re.search(r'\d+', ref)
                if row_match:
                    old_row = int(row_match.group())
                    new_ref = ref.replace(str(old_row), str(old_row + 1))
                    return new_ref
                return ref
            ws2.cell(row=r, column=c).value = re.sub(r'[A-Z]+\d+', adj_row, old_formula)
        else:
            if c == 1:
                ws2.cell(row=r, column=c).value = f"='供应商基础信息'!B{r-7}"
        cell = ws2.cell(row=r, column=c)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        if c == 20:
            cell.font = Font(name='Calibri', bold=True, color='1F3864', size=12)

for r in range(11, 37):
    ws2.row_dimensions[r].height = 22

# ════════════════════════════════════════════════════════
# Sheet 3: 供应商管理评分（30分）
# ════════════════════════════════════════════════════════
ws3 = wb.create_sheet('供应商管理评分(30分)')
ws3.sheet_properties.tabColor = 'BF8F00'

ws3.merge_cells('A1:G1')
ws3['A1'].value = '供应商管理评分表（定性 · 满分30分）'
ws3['A1'].font = TITLE_FONT
ws3.row_dimensions[1].height = 30

# Scoring rules
ws3.merge_cells('A3:G3')
ws3['A3'].value = '评分规则参考'
ws3['A3'].font = SUBHEAD_FONT
ws3['A3'].fill = SUBHEAD_FILL

rules3 = [
    ['响应速度(8分)', '5档=8分, 4档=6分, 3档=4分, 2档=2分, 1档=0分'],
    ['配合度(8分)', '5档=8分, 4档=6分, 3档=4分, 2档=2分, 1档=0分'],
    ['团队稳定性(6分)', '流失率<5%=6分, 5-10%=4分, 10-15%=3分, >15%=1分'],
    ['整改执行力(5分)', '闭环率100%=5分, 90-99%=4分, 80-89%=3分, <80%=1分'],
    ['管理规范性(3分)', '无违规=3分, 轻微=2分, 一般=1分, 严重=0分'],
]
for i, (name, rule) in enumerate(rules3):
    ws3.cell(row=4+i, column=1, value=name).font = Font(name='Calibri', bold=True, size=10)
    ws3.cell(row=4+i, column=2, value=rule).font = Font(name='Calibri', size=10)
    ws3.merge_cells(start_row=4+i, start_column=2, end_row=4+i, end_column=7)

data_row3 = 10
headers3 = ['供应商名称', '评估季度', '响应速度(1-5档)', '响应得分(8分)',
            '配合度(1-5档)', '配合得分(8分)', '团队流失率(%)', '稳定性得分(6分)',
            '整改闭环率(%)', '整改得分(5分)', '管理规范性(1-5档)', '规范得分(3分)', '供应商管理总分']
col_widths3 = [18, 15, 15, 12, 15, 12, 15, 12, 15, 12, 18, 12, 15]
for i, (h, w) in enumerate(zip(headers3, col_widths3), 1):
    ws3.cell(row=data_row3, column=i, value=h)
    ws3.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws3, data_row3, len(headers3))

r = 11
ws3.cell(row=r, column=1).value = "='供应商基础信息'!B4"
ws3.cell(row=r, column=2).value = '2026-Q1'
ws3.cell(row=r, column=2).font = INPUT_FONT
# Response speed (1-5档 input)
ws3.cell(row=r, column=3).font = INPUT_FONT
ws3.cell(row=r, column=4).value = '=IF(C11>=5,8,IF(C11>=4,6,IF(C11>=3,4,IF(C11>=2,2,0))))'
# Cooperation (1-5档 input)
ws3.cell(row=r, column=5).font = INPUT_FONT
ws3.cell(row=r, column=6).value = '=IF(E11>=5,8,IF(E11>=4,6,IF(E11>=3,4,IF(E11>=2,2,0))))'
# Turnover rate (input)
ws3.cell(row=r, column=7).font = INPUT_FONT
ws3.cell(row=r, column=7).number_format = '0.0%'
ws3.cell(row=r, column=8).value = '=IF(G11<0.05,6,IF(G11<0.1,4,IF(G11<0.15,3,1)))'
# Rectification rate (input)
ws3.cell(row=r, column=9).font = INPUT_FONT
ws3.cell(row=r, column=9).number_format = '0.0%'
ws3.cell(row=r, column=10).value = '=IF(I11>=1,5,IF(I11>=0.9,4,IF(I11>=0.8,3,1)))'
# Management norm (1-5档 input)
ws3.cell(row=r, column=11).font = INPUT_FONT
ws3.cell(row=r, column=12).value = '=IF(K11>=5,3,IF(K11>=4,2,IF(K11>=3,1,0)))'
# Total
ws3.cell(row=r, column=13).value = '=D11+F11+H11+J11+L11'
ws3.cell(row=r, column=13).font = Font(name='Calibri', bold=True, color='1F3864', size=12)

for c in range(1, 14):
    cell = ws3.cell(row=r, column=c)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER

for r in range(12, 37):
    src = r - 1
    for c in range(1, 14):
        old_formula = ws3.cell(row=src, column=c).value
        if old_formula and isinstance(old_formula, str) and old_formula.startswith('='):
            import re
            def adj_row2(match):
                ref = match.group(0)
                row_match = re.search(r'\d+', ref)
                if row_match:
                    return ref.replace(str(int(row_match.group())), str(int(row_match.group()) + 1))
                return ref
            ws3.cell(row=r, column=c).value = re.sub(r'[A-Z]+\d+', adj_row2, old_formula)
        elif c == 1:
            ws3.cell(row=r, column=c).value = f"='供应商基础信息'!B{r-7}"
        cell = ws3.cell(row=r, column=c)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        if c == 13:
            cell.font = Font(name='Calibri', bold=True, color='1F3864', size=12)

for r in range(11, 37):
    ws3.row_dimensions[r].height = 22

# ════════════════════════════════════════════════════════
# Sheet 4: 质检评分（10分）
# ════════════════════════════════════════════════════════
ws4 = wb.create_sheet('质检评分(10分)')
ws4.sheet_properties.tabColor = '548235'

ws4.merge_cells('A1:E1')
ws4['A1'].value = '质检评分表（定量 · 满分10分）'
ws4['A1'].font = TITLE_FONT
ws4.row_dimensions[1].height = 30

ws4.merge_cells('A3:E3')
ws4['A3'].value = '评分规则参考'
ws4['A3'].font = SUBHEAD_FONT
ws4['A3'].fill = SUBHEAD_FILL

rules4 = [
    ['质检合格率(5分)', '≥98%=5分, 95-97%=4分, 90-94%=3分, 85-89%=2分, <85%=1分'],
    ['客户投诉率(3分)', '≤0.1%=3分, 0.1-0.3%=2分, 0.3-0.5%=1分, >0.5%=0分'],
    ['红线事件(2分)', '0次=2分, 1次=0分, ≥2次=0分+直接降为C级'],
]
for i, (name, rule) in enumerate(rules4):
    ws4.cell(row=4+i, column=1, value=name).font = Font(name='Calibri', bold=True, size=10)
    ws4.cell(row=4+i, column=2, value=rule).font = Font(name='Calibri', size=10)
    ws4.merge_cells(start_row=4+i, start_column=2, end_row=4+i, end_column=5)

data_row4 = 8
headers4 = ['供应商名称', '评估季度', '质检合格率(%)', '合格率得分(5分)',
            '投诉工单数', '总工单数', '投诉率(%)', '投诉得分(3分)',
            '红线事件次数', '红线得分(2分)', '质检总分']
col_widths4 = [18, 15, 15, 15, 12, 12, 12, 12, 15, 12, 12]
for i, (h, w) in enumerate(zip(headers4, col_widths4), 1):
    ws4.cell(row=data_row4, column=i, value=h)
    ws4.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws4, data_row4, len(headers4))

r = 9
ws4.cell(row=r, column=1).value = "='供应商基础信息'!B4"
ws4.cell(row=r, column=2).value = '2026-Q1'
ws4.cell(row=r, column=2).font = INPUT_FONT
# Quality rate (input)
ws4.cell(row=r, column=3).font = INPUT_FONT
ws4.cell(row=r, column=3).number_format = '0.0%'
# Quality score
ws4.cell(row=r, column=4).value = '=IF(C9>=0.98,5,IF(C9>=0.95,4,IF(C9>=0.9,3,IF(C9>=0.85,2,1))))'
# Complaint orders (input)
ws4.cell(row=r, column=5).font = INPUT_FONT
# Total orders (input)
ws4.cell(row=r, column=6).font = INPUT_FONT
# Complaint rate
ws4.cell(row=r, column=7).value = '=IF(F9=0,0,E9/F9)'
ws4.cell(row=r, column=7).number_format = '0.00%'
# Complaint score
ws4.cell(row=r, column=8).value = '=IF(G9<=0.001,3,IF(G9<=0.003,2,IF(G9<=0.005,1,0)))'
# Redline events (input)
ws4.cell(row=r, column=9).font = INPUT_FONT
# Redline score
ws4.cell(row=r, column=10).value = '=IF(I9=0,2,0)'
# Total
ws4.cell(row=r, column=11).value = '=D9+H9+J9'
ws4.cell(row=r, column=11).font = Font(name='Calibri', bold=True, color='1F3864', size=12)

for c in range(1, 12):
    cell = ws4.cell(row=r, column=c)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER

for r in range(10, 36):
    src = r - 1
    for c in range(1, 12):
        old_formula = ws4.cell(row=src, column=c).value
        if old_formula and isinstance(old_formula, str) and old_formula.startswith('='):
            import re
            def adj_row3(match):
                ref = match.group(0)
                row_match = re.search(r'\d+', ref)
                if row_match:
                    return ref.replace(str(int(row_match.group())), str(int(row_match.group()) + 1))
                return ref
            ws4.cell(row=r, column=c).value = re.sub(r'[A-Z]+\d+', adj_row3, old_formula)
        elif c == 1:
            ws4.cell(row=r, column=c).value = f"='供应商基础信息'!B{r-5}"
        cell = ws4.cell(row=r, column=c)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        if c == 11:
            cell.font = Font(name='Calibri', bold=True, color='1F3864', size=12)

for r in range(9, 36):
    ws4.row_dimensions[r].height = 22

# ════════════════════════════════════════════════════════
# Sheet 5: 综合评定
# ════════════════════════════════════════════════════════
ws5 = wb.create_sheet('综合评定')
ws5.sheet_properties.tabColor = 'C00000'

ws5.merge_cells('A1:L1')
ws5['A1'].value = '供应商ABC分层分级 — 综合评定表'
ws5['A1'].font = TITLE_FONT
ws5.row_dimensions[1].height = 30

ws5.merge_cells('A3:L3')
ws5['A3'].value = '分级规则：≥85分=A级，70-84分=B级，<70分=C级 | 排名后30%降一级 | 连续2季度C级启动淘汰'
ws5['A3'].font = Font(name='Calibri', bold=True, italic=True, color='C00000', size=10)
ws5['A3'].alignment = Alignment(wrap_text=True)

data_row5 = 5
headers5 = ['供应商名称', '业务督导得分(60)', '供应商管理得分(30)', '质检得分(10)',
            '综合得分(100)', '综合排名', '基准等级', '排名是否后30%',
            '最终等级', '连续C季度数', '是否触发PIP', '是否触发淘汰']
col_widths5 = [18, 18, 18, 12, 12, 10, 10, 14, 10, 14, 12, 12]
for i, (h, w) in enumerate(zip(headers5, col_widths5), 1):
    ws5.cell(row=data_row5, column=i, value=h)
    ws5.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws5, data_row5, len(headers5))

r = 6
ws5.cell(row=r, column=1).value = "='供应商基础信息'!B4"
# Link scores from other sheets
ws5.cell(row=r, column=2).value = "='业务督导评分(60分)'!T11"
ws5.cell(row=r, column=3).value = "='供应商管理评分(30分)'!M11"
ws5.cell(row=r, column=4).value = "='质检评分(10分)'!K11"
# Total score
ws5.cell(row=r, column=5).value = '=B6+C6+D6'
ws5.cell(row=r, column=5).font = Font(name='Calibri', bold=True, size=12, color='1F3864')
# Rank
ws5.cell(row=r, column=6).value = '=RANK(E6,$E$6:$E$38,0)'
# Base grade
ws5.cell(row=r, column=7).value = '=IF(E6>=85,"A",IF(E6>=70,"B","C"))'
# Is bottom 30%
total_suppliers = 33
bottom30_threshold = int(total_suppliers * 0.7)
ws5.cell(row=r, column=8).value = f'=IF(F6>{bottom30_threshold},"是","否")'
# Final grade (downgrade if bottom 30%)
ws5.cell(row=r, column=9).value = '=IF(H6="是",IF(G6="A","B",IF(G6="B","C","C")),G6)'
# Consecutive C quarters (input)
ws5.cell(row=r, column=10).font = INPUT_FONT
ws5.cell(row=r, column=10).value = 0
# Trigger PIP
ws5.cell(row=r, column=11).value = '=IF(I6="C","是","否")'
# Trigger elimination
ws5.cell(row=r, column=12).value = '=IF(OR(I6="C",J6>=2),"是","否")'

for c in range(1, 13):
    cell = ws5.cell(row=r, column=c)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER

for r in range(7, 39):
    src = r - 1
    for c in range(1, 13):
        old_formula = ws5.cell(row=src, column=c).value
        if old_formula and isinstance(old_formula, str) and old_formula.startswith('='):
            import re
            def adj_row4(match):
                ref = match.group(0)
                row_match = re.search(r'\d+', ref)
                if row_match:
                    old_r = int(row_match.group())
                    # Handle absolute references
                    if '$' in ref:
                        col_part = ref.split('$')[0]
                        return ref  # keep absolute
                    return ref.replace(str(old_r), str(old_r + 1))
                return ref
            ws5.cell(row=r, column=c).value = re.sub(r'[A-Z]+\$?\d+', adj_row4, old_formula)
        elif c == 1:
            ws5.cell(row=r, column=c).value = f"='供应商基础信息'!B{r-2}"
        cell = ws5.cell(row=r, column=c)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        if c == 5:
            cell.font = Font(name='Calibri', bold=True, size=12, color='1F3864')

# Conditional formatting via colors for final grade
for r in range(6, 39):
    cell = ws5.cell(row=r, column=9)
    # Will be colored after recalc, set up data validation note

# Add summary section
sum_row = 41
ws5.cell(row=sum_row, column=1, value='分级统计').font = SUBHEAD_FONT
ws5.cell(row=sum_row, column=1).fill = SUBHEAD_FILL

ws5.cell(row=sum_row+1, column=1, value='A级数量').font = Font(name='Calibri', bold=True, size=11)
ws5.cell(row=sum_row+1, column=2).value = '=COUNTIF(I6:I38,"A")'
ws5.cell(row=sum_row+1, column=3, value='占比').font = Font(name='Calibri', size=11)
ws5.cell(row=sum_row+1, column=4).value = '=B42/COUNTA(I6:I38)'
ws5.cell(row=sum_row+1, column=4).number_format = '0.0%'
ws5.cell(row=sum_row+1, column=4).fill = A_FILL

ws5.cell(row=sum_row+2, column=1, value='B级数量').font = Font(name='Calibri', bold=True, size=11)
ws5.cell(row=sum_row+2, column=2).value = '=COUNTIF(I6:I38,"B")'
ws5.cell(row=sum_row+2, column=3, value='占比').font = Font(name='Calibri', size=11)
ws5.cell(row=sum_row+2, column=4).value = '=B43/COUNTA(I6:I38)'
ws5.cell(row=sum_row+2, column=4).number_format = '0.0%'
ws5.cell(row=sum_row+2, column=4).fill = B_FILL

ws5.cell(row=sum_row+3, column=1, value='C级数量').font = Font(name='Calibri', bold=True, size=11)
ws5.cell(row=sum_row+3, column=2).value = '=COUNTIF(I6:I38,"C")'
ws5.cell(row=sum_row+3, column=3, value='占比').font = Font(name='Calibri', size=11)
ws5.cell(row=sum_row+3, column=4).value = '=B44/COUNTA(I6:I38)'
ws5.cell(row=sum_row+3, column=4).number_format = '0.0%'
ws5.cell(row=sum_row+3, column=4).fill = C_FILL

ws5.cell(row=sum_row+5, column=1, value='触发PIP数量').font = Font(name='Calibri', bold=True, size=11, color='C00000')
ws5.cell(row=sum_row+5, column=2).value = '=COUNTIF(K6:K38,"是")'

ws5.cell(row=sum_row+6, column=1, value='触发淘汰数量').font = Font(name='Calibri', bold=True, size=11, color='C00000')
ws5.cell(row=sum_row+6, column=2).value = '=COUNTIF(L6:L38,"是")'

for c in range(1, 5):
    for r in range(sum_row+1, sum_row+4):
        cell = ws5.cell(row=r, column=c)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

# ════════════════════════════════════════════════════════
# Freeze panes
# ════════════════════════════════════════════════════════
ws1.freeze_panes = 'A4'
ws2.freeze_panes = 'A11'
ws3.freeze_panes = 'A11'
ws4.freeze_panes = 'A9'
ws5.freeze_panes = 'A6'

# ════════════════════════════════════════════════════════
# Print settings
# ════════════════════════════════════════════════════════
for ws in [ws1, ws2, ws3, ws4, ws5]:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr = None

output_path = '/Users/sundanian/Documents/projects/ai-agents/my-agent/plans/2026-04-13-供应商分层分级方案/打分模板.xlsx'
wb.save(output_path)
print(f'Excel saved: {output_path}')
