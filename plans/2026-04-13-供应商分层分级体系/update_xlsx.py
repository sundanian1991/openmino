from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

path = '/Users/sundanian/Documents/projects/ai-agents/my-agent/plans/2026-04-13-供应商分层分级体系/供应商打分模板.xlsx'
wb = load_workbook(path)

header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='1F4E79')
data_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', bold=True, size=10)
alt_fill = PatternFill('solid', fgColor='F2F7FB')
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
green_fill = PatternFill('solid', fgColor='C6EFCE')
yellow_fill = PatternFill('solid', fgColor='FFEB9C')
orange_fill = PatternFill('solid', fgColor='F4B084')
red_fill = PatternFill('solid', fgColor='FFC7CE')

SUPPLIER_COUNT = 30
data_start = 2
data_end = data_start + SUPPLIER_COUNT - 1

# ========== 新增 Sheet: 月度跟踪 ==========
ws = wb.create_sheet('月度跟踪', 1)  # 插入到第2位
ws.freeze_panes = 'B2'

# 单行表头
headers = ['供应商名称', 'M1排名', 'M1预警', 'M2排名', 'M2预警', 'M3排名', 'M3预警', '季度累计预警', '季度分级']
for i, h in enumerate(headers):
    cell = ws.cell(row=1, column=i+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

set_col_widths = [24, 12, 14, 12, 14, 12, 14, 16, 14]
for i, w in enumerate(set_col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 数据行
for r in range(data_start, data_end + 1):  # row 2 to 31
    idx = r - 1
    ws.cell(row=r, column=1, value=f'供应商{idx}').font = data_font
    ws.cell(row=r, column=1).alignment = left_align
    ws.cell(row=r, column=1).border = thin_border

    # M1排名（手动输入）
    ws.cell(row=r, column=2, value='').font = data_font
    ws.cell(row=r, column=2).alignment = center_align
    ws.cell(row=r, column=2).border = thin_border

    # M1预警 = IF排名/总数>70%则"正常"，<=70%且>30%不显示，<=30%则"黄色预警"
    c = f'B{r}'
    total = str(SUPPLIER_COUNT)
    ws.cell(row=r, column=3).value = f'=IF({c}="","",IF({c}/{total}>0.7,"",IF({c}/{total}>0.3,"","黄色预警")))'
    ws.cell(row=r, column=3).font = data_font
    ws.cell(row=r, column=3).alignment = center_align
    ws.cell(row=r, column=3).border = thin_border

    # M2排名（手动输入）
    ws.cell(row=r, column=4, value='').font = data_font
    ws.cell(row=r, column=4).alignment = center_align
    ws.cell(row=r, column=4).border = thin_border

    # M2预警 = 连续2月后30%则橙色
    ws.cell(row=r, column=5).value = f'=IF(D{r}="","",IF(AND(D{r}/{total}<=0.3,C{r}="黄色预警"),"橙色预警",IF(D{r}/{total}<=0.3,"黄色预警","")))'
    ws.cell(row=r, column=5).font = data_font
    ws.cell(row=r, column=5).alignment = center_align
    ws.cell(row=r, column=5).border = thin_border

    # M3排名（手动输入）
    ws.cell(row=r, column=6, value='').font = data_font
    ws.cell(row=r, column=6).alignment = center_align
    ws.cell(row=r, column=6).border = thin_border

    # M3预警 = 连续3月后30%则红色PIP
    ws.cell(row=r, column=7).value = f'=IF(F{r}="","",IF(AND(F{r}/{total}<=0.3,E{r}="橙色预警"),"红色PIP",IF(AND(F{r}/{total}<=0.3,COUNTIF(C{r}:E{r},"黄色预警")>=1),"橙色预警",IF(F{r}/{total}<=0.3,"黄色预警",""))))'
    ws.cell(row=r, column=7).font = data_font
    ws.cell(row=r, column=7).alignment = center_align
    ws.cell(row=r, column=7).border = thin_border

    # 季度累计预警 = 取最严重的
    ws.cell(row=r, column=8).value = f'=IF(G{r}="红色PIP","红色PIP",IF(E{r}="橙色预警","橙色预警",IF(COUNTIF(C{r}:E{r},"黄色预警")>=1,"黄色预警","正常")))'
    ws.cell(row=r, column=8).font = data_font
    ws.cell(row=r, column=8).alignment = center_align
    ws.cell(row=r, column=8).border = thin_border

    # 季度分级 = 引用打分总表
    ws.cell(row=r, column=9).value = f"='打分总表'!H{r}"
    ws.cell(row=r, column=9).font = data_font
    ws.cell(row=r, column=9).alignment = center_align
    ws.cell(row=r, column=9).border = thin_border

    # 隔行变色
    if idx % 2 == 0:
        for col in range(1, 10):
            ws.cell(row=r, column=col).fill = alt_fill

# 条件格式：预警级别
for col_letter in ['C', 'E', 'G', 'H']:
    rng = f'{col_letter}3:{col_letter}{data_end + 1}'
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"黄色预警"'], fill=yellow_fill, font=Font(color='9C6500')))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"橙色预警"'], fill=orange_fill, font=Font(color='974706')))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"红色PIP"'], fill=red_fill, font=Font(color='9C0006', bold=True)))

# 分级条件格式
ws.conditional_formatting.add(f'I3:I{data_end+1}', CellIsRule(operator='equal', formula=['"A"'], fill=green_fill, font=Font(color='006100', bold=True)))
ws.conditional_formatting.add(f'I3:I{data_end+1}', CellIsRule(operator='equal', formula=['"B"'], fill=yellow_fill, font=Font(color='9C6500')))
ws.conditional_formatting.add(f'I3:I{data_end+1}', CellIsRule(operator='equal', formula=['"C"'], fill=red_fill, font=Font(color='9C0006', bold=True)))

# 参考说明（数据区下方）
ref_start = data_end + 4
ws.cell(row=ref_start, column=1, value='月度介入管理说明').font = Font(name='微软雅黑', bold=True, size=11, color='1F4E79')

notes = [
    ['预警级别', '触发条件', '管理动作', '交付物'],
    ['黄色预警', '单月赛马排名后30%', '发数据诊断报告+电话沟通+口头提醒', '预警通知（微信/邮件）'],
    ['橙色预警', '连续2月后30%', '正式约谈+要求改善计划+双周跟踪+暂停新项目', '约谈纪要+改善计划'],
    ['红色PIP', '连续3月后30%（=季度C级）', '按PIP流程执行（30天）', 'PIP全套文档'],
    ['退出', '连续4月后30%（PIP不通过）', '启动退出流程', '退出通知书'],
]

for ri, row_data in enumerate(notes):
    for ci, val in enumerate(row_data):
        cell = ws.cell(row=ref_start + 1 + ri, column=ci + 1, value=val)
        cell.font = bold_font if ri == 0 else data_font
        cell.fill = header_fill if ri == 0 else (alt_fill if ri % 2 == 0 else PatternFill())
        cell.alignment = left_align
        cell.border = thin_border
        if ri == 0:
            cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)

ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

wb.save(path)
print(f'Updated: {path}')
