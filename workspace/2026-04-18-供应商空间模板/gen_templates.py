#!/usr/bin/env python3
"""生成供应商空间 Excel 模板包"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE = "/Users/sundanian/Documents/projects/ai-agents/my-agent/workspace/2026-04-18-供应商空间模板/[供应商名称]-工作空间"

# 样式常量
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E8B6E")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="2E8B6E")
DESC_FONT = Font(name="微软雅黑", size=10, color="666666")

def style_header(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

def style_data_row(ws, max_col, row):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_dv(ws, cells, items):
    formula = f'"{",".join(items)}"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = f"请从以下选项中选择: {', '.join(items)}"
    dv.errorTitle = "输入无效"
    dv.prompt = f"选项: {', '.join(items)}"
    dv.promptTitle = "请选择"
    ws.add_data_validation(dv)
    dv.add(cells)

def add_title_desc(ws, title, desc, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1).value = desc
    ws.cell(row=2, column=1).font = DESC_FONT
    return 3  # header starts at row 3

# ============================================================
# 模板0：供应商基础信息表（准入时填写，后续变更时更新）
# ============================================================
def gen_basic_info_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "基础信息"
    ws.sheet_properties.tabColor = "4A90D9"

    header_row = add_title_desc(ws, "供应商基础信息表", "填报频率：准入时填写 / 人员变更时24小时内更新", 12)

    headers = ["供应商名称", "统一社会信用代码", "法人代表", "联系人姓名", "联系人职务",
               "联系电话", "微信号", "邮箱", "合作起始日期", "服务业务线", "团队规模", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, len(headers))

    examples = [
        ["XX科技有限公司", "91110000XXXXXXXX", "张三", "李四", "项目对接人", "13800000000", "lisi_wx", "lisi@example.com", "2025-06-01", "金条-电销", 190, ""],
        ["XX科技有限公司", "91110000XXXXXXXX", "张三", "王五", "项目主管", "13900000000", "wwang", "wang@example.com", "2025-06-01", "金条-电销", 190, "2026-04-15更换主管"],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    for r_idx, row_data in enumerate(examples, header_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
        style_data_row(ws, len(headers), r_idx)

    add_dv(ws, f"J3:J100", ["金条-电销", "金条-审批", "企金-电销", "信用卡-电销", "财富-电销", "多业务线"])

    set_col_widths(ws, [18, 22, 10, 10, 12, 14, 12, 20, 12, 14, 10, 20])
    ws.freeze_panes = "A4"

    path = os.path.join(BASE, "01-人力管理", "供应商基础信息表.xlsx")
    wb.save(path)
    print(f"OK: {path}")

# ============================================================
# 模板1：人力明细表
# ============================================================
def gen_hr_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "人力明细"
    ws.sheet_properties.tabColor = "2E8B6E"

    header_row = add_title_desc(ws, "人力明细表", "填报频率：每周一 12:00 前 | 填写上周数据", 12)

    headers = ["填报日期", "填报人", "周期起始", "周期结束", "总人数", "熟手人数(入职>30天)",
               "新人人数(入职≤30天)", "本周新入职", "本周离职", "日均有效出勤率", "日均有效出勤人数", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, len(headers))

    # 3行示例数据
    examples = [
        ["2026-04-14", "张三", "2026-04-07", "2026-04-13", 190, 140, 50, 8, 5, "92%", 175, ""],
        ["2026-04-21", "张三", "2026-04-14", "2026-04-20", 193, 142, 51, 6, 3, "93%", 178, ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    for r_idx, row_data in enumerate(examples, header_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
        style_data_row(ws, len(headers), r_idx)

    # 出勤率列设为百分比格式
    for r in range(header_row + 1, header_row + 4):
        ws.cell(row=r, column=10).number_format = '0%'

    set_col_widths(ws, [12, 10, 12, 12, 10, 16, 16, 12, 10, 16, 16, 25])
    ws.freeze_panes = "A4"

    path = os.path.join(BASE, "01-人力管理", "人力明细表.xlsx")
    wb.save(path)
    print(f"OK: {path}")

# ============================================================
# 模板2：主管信息表
# ============================================================
def gen_manager_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "主管信息"
    ws.sheet_properties.tabColor = "3A7AB8"

    header_row = add_title_desc(ws, "主管信息表", "填报频率：每月5日前确认 / 变更时24小时内更新", 12)

    headers = ["供应商名称", "更新日期", "主管姓名", "联系电话", "微信号", "邮箱",
               "管理团队规模", "在职时长(月)", "本次是否变更", "变更类型", "变更原因", "生效日期"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, len(headers))

    examples = [
        ["XX科技有限公司", "2026-04-01", "李四", "13800000000", "", "", 190, 18, "否", "", "", ""],
        ["XX科技有限公司", "2026-04-15", "王五", "13900000000", "wwang", "wang@example.com", 190, 2, "是", "更换主管", "原主管离职", "2026-04-15"],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    for r_idx, row_data in enumerate(examples, header_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
        style_data_row(ws, len(headers), r_idx)

    add_dv(ws, f"I3:I100", ["是", "否"])
    add_dv(ws, f"J3:J100", ["", "更换主管", "联系方式变更"])

    set_col_widths(ws, [18, 12, 10, 14, 12, 20, 12, 12, 12, 14, 20, 12])
    ws.freeze_panes = "A4"

    path = os.path.join(BASE, "01-人力管理", "主管信息表.xlsx")
    wb.save(path)
    print(f"OK: {path}")

# ============================================================
# 模板3：离职原因登记表
# ============================================================
def gen_resignation_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "离职登记"
    ws.sheet_properties.tabColor = "C25030"

    header_row = add_title_desc(ws, "离职原因登记表", "填报频率：事件触发，员工离职当日登记", 10)

    headers = ["离职日期", "姓名", "岗位", "入职日期", "在职天数", "离职类型",
               "离职原因(一级)", "离职原因(二级)", "主管是否面谈", "面谈记录摘要"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, len(headers))

    examples = [
        ["2026-04-10", "王五", "坐席", "2025-12-01", '=IF(D4<>"",A4-D4,"")', "主动离职", "薪酬不满", "觉得绩效提成太低", "是", "表示同行提成比例更高"],
        ["2026-04-11", "赵六", "坐席", "2026-03-15", '=IF(D5<>"",A5-D5,"")', "主动离职", "工作强度", "每天拨打量压力大", "是", "适应不了节奏"],
        ["", "", "", "", "", "", "", "", "", ""],
    ]
    for r_idx, row_data in enumerate(examples, header_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
        style_data_row(ws, len(headers), r_idx)

    add_dv(ws, f"F3:F100", ["主动离职", "被动辞退", "合同到期不续签", "其他"])
    add_dv(ws, f"G3:G100", ["薪酬不满", "工作强度", "管理问题", "个人原因", "找到更好工作", "其他"])
    add_dv(ws, f"I3:I100", ["是", "否"])

    # 在职天数列设为数字格式
    for r in range(header_row + 1, header_row + 4):
        ws.cell(row=r, column=5).number_format = '0'

    set_col_widths(ws, [12, 8, 8, 12, 10, 14, 14, 20, 12, 25])
    ws.freeze_panes = "A4"

    path = os.path.join(BASE, "01-人力管理", "离职管理", "离职原因登记表.xlsx")
    wb.save(path)
    print(f"OK: {path}")

# ============================================================
# 模板4：质检台账（日常记录，有问题就记）
# ============================================================
def gen_quality_ledger_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "质检台账"
    ws.sheet_properties.tabColor = "B88018"

    header_row = add_title_desc(ws, "质检台账", "填报频率：日常记录，发现问题当日登记", 12)

    headers = ["登记日期", "业务线", "坐席姓名", "工号", "通话/工单编号",
               "问题类型", "问题描述", "严重程度", "处理结果", "整改措施", "质检来源", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, len(headers))

    examples = [
        ["2026-04-10", "金条-电销", "赵六", "ZL001", "TN20260410-001", "A类错误", "未核实客户身份即提供账户信息", "A类", "已警告+重新培训", "安排一对一辅导", "质检抽检", ""],
        ["2026-04-11", "金条-电销", "孙七", "SQ002", "TN20260411-005", "B类错误", "话术不完整，缺少风险提示", "B类", "口头提醒", "晨会统一强调", "质检抽检", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    for r_idx, row_data in enumerate(examples, header_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
        style_data_row(ws, len(headers), r_idx)

    add_dv(ws, f"F3:F100", ["A类错误", "B类错误", "C类错误", "服务态度", "流程违规", "其他"])
    add_dv(ws, f"H3:H100", ["A类", "B类", "C类", "提醒"])
    add_dv(ws, f"K3:K100", ["质检抽检", "客户投诉", "内部自查", "甲方反馈"])

    set_col_widths(ws, [12, 12, 8, 8, 18, 12, 30, 10, 20, 20, 10, 15])
    ws.freeze_panes = "A4"

    path = os.path.join(BASE, "03-质检合规", "质检台账.xlsx")
    wb.save(path)
    print(f"OK: {path}")

# ============================================================
# 模板5：合规自查表
# ============================================================
def gen_compliance_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "合规自查"
    ws.sheet_properties.tabColor = "B88018"

    header_row = add_title_desc(ws, "合规自查表", "填报频率：月度，每月5日前填报上月自查结果", 10)

    headers = ["供应商名称", "自查月份", "填报日期", "检查项编号", "检查项名称", "检查内容",
               "自查结果", "不符合说明", "整改计划", "整改完成时间"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, len(headers))

    # 预设检查项
    checks = [
        ("A1", "信息安全", "外网管控是否全覆盖"),
        ("A2", "信息安全", "手机管控是否执行到位"),
        ("A3", "信息安全", "监控设备是否正常运行"),
        ("B1", "资质管理", "经营许可证是否在有效期内"),
        ("B2", "资质管理", "员工持证率是否达标"),
        ("C1", "质检管理", "当月是否收到A类错误通知"),
        ("C2", "质检管理", "A类错误是否已完成整改"),
    ]
    for r_idx, (code, name, content) in enumerate(checks, header_row + 1):
        ws.cell(row=r_idx, column=1).value = ""
        ws.cell(row=r_idx, column=2).value = ""
        ws.cell(row=r_idx, column=3).value = ""
        ws.cell(row=r_idx, column=4).value = code
        ws.cell(row=r_idx, column=5).value = name
        ws.cell(row=r_idx, column=6).value = content
        ws.cell(row=r_idx, column=7).value = ""
        style_data_row(ws, len(headers), r_idx)

    # 空白行
    for r_idx in range(header_row + 1 + len(checks), header_row + 1 + len(checks) + 3):
        style_data_row(ws, len(headers), r_idx)

    add_dv(ws, f"G{header_row+1}:G100", ["是", "否", "不适用"])

    set_col_widths(ws, [18, 10, 12, 12, 12, 25, 10, 25, 25, 12])
    ws.freeze_panes = "A4"

    path = os.path.join(BASE, "03-质检合规", "合规自查表.xlsx")
    wb.save(path)
    print(f"OK: {path}")

# ============================================================
# 模板6：整改计划表
# ============================================================
def gen_rectification_plan_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "整改计划"
    ws.sheet_properties.tabColor = "3A7AB8"

    header_row = add_title_desc(ws, "整改计划表", "填报频率：事件触发，约谈后48小时内提交", 11)

    headers = ["供应商名称", "约谈日期", "提交日期", "触发原因", "核心问题描述",
               "整改项编号", "整改目标", "具体动作", "责任人", "完成时间", "验收标准"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, len(headers))

    examples = [
        ["XX科技有限公司", "2026-04-30", "2026-05-03", "连续2月后30%", "人均拨打量从80降至58通/天，管理松弛",
         "1", "人均拨打量恢复到75通/天以上", "1.恢复晨会制度\n2.每日目标拆解到人\n3.下午3点复盘", "李四", "2026-05-15", "连续5天人均≥75通"],
        ["XX科技有限公司", "2026-04-30", "2026-05-03", "连续2月后30%", "人均拨打量从80降至58通/天，管理松弛",
         "2", "流失率降至8%/月以下", "1.主管1v1沟通每位员工\n2.了解离职倾向", "李四", "2026-05-20", "5月流失率<8%"],
        ["", "", "", "", "", "", "", "", "", "", ""],
    ]
    for r_idx, row_data in enumerate(examples, header_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
        style_data_row(ws, len(headers), r_idx)
        # 自动换行
        ws.cell(row=r_idx, column=8).alignment = Alignment(vertical="center", wrap_text=True)

    add_dv(ws, f"D3:D100", ["连续2月后30%", "A级骤降", "巡检红色项", "出勤持续异常", "业绩断崖", "其他"])

    set_col_widths(ws, [18, 12, 12, 16, 30, 10, 22, 35, 10, 12, 20])
    ws.freeze_panes = "A4"
    ws.row_dimensions[4].height = 50
    ws.row_dimensions[5].height = 50

    path = os.path.join(BASE, "05-整改跟踪", "整改计划表.xlsx")
    wb.save(path)
    print(f"OK: {path}")

# ============================================================
# 模板7：整改进度跟踪表
# ============================================================
def gen_rectification_progress_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "整改进度"
    ws.sheet_properties.tabColor = "2E8B6E"

    header_row = add_title_desc(ws, "整改进度跟踪表", "填报频率：周报，每周五 18:00 前更新", 8)

    headers = ["填报日期", "整改项编号", "整改目标", "本周进度(%)", "本周具体完成情况", "是否按期", "阻碍/风险", "预计完成日期"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, len(headers))

    examples = [
        ["2026-05-09", "1", "人均拨打量恢复到75通/天", 60, "晨会已恢复，人均达70通/天", "是", "", "2026-05-15"],
        ["2026-05-09", "2", "流失率降至8%/月以下", 30, "已沟通50%员工，2名有离职倾向", "是", "2名员工可能离职", "2026-05-20"],
        ["", "", "", "", "", "", "", ""],
    ]
    for r_idx, row_data in enumerate(examples, header_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
        style_data_row(ws, len(headers), r_idx)

    add_dv(ws, f"F3:F100", ["是", "否"])

    # 进度列格式
    for r in range(header_row + 1, header_row + 4):
        ws.cell(row=r, column=4).number_format = '0"%"'

    set_col_widths(ws, [12, 12, 25, 12, 35, 10, 20, 14])
    ws.freeze_panes = "A4"

    path = os.path.join(BASE, "05-整改跟踪", "整改进度跟踪.xlsx")
    wb.save(path)
    print(f"OK: {path}")

# ============================================================
# 主函数
# ============================================================
if __name__ == "__main__":
    gen_basic_info_sheet()
    gen_hr_sheet()
    gen_manager_sheet()
    gen_resignation_sheet()
    gen_quality_ledger_sheet()
    gen_compliance_sheet()
    gen_rectification_plan_sheet()
    gen_rectification_progress_sheet()
    print("\n全部 7 个模板生成完毕!")
