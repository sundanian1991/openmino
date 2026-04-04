from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# Sheet 1: 排名汇总
ws1 = wb.active
ws1.title = "排名汇总"

# 表头
headers = ["排名", "供应商", "现有人数", "调整人数", "调整后人数", "调整比例", "1月得分", "2月得分", "合计得分"]
ws1.append(headers)

# 样式
header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF')
center_align = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col in range(1, len(headers) + 1):
    cell = ws1.cell(1, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# 数据
data = [
    [1, "中乾", 25, 4, 21, "16.0%", 93.80, 105.65, 100.91],
    [2, "汇讯", 11, 3, 8, "27.3%", 109.24, 95.49, 100.99],
    [3, "锦瑞", 25, 7, 18, "28.0%", 102.42, 92.91, 96.71],
    [4, "德辰", 31, 19, 12, "61.3%", 93.19, 90.31, 91.46],
    [5, "小蜜蜂", 2, 2, 0, "100.0%", 95.32, 82.89, 87.86],
    ["-", "衍生", 0, 0, 0, "-", 89.43, 45.80, 61.61],
    ["-", "广达", 0, 0, 0, "-", "-", "-", "-"]
]

# 合计行
total_people = sum(row[2] for row in data if isinstance(row[2], int))
total_adjust = sum(row[3] for row in data if isinstance(row[3], int))

row_num = 2
for row in data:
    ws1.append(row)
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row_num, col)
        cell.alignment = center_align
        cell.border = border
        if col == 6 and isinstance(cell.value, str) and '%' in cell.value:
            cell.number_format = '0.0%'
        elif col in [7, 8, 9] and isinstance(cell.value, (int, float)):
            cell.number_format = '0.00'
    row_num += 1

# 合计行
ws1.append(["", "合计", total_people, total_adjust, total_people - total_adjust,
           f"{total_adjust/total_people:.1%}", "", "", ""])
for col in range(1, len(headers) + 1):
    cell = ws1.cell(row_num, col)
    cell.alignment = center_align
    cell.border = border
    if col == 2:
        cell.font = Font(bold=True)
    if col in [3, 4, 5]:
        cell.font = Font(bold=True)

# 列宽
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 12
ws1.column_dimensions['H'].width = 12
ws1.column_dimensions['I'].width = 12

# 冻结首行
ws1.freeze_panes = 'A2'

# Sheet 2: 原始数据
ws2 = wb.create_sheet("原始数据")

headers2 = ["供应商", "1月GMV达成率", "2月GMV达成率", "1月权重", "2月权重", "1月加权得分", "2月加权得分", "合计得分"]
ws2.append(headers2)

for col in range(1, len(headers2) + 1):
    cell = ws2.cell(1, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

data2 = [
    ["中乾", 93.80, 105.65, 40, 60, "=B2*D2/100", "=C2*E2/100", "=F2+G2"],
    ["汇讯", 109.24, 95.49, 40, 60, "=B3*D3/100", "=C3*E3/100", "=F3+G3"],
    ["锦瑞", 102.42, 92.91, 40, 60, "=B4*D4/100", "=C4*E4/100", "=F4+G4"],
    ["德辰", 93.19, 90.31, 40, 60, "=B5*D5/100", "=C5*E5/100", "=F5+G5"],
    ["小蜜蜂", 95.32, 82.89, 40, 60, "=B6*D6/100", "=C6*E6/100", "=F6+G6"],
    ["衍生", 89.43, 45.80, 40, 60, "=B7*D7/100", "=C7*E7/100", "=F7+G7"]
]

row_num = 2
for row in data2:
    ws2.append(row)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row_num, col)
        cell.alignment = center_align
        if col in [2, 3, 6, 7, 8]:
            cell.number_format = '0.00'
    row_num += 1

ws2.column_dimensions['A'].width = 12
for col in range(2, 9):
    ws2.column_dimensions[get_column_letter(col)].width = 14

# Sheet 3: 计算说明
ws3 = wb.create_sheet("计算说明")

ws3['A1'] = "供应商排名计算说明"
ws3['A1'].font = Font(bold=True, size=14)

ws3['A3'] = "一、计算公式"
ws3['A3'].font = Font(bold=True)

ws3['A4'] = "合计得分 = 1月GMV达成率 × 40% + 2月GMV达成率 × 60%"

ws3['A6'] = "二、排名规则"
ws3['A6'].font = Font(bold=True)

ws3['A7'] = "按合计得分从高到低排名"
ws3['A8'] = "得分相同，按2月表现优先"

ws3['A10'] = "三、调整原则"
ws3['A10'].font = Font(bold=True)

ws3['A11'] = "保优汰劣：按排名，尾部优先减"
ws3['A12'] = "断崖缓冲：头部（排名第1-2）减量≤30%"
ws3['A13'] = "腰部（排名第3）适度减量"
ws3['A14'] = "尾部（排名第4-5）大幅减量"

ws3['A16'] = "四、数据来源"
ws3['A16'].font = Font(bold=True)

ws3['A17'] = "业务系统统计"
ws3['A18'] = f"导出时间：{datetime.now().strftime('%Y-%m-%d')}"

ws3.column_dimensions['A'].width = 50

# 保存文件
output_path = '/Users/sundanian/Documents/projects/ai-agents/my-agent/docs/linshi/数据资料/供应商排名分析_1-2月.xlsx'
wb.save(output_path)
print(f"Excel文件已保存：{output_path}")
