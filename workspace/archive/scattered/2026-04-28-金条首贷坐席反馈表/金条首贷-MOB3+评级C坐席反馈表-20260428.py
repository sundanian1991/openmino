from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "MOB3+评级C坐席反馈"

names = [
    "刘婵", "张旭", "郭利民", "李雅晴", "蒋纯池", "李帅", "王秀兰", "尹咪咪",
    "潘文静", "李运成", "刘静波", "马淑艳", "卜晓佩", "李少青", "李永娟", "白力粉",
    "杜海欧", "南云辉", "沈成奇", "付冬梅", "王亚婷", "高玉娇", "杨正", "赵静",
    "张俊杰", "陈宇豪", "孙丽红", "季芳", "于思楠", "高鹏", "豆海晓", "丁雯倩",
]

headers = [
    ("序号", 6),
    ("职场", 14),
    ("坐席姓名", 12),
    ("MOB月数", 10),
    ("当前业绩\n(月均件数)", 14),
    ("业绩排名\n(职场内)", 12),
    ("核心短板", 16),
    ("短板数据支撑", 28),
    ("改进措施", 28),
    ("培训计划\n(课程/频次)", 22),
    ("辅导方式\n(一对一/旁听/小组/频次)", 24),
    ("目标值\n(下月)", 14),
    ("跟踪人", 10),
    ("备注", 18),
]

header_font = Font(bold=True, size=11)
header_fill = PatternFill("solid", fgColor="B4C6E7")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col_idx, (name, width) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 36

workplaces = ["毅航合肥职场", "赛维斯职场", "伽玛职场", "岐力职场", "毛毛虫职场", "翰锐曲靖职场", "广达陕西职场"]
dv_workplace = DataValidation(type="list", formula1='"' + ",".join(workplaces) + '"', allow_blank=True)
ws.add_data_validation(dv_workplace)

weaknesses = ["拨打量不足", "接通率低", "转化率差", "话术弱", "质检扣分多"]
dv_weakness = DataValidation(type="list", formula1='"' + ",".join(weaknesses) + '"', allow_blank=True)
ws.add_data_validation(dv_weakness)

example_fill = PatternFill("solid", fgColor="FFF2CC")
example_font = Font(color="666666", italic=True)
example = [
    1, "赛维斯职场", "张三", 4, 8, 15,
    "转化率差",
    "接通率42%正常，但转化率仅3.2%，低于团队均值5.8%，主要在异议处理环节流失",
    "1. 每日旁听优秀坐席2通成交录音\n2. 针对'利率高'异议专项演练\n3. 促成话术强化（二选一法）",
    "异议处理专项课（本周）\n促成技巧课（下周）\n每周2次，持续3周",
    "主管一对一辅导，每周2次，每次30分钟，含录音复盘",
    "转化率提升至5%，月均件数12件",
    "李四（主管）",
    "已安排4月29日开始旁听"
]

data_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")

for col_idx, val in enumerate(example, 1):
    cell = ws.cell(row=2, column=col_idx, value=val)
    cell.fill = example_fill
    cell.font = example_font
    cell.alignment = data_align
    cell.border = thin_border

for i, name in enumerate(names):
    row = i + 3
    ws.cell(row=row, column=1, value=i + 1).alignment = center_align
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = data_align
    dv_workplace.add(ws.cell(row=row, column=2))
    cell_name = ws.cell(row=row, column=3, value=name)
    cell_name.alignment = center_align
    cell_name.border = thin_border
    dv_weakness.add(ws.cell(row=row, column=7))
    for col in range(4, 15):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = data_align

ws.freeze_panes = "A3"

output = "/Users/sundanian/Documents/projects/ai-agents/my-agent/workspace/金条首贷-MOB3+评级C坐席反馈表-20260428.xlsx"
wb.save(output)
print(f"已生成: {output}")
