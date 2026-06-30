from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# Styles
header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='1F4E79')
data_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', bold=True, size=10)
alt_fill = PatternFill('solid', fgColor='F2F7FB')
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

green_fill = PatternFill('solid', fgColor='C6EFCE')
yellow_fill = PatternFill('solid', fgColor='FFEB9C')
red_fill = PatternFill('solid', fgColor='FFC7CE')

def style_header(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_range(ws, start_row, end_row, max_col):
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = center_align if col > 1 else left_align
            cell.border = thin_border
            if (row - start_row) % 2 == 1:
                cell.fill = alt_fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

SUPPLIER_COUNT = 30
data_start = 2
data_end = data_start + SUPPLIER_COUNT - 1

# ========== Sheet 1: 打分总表 ==========
ws1 = wb.active
ws1.title = '打分总表'
ws1.freeze_panes = 'B2'
set_col_widths(ws1, [4, 24, 16, 16, 14, 12, 12, 12, 30])

headers1 = ['序号', '供应商名称', '业务督导(60分)', '供应商管理(30分)', '质检(10分)', '总分(100分)', '排名', '分级', '管理动作']
for col, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=col, value=h)
style_header(ws1, len(headers1))

for r in range(data_start, data_end + 1):
    idx = r - data_start + 1
    ws1.cell(row=r, column=1, value=idx)
    ws1.cell(row=r, column=2, value=f'供应商{idx}')
    # B/C/D reference sheets 2/3/4
    ws1.cell(row=r, column=3).value = f"='业务督导打分'!{get_column_letter(9)}{r}"
    ws1.cell(row=r, column=4).value = f"='供应商管理打分'!{get_column_letter(9)}{r}"
    ws1.cell(row=r, column=5).value = f"='质检打分'!{get_column_letter(7)}{r}"
    ws1.cell(row=r, column=6).value = f'=SUM(C{r}:E{r})'
    ws1.cell(row=r, column=7).value = f'=RANK(F{r},F{data_start}:F{data_end},0)'
    ws1.cell(row=r, column=8).value = f'=IF(G{r}/COUNTA(B2:B31)<=0.3,"A",IF(G{r}/COUNTA(B2:B31)<=0.7,"B","C"))'
    ws1.cell(row=r, column=9).value = '=IF(H2="A","分量优先|月度回顾|战略对话",IF(H2="B","分量稳定|双周回顾|能力提升","分量压缩|周度跟踪|强制PIP"))'

style_data_range(ws1, data_start, data_end, len(headers1))

# Conditional formatting for 总分
ws1.conditional_formatting.add(
    f'F{data_start}:F{data_end}',
    FormulaRule(formula=[f'F{data_start}>=80'], fill=green_fill, font=Font(color='006100'))
)
ws1.conditional_formatting.add(
    f'F{data_start}:F{data_end}',
    FormulaRule(formula=[f'AND(F{data_start}>=60,F{data_start}<80)'], fill=yellow_fill, font=Font(color='9C6500'))
)
ws1.conditional_formatting.add(
    f'F{data_start}:F{data_end}',
    FormulaRule(formula=[f'F{data_start}<60'], fill=red_fill, font=Font(color='9C0006'))
)

# Conditional formatting for 分级
ws1.conditional_formatting.add(
    f'H{data_start}:H{data_end}',
    CellIsRule(operator='equal', formula=['"A"'], fill=green_fill, font=Font(color='006100', bold=True))
)
ws1.conditional_formatting.add(
    f'H{data_start}:H{data_end}',
    CellIsRule(operator='equal', formula=['"B"'], fill=yellow_fill, font=Font(color='9C6500'))
)
ws1.conditional_formatting.add(
    f'H{data_start}:H{data_end}',
    CellIsRule(operator='equal', formula=['"C"'], fill=red_fill, font=Font(color='9C0006', bold=True))
)

# ========== Sheet 2: 业务督导打分 ==========
ws2 = wb.create_sheet('业务督导打分')
ws2.freeze_panes = 'B2'
set_col_widths(ws2, [4, 24, 18, 18, 18, 18, 18, 18, 16])

headers2 = ['序号', '供应商名称', '目标达成率(15分)', '业绩贡献占比(10分)', '转化效率指数(10分)', '月度增长趋势(10分)', '人力稳定贡献(8分)', '有效人力占比(7分)', '合计']
for col, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=col, value=h)
style_header(ws2, len(headers2))

max_scores = [15, 10, 10, 10, 8, 7]
for r in range(data_start, data_end + 1):
    ws2.cell(row=r, column=1, value=r - data_start + 1)
    ws2.cell(row=r, column=2, value=f'供应商{r - data_start + 1}')
    for ci, ms in enumerate(max_scores):
        ws2.cell(row=r, column=ci + 3).value = 0
    ws2.cell(row=r, column=9).value = f'=SUM(C{r}:H{r})'

style_data_range(ws2, data_start, data_end, len(headers2))

# Data validation
dv1 = DataValidation(type='decimal', operator='between', formula1='0', formula2='15', allow_blank=True)
dv1.error = '请输入0-15之间的数字'
dv1.errorTitle = '输入错误'
ws2.add_data_validation(dv1)
dv1.add(f'C{data_start}:C{data_end}')

dv2 = DataValidation(type='decimal', operator='between', formula1='0', formula2='10', allow_blank=True)
dv2.error = '请输入0-10之间的数字'
ws2.add_data_validation(dv2)
dv2.add(f'D{data_start}:F{data_end}')

dv3 = DataValidation(type='decimal', operator='between', formula1='0', formula2='8', allow_blank=True)
ws2.add_data_validation(dv3)
dv3.add(f'G{data_start}:G{data_end}')

dv4 = DataValidation(type='decimal', operator='between', formula1='0', formula2='7', allow_blank=True)
ws2.add_data_validation(dv4)
dv4.add(f'H{data_start}:H{data_end}')

# Scoring reference table below data
ref_start = data_end + 3
ws2.cell(row=ref_start, column=1, value='评分标准参考').font = Font(name='微软雅黑', bold=True, size=11, color='1F4E79')

standards = [
    ['细项', '满分条件', '及格条件', '零分条件'],
    ['季度目标达成率(15分)', '达成率≥100%', '85%≤达成率<100%（线性插值）', '达成率<85%'],
    ['季度业绩贡献占比(10分)', '人均产出≥均值110%', '90%≤人均产出<110%（线性插值）', '人均产出<90%'],
    ['转化效率指数(10分)', '转化率达均值110%以上', '95%≤指数<110%（线性插值）', '指数<95%'],
    ['月度业绩增长趋势(10分)', '月均增长率≥5%', '-2%≤增长<5%（线性插值）', '月均增长率<-2%'],
    ['季度人力稳定贡献(8分)', '稳定人力留存率≥90%', '75%≤留存率<90%（线性插值）', '留存率<75%'],
    ['有效人力占比(7分)', '有效占比≥95%', '85%≤有效占比<95%（线性插值）', '有效占比<85%'],
]

for ri, row_data in enumerate(standards):
    for ci, val in enumerate(row_data):
        cell = ws2.cell(row=ref_start + 1 + ri, column=ci + 1, value=val)
        cell.font = bold_font if ri == 0 else data_font
        cell.fill = header_fill if ri == 0 else (alt_fill if ri % 2 == 0 else PatternFill())
        cell.alignment = center_align
        cell.border = thin_border
        if ri == 0:
            cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)

# ========== Sheet 3: 供应商管理打分 ==========
ws3 = wb.create_sheet('供应商管理打分')
ws3.freeze_panes = 'B2'
set_col_widths(ws3, [4, 24, 16, 16, 16, 16, 16, 16, 16])

headers3 = ['序号', '供应商名称', '配合意愿(6分)', '管理团队(5分)', '招聘培训(5分)', '合规风险(5分)', '资源稳定(5分)', '创新改进(4分)', '合计']
for col, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=h)
style_header(ws3, len(headers3))

weights = [6, 5, 5, 5, 5, 4]
for r in range(data_start, data_end + 1):
    ws3.cell(row=r, column=1, value=r - data_start + 1)
    ws3.cell(row=r, column=2, value=f'供应商{r - data_start + 1}')
    for ci in range(len(weights)):
        ws3.cell(row=r, column=ci + 3).value = 3  # default 合格
    # Actual score = SUM(raw_score/5 * weight)
    parts = [f'C{r}/5*6', f'D{r}/5*5', f'E{r}/5*5', f'F{r}/5*5', f'G{r}/5*5', f'H{r}/5*4']
    ws3.cell(row=r, column=9).value = '=' + '+'.join(parts)

style_data_range(ws3, data_start, data_end, len(headers3))

# Data validation: 1-5 integer
dv5 = DataValidation(type='whole', operator='between', formula1='1', formula2='5', allow_blank=True)
dv5.error = '请输入1-5的整数（1=差, 2=需改进, 3=合格, 4=良好, 5=优秀）'
ws3.add_data_validation(dv5)
dv5.add(f'C{data_start}:H{data_end}')

# Scoring reference
ref_start3 = data_end + 3
ws3.cell(row=ref_start3, column=1, value='5分制评估标准').font = Font(name='微软雅黑', bold=True, size=11, color='1F4E79')

standards3 = [
    ['评分', '标准定义'],
    ['5分-优秀', '超出预期，行业标杆水平，可作最佳实践推广'],
    ['4分-良好', '达到并部分超出要求，表现稳定可靠'],
    ['3分-合格', '满足基本要求，无重大问题，有提升空间'],
    ['2分-需改进', '未达基本要求，存在明显不足，需要辅导'],
    ['1分-差', '严重不达标，存在重大风险，需立即整改'],
    ['', ''],
    ['细项', '5分标准'],
    ['配合意愿(6分)', '需求2h内响应，24h出方案，主动跟进，季度零推诿'],
    ['管理团队(5分)', '驻场经理经验丰富，独立解决问题，管理例会高质量'],
    ['招聘培训(5分)', '满编率≥95%，新人7天达标率≥85%，有多元招聘渠道'],
    ['合规风险(5分)', '季度零合规事故，主动识别上报风险，合规100%'],
    ['资源稳定(5分)', '资源充足，核心团队年度流失<10%，产能波动<5%'],
    ['创新改进(4分)', '季度≥3条优化建议被采纳，积极参与试点项目'],
]

for ri, row_data in enumerate(standards3):
    for ci, val in enumerate(row_data):
        cell = ws3.cell(row=ref_start3 + 1 + ri, column=ci + 1, value=val)
        cell.font = bold_font if ri in (0, 7) else data_font
        cell.fill = header_fill if ri == 0 else (alt_fill if ri <= 5 and ri % 2 == 0 else PatternFill())
        cell.alignment = left_align
        cell.border = thin_border
        if ri == 0:
            cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)

# ========== Sheet 4: 质检打分 ==========
ws4 = wb.create_sheet('质检打分')
ws4.freeze_panes = 'B2'
set_col_widths(ws4, [4, 24, 20, 20, 18, 20, 16])

headers4 = ['序号', '供应商名称', '质检合格率(4分)', '合规违规次数(3分)', '客户投诉率(2分)', '话术规范执行率(1分)', '合计']
for col, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=col, value=h)
style_header(ws4, len(headers4))

for r in range(data_start, data_end + 1):
    ws4.cell(row=r, column=1, value=r - data_start + 1)
    ws4.cell(row=r, column=2, value=f'供应商{r - data_start + 1}')
    for ci in range(4):
        ws4.cell(row=r, column=ci + 3).value = 0
    ws4.cell(row=r, column=7).value = f'=SUM(C{r}:F{r})'

style_data_range(ws4, data_start, data_end, len(headers4))

# Data validation
dv_q1 = DataValidation(type='decimal', operator='between', formula1='0', formula2='4', allow_blank=True)
ws4.add_data_validation(dv_q1)
dv_q1.add(f'C{data_start}:C{data_end}')

dv_q2 = DataValidation(type='decimal', operator='between', formula1='0', formula2='3', allow_blank=True)
ws4.add_data_validation(dv_q2)
dv_q2.add(f'D{data_start}:D{data_end}')

dv_q3 = DataValidation(type='decimal', operator='between', formula1='0', formula2='2', allow_blank=True)
ws4.add_data_validation(dv_q3)
dv_q3.add(f'E{data_start}:E{data_end}')

dv_q4 = DataValidation(type='decimal', operator='between', formula1='0', formula2='1', allow_blank=True)
ws4.add_data_validation(dv_q4)
dv_q4.add(f'F{data_start}:F{data_end}')

# Scoring reference
ref_start4 = data_end + 3
ws4.cell(row=ref_start4, column=1, value='质检评分标准').font = Font(name='微软雅黑', bold=True, size=11, color='1F4E79')

standards4 = [
    ['细项', '满分', '评分规则'],
    ['质检合格率(4分)', '4分', '合格率≥95%满分；90%-95%每低1%扣0.4分；85%-90%每低1%扣0.8分；<85%零分'],
    ['合规违规次数(3分)', '3分', '零A类且B类≤2次满分；零A类B类3-5次每多1次扣0.5分；出现A类违规直接零分'],
    ['客户投诉率(2分)', '2分', '≤3/万满分；3-8/万每高1/万扣0.3分；>8/万零分'],
    ['话术规范执行率(1分)', '1分', '≥95%满分；85%-95%每低2%扣0.1分；<85%零分'],
]

for ri, row_data in enumerate(standards4):
    for ci, val in enumerate(row_data):
        cell = ws4.cell(row=ref_start4 + 1 + ri, column=ci + 1, value=val)
        cell.font = bold_font if ri == 0 else data_font
        cell.fill = header_fill if ri == 0 else (alt_fill if ri % 2 == 0 else PatternFill())
        cell.alignment = left_align
        cell.border = thin_border
        if ri == 0:
            cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
            cell.alignment = center_align

# ========== Sheet 5: 分级结果与管理动作 ==========
ws5 = wb.create_sheet('分级结果与管理动作')
ws5.freeze_panes = 'B2'
set_col_widths(ws5, [4, 24, 14, 12, 14, 50])

headers5 = ['序号', '供应商名称', '总分', '排名', '分级', '管理动作']
for col, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=col, value=h)
style_header(ws5, len(headers5))

for r in range(data_start, data_end + 1):
    ws5.cell(row=r, column=1, value=r - data_start + 1)
    ws5.cell(row=r, column=2, value=f"='打分总表'!B{r}")
    ws5.cell(row=r, column=3, value=f"='打分总表'!F{r}")
    ws5.cell(row=r, column=4, value=f"='打分总表'!G{r}")
    ws5.cell(row=r, column=5, value=f"='打分总表'!H{r}")
    formula_text = '分量优先倾斜|月度经营回顾|季度战略对话|新项目优先准入|质检频次50%'
    formula_text2 = '分量维持稳定|双周运营回顾|针对性辅导|季度能力提升计划|风险预警'
    formula_text3 = '分量压缩至15%以下|周度改善会|强制PIP|质检频次翻倍|负责人约谈'
    ws5.cell(row=r, column=6).value = '=IF(E' + str(r) + '="A","' + formula_text + '",IF(E' + str(r) + '="B","' + formula_text2 + '","' + formula_text3 + '"))'

style_data_range(ws5, data_start, data_end, len(headers5))

# Same conditional formatting
ws5.conditional_formatting.add(
    f'E{data_start}:E{data_end}',
    CellIsRule(operator='equal', formula=['"A"'], fill=green_fill, font=Font(color='006100', bold=True))
)
ws5.conditional_formatting.add(
    f'E{data_start}:E{data_end}',
    CellIsRule(operator='equal', formula=['"B"'], fill=yellow_fill, font=Font(color='9C6500'))
)
ws5.conditional_formatting.add(
    f'E{data_start}:E{data_end}',
    CellIsRule(operator='equal', formula=['"C"'], fill=red_fill, font=Font(color='9C0006', bold=True))
)

# Print settings for all sheets
for ws in [ws1, ws2, ws3, ws4, ws5]:
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

output_path = '/Users/sundanian/Documents/projects/ai-agents/my-agent/plans/2026-04-13-供应商分层分级体系/供应商打分模板.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
