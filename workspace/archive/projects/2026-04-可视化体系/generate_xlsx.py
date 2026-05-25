#!/usr/bin/env python3
"""生成供应商管理可视化体系数据收集表"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# ── 样式定义 ──
header_font = Font(name="system-ui", size=12, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
sub_header_font = Font(name="system-ui", size=10, bold=True, color="1A1A1A")
sub_header_fill = PatternFill(start_color="D4D1C7", end_color="D4D1C7", fill_type="solid")
accent_fill = PatternFill(start_color="C98A6A", end_color="C98A6A", fill_type="solid")
olive_fill = PatternFill(start_color="8F9B7F", end_color="8F9B7F", fill_type="solid")
light_fill = PatternFill(start_color="F5F4F1", end_color="F5F4F1", fill_type="solid")
normal_font = Font(name="system-ui", size=10)
hint_font = Font(name="system-ui", size=9, color="6B6B6B", italic=True)
title_font = Font(name="system-ui", size=14, bold=True, color="1A1A1A")
thin_border = Border(
    left=Side(style="thin", color="D4D1C7"),
    right=Side(style="thin", color="D4D1C7"),
    top=Side(style="thin", color="D4D1C7"),
    bottom=Side(style="thin", color="D4D1C7"),
)
wrap_align = Alignment(wrap_text=True, vertical="center")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_sub_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_data_row(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
        if alt:
            cell.fill = light_fill


def add_hint(ws, row, col, text):
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = hint_font


# ════════════════════════════════════════════════════════════════
# Sheet 字段定义：每个字段的完整说明
# ════════════════════════════════════════════════════════════════
ws_dict = wb.active
ws_dict.title = "0-字段定义"
ws_dict.sheet_properties.tabColor = "1A1A1A"

ws_dict.merge_cells("A1:F1")
ws_dict.cell(row=1, column=1, value="字段定义与填写说明").font = Font(name="system-ui", size=16, bold=True, color="1A1A1A")
ws_dict.row_dimensions[1].height = 36

ws_dict.merge_cells("A2:F2")
ws_dict.cell(row=2, column=1, value="每个字段一行：它是什么、怎么算、从哪来、填什么值、举个例子。拿不准就看这页。").font = hint_font

dict_headers = [("所属Sheet", 14), ("字段名", 18), ("定义", 28), ("计算方式/取值范围", 26), ("数据来源", 18), ("填写示例", 14)]
for i, (name, width) in enumerate(dict_headers, 1):
    ws_dict.cell(row=3, column=i, value=name)
    ws_dict.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws_dict, 3, 6)

# ── 字段定义数据 ──
fields = [
    # Sheet 1
    ("1-月度核心指标", "供应商名称", "供应商公司全称或简称", "文本", "固定列表", "毅航"),
    ("1-月度核心指标", "年月", "数据所属的年月", "格式：YYYY-MM", "手动填写", "2026-04"),
    ("1-月度核心指标", "业绩达成率(%)", "当月实际产出占目标产出的百分比。衡量供应商是否完成既定目标",
     "实际产出 / 目标产出 x 100\n范围：0-200%，100%为达标线", "业务系统导出，或供应商报送", "92（表示完成目标的92%）"),
    ("1-月度核心指标", "综合排名", "该供应商在所有供应商中的综合排名位置",
     "基于业绩+质量+产能+合规的加权综合得分排序\n范围：正整数，越小越好", "排名系统自动计算", "1（第1名）"),
    ("1-月度核心指标", "人效排名", "按人效指标单独排名的位置",
     "按人效值从高到低排序\n范围：正整数", "排名系统", "3（人效排第3）"),
    ("1-月度核心指标", "人效(实际产出/目标)", "每单位人力的产出效率，即FCI中的F值",
     "实际产出人数 / 总投入人数\n>1 = 超额，<1 = 不足\n范围：0.00-2.00", "业务系统导出", "1.05（超出目标5%）"),
    ("1-月度核心指标", "质量评分(0-100)", "质检部门对该供应商当月作业质量的综合评分",
     "质检部门月度打分\n范围：0-100，85分为良好线", "质检部门提供", "88"),
    ("1-月度核心指标", "产能利用率(%)", "实际使用的产能占理论最大产能的比例",
     "实际产能 / 理论最大产能 x 100\n范围：0-100%\n80%以上为健康", "业务系统/供应商报送", "82（产能用了82%）"),
    ("1-月度核心指标", "合规率(%)", "合规通话（或合规作业）占总量的比例",
     "合规通话数 / 总通话数 x 100\n范围：0-100%\n95%为底线", "质检/合规系统", "96（96%的通话合规）"),
    ("1-月度核心指标", "配合度评分(0-100)", "对该供应商配合意愿和响应速度的综合主观评价",
     "综合打分，考虑：\n· 响应速度（多久回复）\n· 主动性（是否主动汇报）\n· 配合意愿（是否接受调整）\n范围：0-100", "服务组业务管理岗主观评估", "90"),
    ("1-月度核心指标", "市场份额(%)", "该供应商在总业务量中的占比",
     "该供应商产出 / 总产出 x 100\n范围：0-100%", "业务系统统计", "38（占总量的38%）"),
    ("1-月度核心指标", "FCI-F值", "产出效率比。衡量实际产出与目标的关系",
     "F = 实际产出 / 目标产出\n>1 超额，=1 刚好，<1 未达标\n同「人效」字段", "自动计算，=人效字段值", "1.05"),
    ("1-月度核心指标", "FCI-C值", "投入产出比。衡量每单位投入的产出效率",
     "C = 总产出 / 总投入\n>1 高效，<1 低效\n反映资源使用效率", "业务系统统计", "1.15（每投入1得到1.15）"),
    ("1-月度核心指标", "FCI-I值", "客户满意度指数。衡量终端客户对该供应商服务的满意程度",
     "I = 满意客户数 / 总客户数\n或取满意度调查均分/100\n范围：0.00-1.00", "客户满意度调查/投诉率反推", "0.98（98%满意度）"),
    ("1-月度核心指标", "综合健康分", "7个维度加权计算的综合健康评分",
     "自动计算，权重：\n· 业绩达成 25%\n· 排名稳定 15%\n· 人效 15%\n· 质量评分 15%\n· 产能利用 10%\n· 合规率 10%\n· 配合度 10%\n范围：0-100", "自动计算（不用手填）", "84"),
    ("1-月度核心指标", "等级(A/B/C/D)", "根据综合健康分和各项指标自动判定的等级",
     "自动判定：\n· A级：业绩≥95% 且 排名Top3 且 质量≥90\n· B级：业绩85-95% 且 排名中上\n· C级：业绩70-85% 且 波动较大\n· D级：业绩<70% 且 连续不达标", "自动计算（不用手填）", "B+"),

    # Sheet 2
    ("2-周产能数据", "W1-W8(%)", "每周的产能利用率",
     "实际产能 / 理论最大产能 x 100\n每周记录一次，连续8周\n范围：0-100%", "业务系统周报", "82\n75\n88\n..."),

    # Sheet 3
    ("3-月度状态追踪", "业绩/排名/质量/产能/人效/合规",
     "该维度当月的整体状态判断",
     "三选一：\n· 达标（绿色）：正常或超出预期\n· 关注（黄色）：临界或轻微偏离\n· 异常（红色）：明显不达标或恶化",
     "根据Sheet 1数据判断\n阈值参考：\n· 业绩≥95%→达标，85-95%→关注，<85%→异常\n· 合规≥95%→达标，90-95%→关注，<90%→异常", "达标"),

    # Sheet 4
    ("4-风险预警", "业绩下滑风险", "该供应商业绩持续下滑的可能性评估",
     "四选一：低 / 中 / 高 / 严重\n判断依据：\n· 低：偶尔波动，趋势平稳\n· 中：连续1-2月下滑\n· 高：连续3月以上下滑\n· 严重：断崖式下跌或持续恶化",
     "基于Sheet 1业绩达成率趋势判断", "中"),
    ("4-风险预警", "人员流失风险", "关键岗位人员流失的可能性",
     "四选一：低 / 中 / 高 / 严重\n判断依据：\n· 低：团队稳定，流失率<5%\n· 中：偶有流失，5-10%\n· 高：频繁流失，10-20%\n· 严重：核心人员离职潮，>20%",
     "供应商报送 + 观察判断", "高"),
    ("4-风险预警", "质量合规风险", "质量评分和合规率出现异常的可能性",
     "四选一：低 / 中 / 高 / 严重", "质检部门反馈", "低"),
    ("4-风险预警", "产能波动风险", "产能利用率大幅波动的可能性",
     "四选一：低 / 中 / 高 / 严重", "Sheet 2 周数据波动判断", "严重"),
    ("4-风险预警", "配合度风险", "供应商配合意愿下降的可能性",
     "四选一：低 / 中 / 高 / 严重", "主观评估", "中"),
    ("4-风险预警", "综合风险等级", "5个维度的综合风险判断",
     "取5个维度中的最高风险等级\n如有2个以上「高」或任意「严重」，升级一级", "自动判断", "高"),

    # Sheet 5
    ("5-触发事件记录", "异常类型", "异常事件的分类",
     "八选一：\n· 业绩不达标\n· 单量骤降（>30%）\n· 转化率下滑\n· 合规率跌破（<95%）\n· 客户投诉激增\n· 人员流失（关键岗位）\n· 产能异常\n· 其他", "人工判断", "业绩不达标"),
    ("5-触发事件记录", "严重程度", "该事件的预警等级",
     "四选一：\n· 红色：业绩+排名双降\n· 橙色：单方面持续恶化\n· 黄色：需要关注但不紧急\n· 绿色：已恢复或误报",
     "对照触发规则矩阵(Unit 07)", "橙色"),
    ("5-触发事件记录", "处理状态", "当前处理进度",
     "三选一：待处理 / 处理中 / 已关闭", "人工更新", "处理中"),

    # Sheet 6
    ("6-整改跟踪", "约谈原因", "触发约谈的具体原因",
     "自由文本，简要说明\n如：连续2月业绩不达标", "人工填写", "连续2月业绩不达标"),
    ("6-整改跟踪", "整改事项", "约定的具体整改内容和目标",
     "自由文本，要具体可量化\n如：6月底前人效提升至0.90", "约谈时双方确认", "6月底前人效从0.75提升至0.90"),
    ("6-整改跟踪", "状态", "整改进度",
     "四选一：\n· 待启动：约谈完还没开始\n· 进行中：正在执行\n· 已完成：达成目标\n· 已逾期：超时未完成", "人工更新", "进行中"),

    # Sheet 7
    ("7-主管胜任力", "战略思维", "该主管理解业务目标、制定团队策略的能力",
     "0-100分\n90+优秀，75-89良好，60-74合格，<60不合格", "服务组评估", "88"),
    ("7-主管胜任力", "执行能力", "落地计划、推动团队完成任务的能力",
     "0-100分", "服务组评估", "82"),
    ("7-主管胜任力", "团队管理", "搭建团队、培养人员、管理情绪的能力",
     "0-100分", "服务组评估", "90"),
    ("7-主管胜任力", "创新意识", "主动提出改进方案、尝试新方法的能力",
     "0-100分", "服务组评估", "75"),
    ("7-主管胜任力", "沟通协调", "与甲方沟通、协调内部资源、向上反馈的能力",
     "0-100分", "服务组评估", "86"),

    # Sheet 8
    ("8-准入评估", "注册资金(万)", "供应商公司的注册资金",
     "单位：万元\n准入线：≥100万", "营业执照/天眼查", "500"),
    ("8-准入评估", "团队人数", "供应商可投入该项目的人数",
     "正整数\n准入线：≥50人", "供应商自报+实地确认", "120"),
    ("8-准入评估", "行业经验(年)", "供应商在同行业的运营年限",
     "正整数\n准入线：≥2年", "供应商自报+背景调查", "5"),
    ("8-准入评估", "资质合规", "是否具备必要的资质证书和合规记录",
     "是/否\n否=一票否决", "资质审核", "是"),
    ("8-准入评估", "技术能力评分", "技术方案、系统对接、数据安全方面的评分",
     "0-100分\n70分以下不建议通过", "技术评估", "82"),
    ("8-准入评估", "报价水平", "该供应商的报价在市场中的水平",
     "偏高/适中/偏低", "与现有供应商对比", "适中"),
    ("8-准入评估", "评估结论", "准入评审最终结论",
     "三选一：\n· 通过进入试运行\n· 不通过\n· 待补充材料", "评审会集体决议", "通过进入试运行"),

    # Sheet 9
    ("9-历史档案", "事件类型", "事件的分类",
     "七选一：\n· 约谈 · 整改 · 人员变动 · 产能调整\n· 合同变更 · 清退预警 · 其他", "人工分类", "约谈"),
    ("9-历史档案", "事件描述", "发生了什么事",
     "自由文本，写清：\n· 什么事\n· 涉及谁\n· 原因是什么", "人工填写", "月度约谈，产能提升方案确认"),
    ("9-历史档案", "处理结果", "怎么处理的，结果如何",
     "自由文本\n如：已制定3月提升计划", "人工填写", "已确认产能提升方案，目标4月底提升15%"),
    ("9-历史档案", "影响评估", "该事件对业务的实际影响",
     "自由文本\n如：导致当月产能下降10%", "人工评估", "轻微影响，已控制在单周内"),
]

row = 4
current_sheet = ""
for idx, (sheet, field, definition, calculation, source, example) in enumerate(fields):
    # 分组标题行
    if sheet != current_sheet:
        current_sheet = sheet
        ws_dict.merge_cells(f"A{row}:F{row}")
        ws_dict.cell(row=row, column=1, value=f"  {sheet}").font = Font(name="system-ui", size=11, bold=True, color="C98A6A")
        ws_dict.cell(row=row, column=1).fill = PatternFill(start_color="FAF9F7", end_color="FAF9F7", fill_type="solid")
        ws_dict.cell(row=row, column=1).border = thin_border
        for c in range(2, 7):
            ws_dict.cell(row=row, column=c).fill = PatternFill(start_color="FAF9F7", end_color="FAF9F7", fill_type="solid")
            ws_dict.cell(row=row, column=c).border = thin_border
        row += 1

    ws_dict.cell(row=row, column=1, value=sheet)
    ws_dict.cell(row=row, column=2, value=field)
    ws_dict.cell(row=row, column=3, value=definition)
    ws_dict.cell(row=row, column=4, value=calculation)
    ws_dict.cell(row=row, column=5, value=source)
    ws_dict.cell(row=row, column=6, value=example)
    style_data_row(ws_dict, row, 6, alt=(idx % 2 == 1))
    # 定义列和计算方式列要高一些以容纳多行
    ws_dict.row_dimensions[row].height = max(30, 16 * (calculation.count('\n') + 1))
    row += 1

# 底部补充说明
row += 1
ws_dict.merge_cells(f"A{row}:F{row}")
ws_dict.cell(row=row, column=1, value="常见问题").font = Font(name="system-ui", size=12, bold=True, color="1A1A1A")
row += 1

faqs = [
    ("Q: 数据从哪来？", "大部分从业务系统导出（排名、业绩、产能）。质检评分找质检部门。配合度评分由业务管理岗主观评估。"),
    ("Q: FCI三个值怎么填？", "F值=人效字段值（同一个数）。C值从业务系统统计投入产出比。I值从客户满意度调查取。如果暂时没有满意度数据，先用合规率代替。"),
    ("Q: 状态追踪（Sheet 3）的「达标/关注/异常」怎么判断？", "看Sheet 1的实际数值：业绩≥95%达标，85-95%关注，<85%异常。合规≥95%达标，90-95%关注，<90%异常。其他维度类似。"),
    ("Q: 风险预警（Sheet 4）多久评一次？", "每月一次，和月度核心指标同步。如果月中出现重大事件，立即更新。"),
    ("Q: 填不完怎么办？", "先填Sheet 1的核心字段（业绩/排名/人效/质量/产能/合规），这6个就能让大部分可视化跑起来。其他Sheet可以逐步补。"),
    ("Q: 这些数据给谁看？", "给自己用的——填完数据后告诉我，我用真实数据替换展示页里的示例数据，生成你的专属可视化体系。"),
]

for q, a in faqs:
    ws_dict.merge_cells(f"A{row}:F{row}")
    ws_dict.cell(row=row, column=1, value=q).font = Font(name="system-ui", size=10, bold=True, color="1A1A1A")
    ws_dict.cell(row=row, column=1).alignment = wrap_align
    ws_dict.row_dimensions[row].height = 20
    row += 1
    ws_dict.merge_cells(f"A{row}:F{row}")
    ws_dict.cell(row=row, column=1, value=f"  {a}").font = Font(name="system-ui", size=9, color="6B6B6B")
    ws_dict.cell(row=row, column=1).alignment = wrap_align
    ws_dict.row_dimensions[row].height = 30
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 1: 供应商月度核心指标
# ════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("1-月度核心指标")
ws1.sheet_properties.tabColor = "C98A6A"

# 标题行
ws1.merge_cells("A1:P1")
ws1.cell(row=1, column=1, value="供应商月度核心指标采集表").font = title_font
ws1.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 30

# 说明行
ws1.merge_cells("A2:P2")
ws1.cell(row=2, column=1, value="说明：每月5日前填完上月数据。黄色列为必填，灰色列为系统自动计算（不用填）。").font = hint_font
ws1.row_dimensions[2].height = 22

# 表头
headers = [
    ("供应商名称", 14),
    ("年月", 10),
    ("业绩达成率(%)", 14),
    ("综合排名", 10),
    ("人效排名", 10),
    ("人效(实际产出/目标)", 18),
    ("质量评分(0-100)", 14),
    ("产能利用率(%)", 14),
    ("合规率(%)", 12),
    ("配合度评分(0-100)", 14),
    ("市场份额(%)", 12),
    ("FCI-F值", 10),
    ("FCI-C值", 10),
    ("FCI-I值", 10),
    ("综合健康分", 12),
    ("等级(A/B/C/D)", 14),
]

for i, (name, width) in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=name)
    ws1.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws1, 3, len(headers))

# 示例数据
suppliers = ["毅航", "毛毛虫", "伽玛", "赛维斯", "岐力", "翰锐"]
sample_data = {
    "毅航": [92, 1, 1, 1.05, 88, 82, 96, 90, 38, 0.92, 1.15, 0.98],
    "毛毛虫": [85, 2, 3, 0.95, 82, 75, 94, 78, 22, 0.85, 1.08, 0.95],
    "伽玛": [72, 5, 6, 0.78, 70, 68, 88, 65, 8, 0.55, 0.78, 0.85],
    "赛维斯": [78, 3, 2, 0.88, 76, 72, 92, 72, 15, 0.78, 0.92, 1.02],
    "岐力": [70, 6, 5, 0.82, 68, 65, 86, 62, 7, 0.70, 0.95, 0.92],
    "翰锐": [68, 4, 4, 0.75, 65, 60, 90, 58, 10, 0.68, 0.85, 0.90],
}

row = 4
for idx, name in enumerate(suppliers):
    ws1.cell(row=row, column=1, value=name)
    ws1.cell(row=row, column=2, value="2026-04")
    d = sample_data[name]
    for c, v in enumerate(d, 3):
        ws1.cell(row=row, column=c, value=v)
    # FCI 和健康分、等级 留空让用户填或备注
    add_hint(ws1, row, 15, "自动计算")
    add_hint(ws1, row, 16, "自动计算")
    style_data_row(ws1, row, len(headers), alt=(idx % 2 == 1))
    row += 1

# 空行继续
for idx in range(6):
    ws1.cell(row=row, column=1, value="")
    ws1.cell(row=row, column=2, value="")
    style_data_row(ws1, row, len(headers), alt=(idx % 2 == 1))
    row += 1

# 底部说明
row += 1
ws1.merge_cells(f"A{row}:P{row}")
ws1.cell(row=row, column=1, value="字段说明：").font = sub_header_font
row += 1
notes = [
    "业绩达成率 = 当月实际产出 / 当月目标产出 x 100%",
    "人效 = 实际产出人数 / 总投入人数（F值）",
    "质量评分 = 质检部门月度评分（0-100分）",
    "产能利用率 = 实际产能 / 理论最大产能 x 100%",
    "合规率 = 合规通话数 / 总通话数 x 100%",
    "配合度 = 主观+客观综合评分（响应速度/主动性/配合意愿）",
    "FCI: F=实际产出/目标产出, C=总产出/总投入, I=客户满意度",
    "等级根据分级标准自动判定：A(≥95/Top3), B(85-95), C(70-85), D(<70)",
]
for note in notes:
    ws1.cell(row=row, column=1, value=f"  {note}").font = hint_font
    ws1.merge_cells(f"A{row}:P{row}")
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 2: 周产能数据
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2-周产能数据")
ws2.sheet_properties.tabColor = "8F9B7F"

ws2.merge_cells("A1:J1")
ws2.cell(row=1, column=1, value="供应商周产能数据采集表").font = title_font
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:J2")
ws2.cell(row=2, column=1, value="说明：每家供应商每周记录一次产能利用率。用于热力图(Unit 06)。").font = hint_font

headers2 = [
    ("供应商名称", 14),
    ("年月", 10),
    ("W1(%)", 10),
    ("W2(%)", 10),
    ("W3(%)", 10),
    ("W4(%)", 10),
    ("W5(%)", 10),
    ("W6(%)", 10),
    ("W7(%)", 10),
    ("W8(%)", 10),
]
for i, (name, width) in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=name)
    ws2.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws2, 3, len(headers2))

row = 4
for idx, name in enumerate(suppliers):
    ws2.cell(row=row, column=1, value=name)
    ws2.cell(row=row, column=2, value="2026-04")
    style_data_row(ws2, row, len(headers2), alt=(idx % 2 == 1))
    row += 1

for idx in range(6):
    style_data_row(ws2, row, len(headers2), alt=(idx % 2 == 1))
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 3: 月度状态追踪
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3-月度状态追踪")
ws3.sheet_properties.tabColor = "D4845A"

ws3.merge_cells("A1:I1")
ws3.cell(row=1, column=1, value="供应商月度状态色块追踪表").font = title_font
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:I2")
ws3.cell(row=2, column=1, value="说明：每个供应商每月填6个维度状态（达标/关注/异常）。用于色块图(Unit 10)。").font = hint_font

headers3 = [
    ("供应商名称", 14),
    ("年月", 10),
    ("业绩", 10),
    ("排名", 10),
    ("质量", 10),
    ("产能", 10),
    ("人效", 10),
    ("合规", 10),
    ("备注", 30),
]
for i, (name, width) in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=name)
    ws3.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws3, 3, len(headers3))

# 状态选项说明
row = 4
ws3.merge_cells(f"A{row}:I{row}")
ws3.cell(row=row, column=1, value="状态选项：达标（绿）/ 关注（黄）/ 异常（红）").font = Font(name="system-ui", size=9, color="8F9B7F")
row += 1

for idx, name in enumerate(suppliers):
    ws3.cell(row=row, column=1, value=name)
    ws3.cell(row=row, column=2, value="2026-04")
    style_data_row(ws3, row, len(headers3), alt=(idx % 2 == 1))
    row += 1

for idx in range(6):
    style_data_row(ws3, row, len(headers3), alt=(idx % 2 == 1))
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 4: 风险预警
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4-风险预警")
ws4.sheet_properties.tabColor = "C98A6A"

ws4.merge_cells("A1:H1")
ws4.cell(row=1, column=1, value="供应商风险预警评估表").font = title_font
ws4.row_dimensions[1].height = 30

ws4.merge_cells("A2:H2")
ws4.cell(row=2, column=1, value="说明：每月对每家供应商做5维风险评级（低/中/高/严重）。用于风险预警模板(Unit 16)。").font = hint_font

headers4 = [
    ("供应商名称", 14),
    ("年月", 10),
    ("业绩下滑风险", 14),
    ("人员流失风险", 14),
    ("质量合规风险", 14),
    ("产能波动风险", 14),
    ("配合度风险", 14),
    ("综合风险等级", 14),
]
for i, (name, width) in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=name)
    ws4.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws4, 3, len(headers4))

row = 4
ws4.merge_cells(f"A{row}:H{row}")
ws4.cell(row=row, column=1, value="风险等级：低 / 中 / 高 / 严重").font = Font(name="system-ui", size=9, color="D4845A")
row += 1

for idx, name in enumerate(suppliers):
    ws4.cell(row=row, column=1, value=name)
    ws4.cell(row=row, column=2, value="2026-04")
    style_data_row(ws4, row, len(headers4), alt=(idx % 2 == 1))
    row += 1

for idx in range(6):
    style_data_row(ws4, row, len(headers4), alt=(idx % 2 == 1))
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 5: 触发事件记录
# ════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5-触发事件记录")
ws5.sheet_properties.tabColor = "D4845A"

ws5.merge_cells("A1:G1")
ws5.cell(row=1, column=1, value="异常触发事件记录表").font = title_font
ws5.row_dimensions[1].height = 30

ws5.merge_cells("A2:G2")
ws5.cell(row=2, column=1, value="说明：发现异常时记录。用于触发规则矩阵(Unit 07)、信号速查表(Unit 14)、异常模式(Unit 17)。").font = hint_font

headers5 = [
    ("日期", 12),
    ("供应商名称", 14),
    ("异常类型", 14),
    ("具体描述", 30),
    ("严重程度", 10),
    ("处理状态", 10),
    ("备注", 20),
]
for i, (name, width) in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=name)
    ws5.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws5, 3, len(headers5))

row = 4
ws5.merge_cells(f"A{row}:G{row}")
ws5.cell(row=row, column=1, value="异常类型：业绩不达标 / 单量骤降 / 转化率下滑 / 合规率跌破 / 客户投诉 / 人员流失 / 产能异常 / 其他").font = hint_font
row += 1
ws5.merge_cells(f"A{row}:G{row}")
ws5.cell(row=row, column=1, value="严重程度：红色（双降）/ 橙色（单降）/ 黄色（关注）/ 绿色（恢复）").font = hint_font
row += 1
ws5.merge_cells(f"A{row}:G{row}")
ws5.cell(row=row, column=1, value="处理状态：待处理 / 处理中 / 已关闭").font = hint_font
row += 1

for idx in range(15):
    style_data_row(ws5, row, len(headers5), alt=(idx % 2 == 1))
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 6: 整改跟踪
# ════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6-整改跟踪")
ws6.sheet_properties.tabColor = "6A8FB5"

ws6.merge_cells("A1:H1")
ws6.cell(row=1, column=1, value="供应商整改事项跟踪表").font = title_font
ws6.row_dimensions[1].height = 30

ws6.merge_cells("A2:H2")
ws6.cell(row=2, column=1, value="说明：约谈/整改后录入。用于整改检查清单(Unit 21)、约谈框架(Unit 20)闭环。").font = hint_font

headers6 = [
    ("供应商名称", 14),
    ("约谈日期", 12),
    ("约谈原因", 20),
    ("整改事项", 25),
    ("完成时限", 12),
    ("责任人", 10),
    ("当前进展", 20),
    ("状态", 10),
]
for i, (name, width) in enumerate(headers6, 1):
    ws6.cell(row=3, column=i, value=name)
    ws6.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws6, 3, len(headers6))

row = 4
ws6.merge_cells(f"A{row}:H{row}")
ws6.cell(row=row, column=1, value="状态：待启动 / 进行中 / 已完成 / 已逾期").font = hint_font
row += 1

for idx in range(12):
    style_data_row(ws6, row, len(headers6), alt=(idx % 2 == 1))
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 7: 主管胜任力
# ════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7-主管胜任力")
ws7.sheet_properties.tabColor = "8F9B7F"

ws7.merge_cells("A1:H1")
ws7.cell(row=1, column=1, value="供应商主管胜任力评分表").font = title_font
ws7.row_dimensions[1].height = 30

ws7.merge_cells("A2:H2")
ws7.cell(row=2, column=1, value="说明：每季度评估一次。用于主管胜任力雷达(Unit 08)。评分0-100。").font = hint_font

headers7 = [
    ("供应商名称", 14),
    ("主管姓名", 12),
    ("评估日期", 12),
    ("战略思维", 10),
    ("执行能力", 10),
    ("团队管理", 10),
    ("创新意识", 10),
    ("沟通协调", 10),
]
for i, (name, width) in enumerate(headers7, 1):
    ws7.cell(row=3, column=i, value=name)
    ws7.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws7, 3, len(headers7))

row = 4
for idx, name in enumerate(suppliers):
    ws7.cell(row=row, column=1, value=name)
    style_data_row(ws7, row, len(headers7), alt=(idx % 2 == 1))
    row += 1

for idx in range(6):
    style_data_row(ws7, row, len(headers7), alt=(idx % 2 == 1))
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 8: 准入评估
# ════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("8-准入评估")
ws8.sheet_properties.tabColor = "6A8FB5"

ws8.merge_cells("A1:J1")
ws8.cell(row=1, column=1, value="新供应商准入评估表").font = title_font
ws8.row_dimensions[1].height = 30

ws8.merge_cells("A2:J2")
ws8.cell(row=2, column=1, value="说明：新供应商引入时填写。用于准入评估(Unit 22)。").font = hint_font

headers8 = [
    ("供应商名称", 14),
    ("申请日期", 12),
    ("注册资金(万)", 12),
    ("团队人数", 10),
    ("行业经验(年)", 12),
    ("资质合规", 10),
    ("技术能力评分", 12),
    ("报价水平", 10),
    ("评估结论", 12),
    ("审批人", 10),
]
for i, (name, width) in enumerate(headers8, 1):
    ws8.cell(row=3, column=i, value=name)
    ws8.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws8, 3, len(headers8))

row = 4
ws8.merge_cells(f"A{row}:J{row}")
ws8.cell(row=row, column=1, value="准入标准参考：注册资金≥100万 · 团队≥50人 · 无重大合规记录 · 行业经验≥2年").font = hint_font
row += 1
ws8.merge_cells(f"A{row}:J{row}")
ws8.cell(row=row, column=1, value="评估结论：通过进入试运行 / 不通过 / 待补充材料").font = hint_font
row += 1

for idx in range(5):
    style_data_row(ws8, row, len(headers8), alt=(idx % 2 == 1))
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 9: 历史档案
# ════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("9-历史档案")
ws9.sheet_properties.tabColor = "1A1A1A"

ws9.merge_cells("A1:G1")
ws9.cell(row=1, column=1, value="供应商历史事件档案").font = title_font
ws9.row_dimensions[1].height = 30

ws9.merge_cells("A2:G2")
ws9.cell(row=2, column=1, value="说明：重要事件随时记录。用于供应商档案(Unit 30)、独立空间(Unit 26)。").font = hint_font

headers9 = [
    ("供应商名称", 14),
    ("事件日期", 12),
    ("事件类型", 14),
    ("事件描述", 30),
    ("处理结果", 20),
    ("影响评估", 15),
    ("关联单元", 12),
]
for i, (name, width) in enumerate(headers9, 1):
    ws9.cell(row=3, column=i, value=name)
    ws9.column_dimensions[get_column_letter(i)].width = width
style_header_row(ws9, 3, len(headers9))

row = 4
ws9.merge_cells(f"A{row}:G{row}")
ws9.cell(row=row, column=1, value="事件类型：约谈 / 整改 / 人员变动 / 产能调整 / 合同变更 / 清退预警 / 其他").font = hint_font
row += 1

for idx in range(15):
    style_data_row(ws9, row, len(headers9), alt=(idx % 2 == 1))
    row += 1


# ════════════════════════════════════════════════════════════════
# Sheet 10: 收集指引（汇总页）
# ════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("0-收集指引")
ws10.sheet_properties.tabColor = "C98A6A"

# 移到最前面（现在有11个sheet）
wb.move_sheet("0-收集指引", offset=-10)

ws10.merge_cells("A1:D1")
ws10.cell(row=1, column=1, value="供应商管理可视化体系 · 数据收集指引").font = Font(name="system-ui", size=16, bold=True, color="1A1A1A")
ws10.row_dimensions[1].height = 36

ws10.merge_cells("A3:D3")
ws10.cell(row=3, column=1, value="Sheet 与可视化单元对应关系").font = Font(name="system-ui", size=12, bold=True, color="1A1A1A")

guide_headers = ["Sheet", "收集内容", "频率", "对应可视化单元"]
for i, h in enumerate(guide_headers, 1):
    ws10.cell(row=4, column=i, value=h)
    ws10.column_dimensions[get_column_letter(i)].width = [12, 30, 10, 30][i - 1]
style_header_row(ws10, 4, 4)

guide_data = [
    ["1-月度核心指标", "业绩/排名/人效/质量/产能/合规/配合度/FCI", "每月", "01 02 04 09 26 28 30"],
    ["2-周产能数据", "近8周产能利用率", "每周", "06"],
    ["3-月度状态追踪", "6维度月度状态（达标/关注/异常）", "每月", "10"],
    ["4-风险预警", "5维风险评级", "每月", "16"],
    ["5-触发事件记录", "异常触发事件明细", "实时", "07 14 17"],
    ["6-整改跟踪", "约谈后的整改事项闭环", "实时", "20 21"],
    ["7-主管胜任力", "主管5维度评分", "每季度", "08"],
    ["8-准入评估", "新供应商准入评估", "按需", "22"],
    ["9-历史档案", "供应商重要事件时间线", "实时", "26 30"],
]

row = 5
for idx, (sheet, content, freq, units) in enumerate(guide_data):
    ws10.cell(row=row, column=1, value=sheet)
    ws10.cell(row=row, column=2, value=content)
    ws10.cell(row=row, column=3, value=freq)
    ws10.cell(row=row, column=4, value=units)
    style_data_row(ws10, row, 4, alt=(idx % 2 == 1))
    row += 1

row += 1
ws10.merge_cells(f"A{row}:D{row}")
ws10.cell(row=row, column=1, value="收集优先级建议").font = Font(name="system-ui", size=12, bold=True, color="1A1A1A")
row += 1

priority_data = [
    ["第一优先", "Sheet 1 月度核心指标", "6家头部供应商，每月5日前"],
    ["第二优先", "Sheet 2 周产能 + Sheet 3 状态追踪", "数据到位后热力图和色块图可运行"],
    ["第三优先", "Sheet 5 触发事件", "发现异常时随手记录"],
    ["按需收集", "Sheet 4/6/7/8/9", "约谈/整改/准入/档案按事件触发"],
]
for idx, (p, sheet, note) in enumerate(priority_data):
    ws10.cell(row=row, column=1, value=p).font = Font(name="system-ui", size=10, bold=True, color="C98A6A")
    ws10.cell(row=row, column=2, value=sheet)
    ws10.cell(row=row, column=3, value=note)
    ws10.merge_cells(f"C{row}:D{row}")
    style_data_row(ws10, row, 4, alt=(idx % 2 == 1))
    row += 1

row += 1
ws10.merge_cells(f"A{row}:D{row}")
ws10.cell(row=row, column=1, value="供应商清单（当前头部6家，可自行添加）").font = Font(name="system-ui", size=12, bold=True, color="1A1A1A")
row += 1
for name in suppliers:
    ws10.cell(row=row, column=1, value=name)
    style_data_row(ws10, row, 4)
    row += 1


# ── 保存 ──
output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "供应商可视化体系_数据收集表.xlsx")
wb.save(output)
print(f"Done: {output}")
