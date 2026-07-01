import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# 样式定义
header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
module_font = Font(name="Microsoft YaHei", bold=True, size=11, color="2F5496")
module_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
normal_font = Font(name="Microsoft YaHei", size=10)
note_font = Font(name="Microsoft YaHei", size=9, color="666666")
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

def style_row(ws, row, cols, font=normal_font, fill=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if fill:
            cell.fill = fill

def add_module_header(ws, row, text, cols):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    style_row(ws, row, cols, font=module_font, fill=module_fill)

def add_questions(ws, start_row, questions, cols):
    """questions: list of (问题, 为什么要问, 可能的追问)"""
    row = start_row
    for q in questions:
        ws.cell(row=row, column=1, value=q[0])
        ws.cell(row=row, column=2, value=q[1])
        ws.cell(row=row, column=3, value=q[2])
        ws.cell(row=row, column=4, value="")  # 现场记录
        ws.cell(row=row, column=5, value="")  # 备注
        style_row(ws, row, cols)
        row += 1
    return row

def create_question_sheet(ws, title, sections, cols=5):
    """sections: list of (模块名, [(问题, 为什么, 追问), ...])"""
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 15

    # 标题行
    ws.cell(row=1, column=1, value=f"调研工具包 — {title}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    title_font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
    ws.cell(row=1, column=1).font = title_font

    # 表头
    headers = ["问题", "为什么要问", "可能的追问", "现场记录", "备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header(ws, 2, cols)

    row = 3
    for section_name, questions in sections:
        add_module_header(ws, row, section_name, cols)
        row += 1
        row = add_questions(ws, row, questions, cols)
        row += 1  # 空行

    return row

# ============================================================
# Sheet 1: 使用说明
# ============================================================
ws = wb.active
ws.title = "使用说明"
ws.column_dimensions["A"].width = 80

instructions = [
    ("供应商深度调研工具包", Font(name="Microsoft YaHei", bold=True, size=16, color="2F5496")),
    ("", None),
    ("一、使用说明", module_font),
    ("1. 调研前：打印本工具包 + 准备好供应商数据", normal_font),
    ('2. 调研中：按模块顺序提问，在"现场记录"列实时记录', normal_font),
    ('3. 调研后：对照"调研后整理"sheet检查覆盖度', normal_font),
    ("4. 汇报前：用数据对比sheet整理核心发现", normal_font),
    ("", None),
    ("二、核心目标", module_font),
    ("1. 验场：站点硬件、团队规模、合规基础", normal_font),
    ("2. 业务诊断：为什么人均收入不理想？根因在哪？", normal_font),
    ("3. 竞品了解：这家同时在做的竞品产品，一线真实偏好", normal_font),
    ("4. 减量决策：给老板汇报的数据支撑", normal_font),
    ("", None),
    ("三、提问节奏", module_font),
    ('• 先开放后聚焦：用"怎么样/怎么看/什么原因"开场，后段用数据追问', normal_font),
    ("• 先轻松后敏感：收入/名单/竞品放到中后段", normal_font),
    ("• 先听再说：沉默3秒，对方往往会补充更多信息", normal_font),
    ('• 数据验证：每个回答都要想到"我有什么数据可以验证"', normal_font),
    ("", None),
    ("四、时间分配（约3小时）", module_font),
    ("• 到达后：验场 30min", normal_font),
    ("• 上午：供应商经营 + 业务运营 1-1.5h", normal_font),
    ("• 午饭前后：一线坐席访谈 45min", normal_font),
    ("• 下午：竞品情报 + 合规质检 + 团队管理 1h", normal_font),
    ("• 离开前：管理层二次沟通 30min", normal_font),
    ("", None),
    ("五、后续动作", module_font),
    ('1. 24h内：完成调研记录整理（见"调研后整理"sheet）', normal_font),
    ("2. 3天内：产出调研报告初稿", normal_font),
    ("3. 回公司后：给老板汇报（4页以内）", normal_font),
]

for i, (text, font) in enumerate(instructions, 1):
    if text:
        ws.cell(row=i, column=1, value=text)
        if font:
            ws.cell(row=i, column=1).font = font

# ============================================================
# Sheet 2: 调研前准备
# ============================================================
ws2 = wb.create_sheet("调研前准备")
ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 30
ws2.column_dimensions["D"].width = 15
ws2.column_dimensions["E"].width = 20

headers = ["类别", "准备事项", "具体要求", "完成状态", "备注"]
for col, h in enumerate(headers, 1):
    ws2.cell(row=1, column=col, value=h)
style_header(ws2, 1, 5)

preps = [
    ("数据准备", "该供应商近6个月业绩数据", "人均产能、结算、达标率、趋势", "", ""),
    ("数据准备", "其他供应商同期数据", "做对比基准，特别是最大那家（人均10000）", "", ""),
    ("数据准备", "企业主贷大额产品要素表", "利率/期限/准入条件/审批时效", "", ""),
    ("数据准备", "竞品产品要素表（如有）", "产品名称/利率/额度/准入", "", ""),
    ("数据准备", "该供应商合同/结算标准", "合同单价、达标标准、激励政策", "", ""),
    ("设备准备", "笔记本电脑/平板", "现场填写本Excel", "", ""),
    ("设备准备", "录音笔（征得同意后）", "关键对话记录", "", ""),
    ("设备准备", "相机/手机", "职场环境拍照", "", ""),
    ("人员协调", "确认供应商对接人", "老板/总经理在场时间", "", ""),
    ("人员协调", "确认业务同事参与", "业务交流环节", "", ""),
    ("人员协调", "提前告知要看一线坐席", "供应商提前安排2-3人", "", ""),
    ("资料准备", "验场检查清单", "打印或存平板", "", ""),
    ("资料准备", "之前合作相关文件", "历史问题、整改记录", "", ""),
]

for i, row_data in enumerate(preps, 2):
    for j, val in enumerate(row_data, 1):
        ws2.cell(row=i, column=j, value=val)
    style_row(ws2, i, 5)

# ============================================================
# Sheet 3: 模块1 - 供应商经营画像
# ============================================================
ws3 = wb.create_sheet("1-供应商经营")
sections_1 = [
    ("A. 人员规模与产能关系", [
        ("目前整个团队多少人？月度人力变化情况？", "摸清真实规模和趋势", "和我们系统数据对得上吗？"),
        ("人均月产能多少？近6个月趋势？", "找到拐点，定位问题时间", "从什么时候开始变化的？有什么事件？"),
        ("TOP 20%和后20%的人均产能差距多少？", "判断是整体问题还是局部问题", "后20%主要是新人还是老人？"),
        ("现在这么多人都能打满吗？还是有冗余？", "冗余人力分析", "如果要调整，你觉得多少人合理？"),
    ]),
    ("B. 经营模型与成本结构", [
        ("目前收入能覆盖成本吗？盈亏平衡点在哪？", "保本线测算", "算上管理成本、场地、设备折旧"),
        ("主要成本构成是什么？占比多少？", "成本结构分析", "人力成本占比？固定vs变动？"),
        ("今年利润和去年比怎么样？", "经营趋势", "主要原因是什么？"),
        ("人员越多利润越高还是越低？为什么？", "规模效应分析", "边际成本是递增还是递减？"),
    ]),
    ("C. 其他业务情况", [
        ("除了我们，还在做哪些甲方的产品？", "业务多元化，判断依赖度", "哪些产品利润高？"),
        ("各产品线分别多少人？利润占比？", "资源分配逻辑", "有没有在调整重心？"),
        ("哪些产品增长好？哪些在萎缩？", "业务趋势", "原因是什么？"),
        ("竞品产品和我们产品，你更愿意做哪个？为什么？", "最真实的信号", "结算、难度、稳定性哪个更重要？"),
    ]),
    ("D. 合作意愿与战略思考", [
        ("对我们这个产品的定位是什么？核心/补充/过渡？", "判断长期合作意愿", "未来6个月打算怎么发展？"),
        ("如果产品调整或市场变化，你会怎么应对？", "风险承受能力", "有没有备选方案？"),
        ("理想的合作模式是什么样的？", "了解诉求", "结算、名单、培训、信息方面"),
        ("对我们有什么建议或意见？", "开放性问题", "产品、名单、政策、沟通"),
    ]),
]
create_question_sheet(ws3, "模块1·供应商经营画像", sections_1)

# ============================================================
# Sheet 4: 模块2 - 业务运营深度
# ============================================================
ws4 = wb.create_sheet("2-业务运营")
sections_2 = [
    ("A. 产能分层", [
        ("TOP 20%坐席月均产能多少？为什么能做好？", "标杆解码", "有没有可复制的方法？"),
        ("后20%坐席月均产能多少？问题在哪？", "找到改进空间", "是态度、能力还是资源问题？"),
        ("中间层占比多少？有没有提升空间？", "中间层是改进杠杆", "用什么方法能推上去？"),
        ("各组产能差异大吗？为什么？", "管理因素分析", "主管风格对产能影响大吗？"),
    ]),
    ("B. 转化漏斗", [
        ("人均日拨打量多少？有效通话量？", "漏斗第一层", "有没有拨打量不够的情况？"),
        ("接通率多少？近6个月趋势？", "漏斗第二层", "接通率下降的原因？"),
        ("有效接通到添加企微/微信的转化率？", "漏斗第三层", "哪一步流失最大？"),
        ("企微到申请提交的转化率？", "漏斗第四层", "客户放弃的原因？"),
        ("申请到放款的通过率？", "漏斗最终层", "拒绝的主要原因？"),
    ]),
    ("C. 流失分析", [
        ("流失率多少？新人和老人分别多少？", "流失结构", "新人流失主要原因？"),
        ("流失的人去了哪里？同行业/跨行？", "流失方向", "是我们的待遇问题还是行业问题？"),
        ("新人多久能出产能？培训周期多长？", "新人效率", "有没有缩短周期的办法？"),
        ("新人流失对公司成本影响多大？", "量化流失成本", "招聘+培训+空窗期成本"),
    ]),
    ("D. 名单管理", [
        ("目前名单来源是什么？转化率如何？", "名单质量", "不同来源的名单差异大吗？"),
        ("名单够用吗？有没有断档的时候？", "名单稳定性", "断档时怎么处理？"),
        ("复打名单占比多少？复打效果怎么样？", "名单利用率", "第N次打的效果衰减多少？"),
        ("名单管理有什么问题或建议？", "一线反馈", "什么名单最好打？"),
    ]),
]
create_question_sheet(ws4, "模块2·业务运营深度", sections_2)

# ============================================================
# Sheet 5: 模块3 - 名单管理
# ============================================================
ws5 = wb.create_sheet("3-名单管理")
sections_3 = [
    ("A. 名单来源与分配", [
        ("目前名单主要来源？各渠道占比？", "名单结构", "哪个渠道质量最高？"),
        ("名单分配机制是什么？按什么标准分？", "分配公平性", "有没有挑名单的情况？"),
        ("不同坐席拿到的名单质量差异大吗？", "公平性", "TOP坐席和普通坐席拿到的一样吗？"),
        ("名单到手后多久开始打？有没有积压？", "时效性", "积压影响转化率吗？"),
    ]),
    ("B. 名单质量", [
        ("不同类型名单的接通率差异？", "质量分层", "最好打的是什么类型？"),
        ("空号/停机/拒接的比例多少？", "名单有效率", "有没有筛号环节？"),
        ("名单更新频率？多久换一批新名单？", "新鲜度", "旧名单还能打吗？"),
        ("对名单质量有什么改进建议？", "一线洞察", "什么样的名单最好打？"),
    ]),
    ("C. 名单使用效率", [
        ("平均每个名单打几次？", "利用深度", "第N次打的接通率衰减多少？"),
        ("有效沟通后的跟进频率？", "跟进机制", "有没有因为跟进不及时丢单的？"),
        ("有没有名单复用的情况？效果怎么样？", "复用价值", "多久可以复打一次？"),
        ("名单管理最大的痛点是什么？", "核心问题", "如果能改一个点，你最想改什么？"),
    ]),
]
create_question_sheet(ws5, "模块3·名单管理", sections_3)

# ============================================================
# Sheet 6: 模块4 - 竞品情报
# ============================================================
ws6 = wb.create_sheet("4-竞品情报")
sections_4 = [
    ("A. 竞品产品信息", [
        ("目前在做的竞品产品名称？做了多久？", "基础信息", "什么时候开始做的？"),
        ("竞品的利率/费率、额度、期限、准入条件？", "产品要素对比", "和我们的差异在哪？"),
        ("竞品的审批时效？通过率？放款速度？", "流程体验", "比我们快还是慢？"),
        ("竞品的目标客群？和我们重叠吗？", "客群分析", "是同一批客户还是不同客户？"),
    ]),
    ("B. 竞品电销运营", [
        ("竞品人均日拨打量、接通率、转化率？", "运营数据对比", "和我们产品比哪个好做？"),
        ("竞品的人均月结算多少？", "结算对比", "比我们高还是低？"),
        ("竞品的名单来源？质量怎么样？", "名单对比", "名单是核心差异吗？"),
        ("竞品的培训周期？上手难度？", "运营成本对比", "新人多久能出产能？"),
        ("竞品的合规要求和我们比？", "合规成本", "更严还是更松？"),
    ]),
    ("C. 一线真实偏好", [
        ("坐席更愿意推哪个产品？为什么？", "最真实的一线信号", "收入、难度、稳定性哪个更重要？"),
        ("推我们产品时客户最常见的拒绝理由？", "产品痛点", "这些理由竞品能解决吗？"),
        ("推竞品时客户最常见的拒绝理由？", "竞品弱点", "我们能解决这些问题吗？"),
        ("有没有客户同时适合两个产品？怎么选？", "交叉场景", "坐席怎么判断推哪个？"),
    ]),
    ("D. 市场洞察", [
        ("你觉得这个市场还有多大空间？", "市场判断", "客户接受度在提升还是下降？"),
        ("竞品在市场上的口碑怎么样？", "品牌感知", "客户/渠道怎么评价？"),
        ("有没有其他竞品在进入这个市场？", "竞争格局", "新进入者有什么特点？"),
        ("如果我们改进一个点，你最建议改什么？", "最优先的改进项", "为什么？"),
    ]),
]
create_question_sheet(ws6, "模块4·竞品情报", sections_4)

# ============================================================
# Sheet 7: 模块5 - 合规质检
# ============================================================
ws7 = wb.create_sheet("5-合规质检")
sections_5 = [
    ("A. 合规管控", [
        ("合规培训怎么做？多久一次？", "合规基础", "新人入职有合规培训吗？"),
        ("违规率多少？常见违规类型？", "风险暴露", "哪类违规最多？"),
        ("重大违规怎么处理？有案例吗？", "处理机制", "处理结果是什么？"),
        ("客户投诉率多少？主要原因？", "客户体验", "投诉怎么处理的？"),
    ]),
    ("B. 质检执行", [
        ("质检覆盖率多少？日均抽检几通？", "质检强度", "够用吗？"),
        ("质检标准是什么？有明确的评分卡吗？", "标准化程度", "标准执行一致吗？"),
        ("质检问题怎么反馈给坐席？有闭环吗？", "改进机制", "反馈后有改善吗？"),
        ("有没有AI质检？效果怎么样？", "技术应用", "和人工质检比准确率如何？"),
    ]),
    ("C. 录音质检", [
        ("录音保存多久？能随时调取吗？", "合规基础", "保存时间够合规要求吗？"),
        ("每月抽听多少通？发现什么问题？", "抽检强度", "问题率是多少？"),
        ("有没有因为录音问题被投诉或处罚的？", "风险案例", "怎么处理的？"),
        ("录音质检最大的难点是什么？", "痛点", "需要什么支持？"),
    ]),
]
create_question_sheet(ws7, "模块5·合规质检", sections_5)

# ============================================================
# Sheet 8: 模块6 - 团队与管理
# ============================================================
ws8 = wb.create_sheet("6-团队管理")
sections_6 = [
    ("A. 团队结构", [
        ("团队怎么分组？每组多少人？", "管理幅度", "组长管多少人合适？"),
        ("管理层级几层？信息传递顺畅吗？", "组织效率", "有没有信息失真的情况？"),
        ("有没有末位淘汰？怎么做的？", "竞争机制", "淘汰率多少？"),
        ("新老员工比例？经验断层吗？", "梯队建设", "有没有骨干流失的情况？"),
    ]),
    ("B. 激励与考核", [
        ("薪酬结构是什么？底薪+提成比例？", "薪酬竞争力", "在市场上有竞争力吗？"),
        ("考核指标有哪些？权重多少？", "考核导向", "指标合理吗？"),
        ("激励政策有效吗？坐席最在意什么？", "激励效果", "钱/荣誉/成长哪个更重要？"),
        ("收入满意度怎么样？和同行比呢？", "留人关键", "不满意的人去了哪？"),
    ]),
    ("C. 培训体系", [
        ("新人培训周期多长？内容是什么？", "培训效率", "能不能缩短？"),
        ("在职培训怎么做？频率？", "能力提升", "效果怎么衡量？"),
        ("培训后上岗达标率多少？", "培训效果", "不达标的主要原因？"),
        ("有没有标杆坐席分享经验？", "知识传承", "好经验能复制吗？"),
    ]),
    ("D. 管理工具与数据", [
        ("用什么系统管理？CRM/外呼系统？", "工具基础", "系统好用吗？有什么问题？"),
        ("数据看板能看到什么数据？", "数据透明度", "够用吗？还需要什么数据？"),
        ("日常管理靠什么？数据还是经验？", "管理方式", "数据驱动程度"),
        ("管理上最大的挑战是什么？", "核心痛点", "需要什么支持？"),
    ]),
]
create_question_sheet(ws8, "模块6·团队与管理", sections_6)

# ============================================================
# Sheet 9: 数据对比
# ============================================================
ws9 = wb.create_sheet("数据对比")
ws9.column_dimensions["A"].width = 20
ws9.column_dimensions["B"].width = 20
ws9.column_dimensions["C"].width = 20
ws9.column_dimensions["D"].width = 20
ws9.column_dimensions["E"].width = 20
ws9.column_dimensions["F"].width = 20

ws9.cell(row=1, column=1, value="调研数据对比表")
ws9.cell(row=1, column=1).font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
ws9.merge_cells("A1:F1")

# 供应商基本情况对比
headers_basic = ["指标", "本次调研供应商", "最大供应商(基准)", "其他供应商均值", "差距", "备注"]
for col, h in enumerate(headers_basic, 1):
    ws9.cell(row=3, column=col, value=h)
style_header(ws9, 3, 6)

basic_rows = [
    ("在岗人数", "", "100人", "", "", ""),
    ("人均月结算", "", "10000元", "~7000元", "", ""),
    ("人均日拨打量", "", "", "", "", ""),
    ("接通率", "", "", "", "", ""),
    ("转化率", "", "", "", "", ""),
    ("月度流失率", "", "", "", "", ""),
    ("新人出单周期", "", "", "", "", ""),
    ("人均利润", "", "", "", "", ""),
    ("盈亏平衡人数", "", "", "", "", ""),
]

for i, row_data in enumerate(basic_rows, 4):
    for j, val in enumerate(row_data, 1):
        ws9.cell(row=i, column=j, value=val)
    style_row(ws9, i, 6)

# 产品对比
add_module_header(ws9, 14, "产品对比：我方 vs 竞品", 6)

prod_headers = ["指标", "我方产品", "竞品产品", "差异分析", "一线感受", "备注"]
for col, h in enumerate(prod_headers, 1):
    ws9.cell(row=15, column=col, value=h)
style_header(ws9, 15, 6)

prod_rows = [
    ("产品名称", "", "", "", "", ""),
    ("利率/费率", "", "", "", "", ""),
    ("额度范围", "", "", "", "", ""),
    ("期限", "", "", "", "", ""),
    ("准入条件", "", "", "", "", ""),
    ("审批时效", "", "", "", "", ""),
    ("通过率", "", "", "", "", ""),
    ("人均月结算", "", "", "", "", ""),
    ("接通率", "", "", "", "", ""),
    ("转化率", "", "", "", "", ""),
    ("坐席偏好", "", "", "", "", "更愿意推哪个？"),
    ("客户接受度", "", "", "", "", ""),
]

for i, row_data in enumerate(prod_rows, 16):
    for j, val in enumerate(row_data, 1):
        ws9.cell(row=i, column=j, value=val)
    style_row(ws9, i, 6)

# 减量决策评估
add_module_header(ws9, 29, "减量决策评估", 6)

decision_headers = ["评估维度", "数据/事实", "判断", "风险", "建议", "备注"]
for col, h in enumerate(decision_headers, 1):
    ws9.cell(row=30, column=col, value=h)
style_header(ws9, 30, 6)

decision_rows = [
    ("当前亏损程度", "", "", "", "", ""),
    ("盈亏平衡人数", "", "", "", "", ""),
    ("近6个月趋势", "", "趋势性下滑/短期波动", "", "", ""),
    ("减量后产能影响", "", "", "", "", ""),
    ("减下人员去向", "", "自然减员/转项目", "", "", ""),
    ("内部转项目风险", "", "", "", "", "年老师核心顾虑"),
    ("加回条件", "", "", "", "", ""),
    ("替代方案", "", "", "", "", ""),
]

for i, row_data in enumerate(decision_rows, 31):
    for j, val in enumerate(row_data, 1):
        ws9.cell(row=i, column=j, value=val)
    style_row(ws9, i, 6)

# ============================================================
# Sheet 10: 调研后整理
# ============================================================
ws10 = wb.create_sheet("调研后整理")
ws10.column_dimensions["A"].width = 12
ws10.column_dimensions["B"].width = 40
ws10.column_dimensions["C"].width = 30
ws10.column_dimensions["D"].width = 15
ws10.column_dimensions["E"].width = 20

headers = ["类别", "整理事项", "具体内容", "完成状态", "备注"]
for col, h in enumerate(headers, 1):
    ws10.cell(row=1, column=col, value=h)
style_header(ws10, 1, 5)

checklist = [
    ("验场", "硬件环境检查结果汇总", "合格/有条件合格/不合格", "", ""),
    ("验场", "团队规模确认", "在岗人数 vs 系统数据", "", ""),
    ("验场", "合规基础评估", "录音/培训/质检", "", ""),
    ("经营", "经营数据核实", "人均产能/结算/趋势", "", ""),
    ("经营", "成本结构梳理", "固定/变动/盈亏平衡点", "", ""),
    ("经营", "其他业务情况", "竞品产品/其他甲方", "", ""),
    ("业务", "转化漏斗数据", "拨打→接通→转化→放款", "", ""),
    ("业务", "问题诊断结论", "根因在哪一步", "", ""),
    ("业务", "流失率及原因", "新人/老人分别", "", ""),
    ("名单", "名单质量评估", "来源/有效率/复打率", "", ""),
    ("名单", "名单管理问题", "痛点/改进建议", "", ""),
    ("竞品", "产品要素对比", "利率/额度/准入/时效", "", ""),
    ("竞品", "运营数据对比", "结算/接通率/转化率", "", ""),
    ("竞品", "一线偏好", "更愿意推哪个/为什么", "", ""),
    ("竞品", "可借鉴的做法", "运营模式/管理方法", "", ""),
    ("合规", "合规风险评估", "违规率/投诉率/处理", "", ""),
    ("合规", "质检执行情况", "覆盖率/标准/闭环", "", ""),
    ("团队", "管理效率评估", "层级/幅度/信息传递", "", ""),
    ("团队", "薪酬竞争力", "在市场上什么水平", "", ""),
    ("减量", "减量诉求评估", "数据支撑+判断+建议", "", ""),
    ("汇报", "核心发现（3-5条）", "", "", "按重要性排序"),
    ("汇报", "行动建议", "", "", "具体可执行"),
    ("汇报", "需要老板支持的事项", "", "", ""),
    ("汇报", "风险点", "", "", ""),
]

for i, row_data in enumerate(checklist, 2):
    for j, val in enumerate(row_data, 1):
        ws10.cell(row=i, column=j, value=val)
    style_row(ws10, i, 5)

# ============================================================
# 设置打印区域和页面设置
# ============================================================
for ws in wb.worksheets:
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"

# 保存
output_path = "/Users/sundanian/Documents/projects/ai-agents/my-agent/workspace/2026-06-09-供应商调研/供应商深度调研工具包.xlsx"
wb.save(output_path)
print(f"已保存: {output_path}")
print(f"共 {len(wb.sheetnames)} 个sheet: {', '.join(wb.sheetnames)}")
